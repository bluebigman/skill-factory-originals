#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
股票预测与推荐分析工具（独立实现）

本脚本根据功能规格独立实现，仅用于学习与参考。
不构成任何投资建议，使用后果由使用者自行承担。
"""

import argparse
import csv
import json
import math
import os
import statistics
import sys
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple


# ============================================================
# 错误码定义
# ============================================================
ERROR_CODES = {
    "E001": "输入文件不存在或无法读取",
    "E002": "输入文件格式不支持（仅支持 CSV / JSON / Excel）",
    "E003": "数据解析失败：缺少必要字段",
    "E004": "数据解析失败：字段类型错误",
    "E005": "数据为空或记录数不足",
    "E006": "日期格式错误",
    "E007": "数值计算异常",
    "E008": "输出目录不可写",
    "E009": "参数错误或冲突",
    "E010": "内部逻辑错误",
}


# ============================================================
# 数据解析模块
# ============================================================
class DataParser:
    """数据解析器：将不同格式的输入转换为统一结构"""

    REQUIRED_FIELDS = ["date", "close"]

    @staticmethod
    def parse_file(filepath: str) -> List[Dict[str, Any]]:
        """解析文件，返回统一格式的数据列表"""
        if not os.path.exists(filepath):
            raise ValueError(ERROR_CODES["E001"])

        ext = os.path.splitext(filepath)[1].lower()
        if ext == ".csv":
            return DataParser._parse_csv(filepath)
        elif ext == ".json":
            return DataParser._parse_json(filepath)
        elif ext in (".xlsx", ".xls"):
            # Excel 解析需要第三方库
            try:
                import openpyxl  # pip install openpyxl
            except ImportError:
                raise ImportError("解析 Excel 需要安装 openpyxl: pip install openpyxl")
            return DataParser._parse_excel(filepath)
        else:
            raise ValueError(ERROR_CODES["E002"])

    @staticmethod
    def _parse_csv(filepath: str) -> List[Dict[str, Any]]:
        """解析 CSV 文件"""
        try:
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception:
            raise ValueError(ERROR_CODES["E003"])

        return DataParser._normalize_rows(rows)

    @staticmethod
    def _parse_json(filepath: str) -> List[Dict[str, Any]]:
        """解析 JSON 文件"""
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            raise ValueError(ERROR_CODES["E003"])

        # 支持两种结构：直接数组或 {"data": [...]}
        if isinstance(data, list):
            rows = data
        elif isinstance(data, dict) and "data" in data:
            rows = data["data"]
        else:
            raise ValueError(ERROR_CODES["E003"])

        return DataParser._normalize_rows(rows)

    @staticmethod
    def _parse_excel(filepath: str) -> List[Dict[str, Any]]:
        """解析 Excel 文件"""
        import openpyxl

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and any(v is not None for v in row):
                    rows.append(dict(zip(headers, row)))
        except Exception:
            raise ValueError(ERROR_CODES["E003"])

        return DataParser._normalize_rows(rows)

    @staticmethod
    def _normalize_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """将原始行数据标准化为内部结构"""
        if not rows:
            raise ValueError(ERROR_CODES["E005"])

        # 字段名兼容处理（支持中英文别名）
        field_map = {
            "date": ["date", "日期", "时间", "trade_date", "datetime"],
            "close": ["close", "收盘", "收盘价", "close_price"],
            "open": ["open", "开盘", "开盘价", "open_price"],
            "high": ["high", "最高", "最高价", "high_price"],
            "low": ["low", "最低", "最低价", "low_price"],
            "volume": ["volume", "成交量", "vol"],
        }

        normalized = []
        for raw_row in rows:
            item = {}
            for std_key, aliases in field_map.items():
                for alias in aliases:
                    if alias in raw_row:
                        item[std_key] = raw_row[alias]
                        break

            # 检查必要字段
            if "date" not in item or "close" not in item:
                raise ValueError(ERROR_CODES["E003"])

            # 类型转换
            try:
                item["date"] = str(item["date"]).strip()
                item["close"] = float(item["close"])
                if "open" in item:
                    item["open"] = float(item["open"])
                if "high" in item:
                    item["high"] = float(item["high"])
                if "low" in item:
                    item["low"] = float(item["low"])
                if "volume" in item:
                    item["volume"] = float(item["volume"])
            except (ValueError, TypeError):
                raise ValueError(ERROR_CODES["E004"])

            normalized.append(item)

        # 按日期排序
        try:
            normalized.sort(key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d"))
        except ValueError:
            try:
                normalized.sort(key=lambda x: datetime.strptime(x["date"], "%Y/%m/%d"))
            except ValueError:
                raise ValueError(ERROR_CODES["E006"])

        return normalized


# ============================================================
# 统计分析模块
# ============================================================
class StockAnalyzer:
    """股票数据分析器：计算统计指标、趋势判断、预测"""

    @staticmethod
    def calculate_returns(prices: List[float]) -> List[float]:
        """计算日收益率序列"""
        if len(prices) < 2:
            return []
        returns = []
        for i in range(1, len(prices)):
            prev = prices[i - 1]
            if prev == 0:
                returns.append(0.0)
            else:
                returns.append((prices[i] - prev) / prev)
        return returns

    @staticmethod
    def moving_average(prices: List[float], window: int) -> List[Optional[float]]:
        """计算移动平均线"""
        if window <= 0 or len(prices) < window:
            return [None] * len(prices)

        result = [None] * len(prices)
        for i in range(window - 1, len(prices)):
            window_data = prices[i - window + 1 : i + 1]
            result[i] = sum(window_data) / window
        return result

    @staticmethod
    def volatility(returns: List[float]) -> float:
        """计算波动率（标准差）"""
        if len(returns) < 2:
            return 0.0
        return statistics.stdev(returns)

    @staticmethod
    def trend_strength(prices: List[float]) -> Dict[str, Any]:
        """判断趋势方向和强度"""
        if len(prices) < 5:
            return {"direction": "unknown", "strength": 0.0, "description": "数据不足"}

        # 线性回归斜率
        n = len(prices)
        x_mean = (n - 1) / 2
        y_mean = sum(prices) / n

        numerator = sum((i - x_mean) * (prices[i] - y_mean) for i in range(n))
        denominator = sum((i - x_mean) ** 2 for i in range(n))

        if denominator == 0:
            slope = 0.0
        else:
            slope = numerator / denominator

        # 归一化斜率（相对价格水平）
        if y_mean != 0:
            norm_slope = slope / y_mean
        else:
            norm_slope = 0.0

        # 判断方向
        if norm_slope > 0.01:
            direction = "up"
        elif norm_slope < -0.01:
            direction = "down"
        else:
            direction = "sideways"

        # 计算 R² 作为趋势强度
        if denominator == 0 or n < 2:
            r_squared = 0.0
        else:
            y_pred = [slope * i + (y_mean - slope * x_mean) for i in range(n)]
            ss_res = sum((prices[i] - y_pred[i]) ** 2 for i in range(n))
            ss_tot = sum((prices[i] - y_mean) ** 2 for i in range(n))
            r_squared = 1.0 - (ss_res / ss_tot) if ss_tot > 0 else 0.0

        # 强度分级
        if r_squared > 0.7:
            strength = "strong"
        elif r_squared > 0.4:
            strength = "moderate"
        elif r_squared > 0.1:
            strength = "weak"
        else:
            strength = "noise"

        return {
            "direction": direction,
            "strength": strength,
            "r_squared": r_squared,
            "norm_slope": norm_slope,
        }

    @staticmethod
    def predict_next_price(prices: List[float]) -> Dict[str, Any]:
        """基于简单线性回归预测下一交易日的价格区间"""
        if len(prices) < 3:
            return {"predicted": None, "low": None, "high": None, "confidence": "low"}

        n = len(prices)
        x_mean = (n - 1) / 2
        y_mean = sum(prices) / n

        numerator = sum((i - x_mean) * (prices[i] - y_mean) for i in range(n))
        denominator = sum((i - x_mean) ** 2 for i in range(n))

        if denominator == 0:
            slope = 0.0
        else:
            slope = numerator / denominator

        intercept = y_mean - slope * x_mean
        predicted = slope * n + intercept

        # 计算残差标准差
        residuals = [prices[i] - (slope * i + intercept) for i in range(n)]
        if len(residuals) > 1:
            residual_std = statistics.stdev(residuals)
        else:
            residual_std = 0.0

        # 置信区间（约 1 倍标准差）
        low = predicted - residual_std
        high = predicted + residual_std

        # 置信度基于数据量和 R²
        if n >= 30:
            confidence = "high"
        elif n >= 10:
            confidence = "medium"
        else:
            confidence = "low"

        return {
            "predicted": predicted,
            "low": low,
            "high": high,
            "confidence": confidence,
        }

    @staticmethod
    def analyze(data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """执行完整分析"""
        if len(data) < 5:
            raise ValueError(ERROR_CODES["E005"])

        prices = [item["close"] for item in data]
        dates = [item["date"] for item in data]

        # 基础统计
        returns = StockAnalyzer.calculate_returns(prices)
        avg_price = sum(prices) / len(prices)
        max_price = max(prices)
        min_price = min(prices)
        total_change = (prices[-1] - prices[0]) / prices[0] if prices[0] != 0 else 0

        # 波动率（年化，假设 250 个交易日）
        daily_vol = StockAnalyzer.volatility(returns)
        annual_vol = daily_vol * math.sqrt(250) if daily_vol else 0

        # 移动平均线
        ma5 = StockAnalyzer.moving_average(prices, 5)
        ma20 = StockAnalyzer.moving_average(prices, 20)

        # 趋势分析
        trend = StockAnalyzer.trend_strength(prices)

        # 预测
        prediction = StockAnalyzer.predict_next_price(prices)

        # 风险评级
        if annual_vol < 0.15:
            risk_level = "低"
        elif annual_vol < 0.35:
            risk_level = "中"
        else:
            risk_level = "高"

        return {
            "data_overview": {
                "stock_code": data[0].get("code", "未知"),
                "start_date": dates[0],
                "end_date": dates[-1],
                "data_points": len(data),
                "avg_price": avg_price,
                "max_price": max_price,
                "min_price": min_price,
                "total_change_percent": total_change * 100,
            },
            "technical_indicators": {
                "ma5": ma5[-1] if ma5 and ma5[-1] else None,
                "ma20": ma20[-1] if ma20 and ma20[-1] else None,
                "daily_volatility": daily_vol,
                "annual_volatility": annual_vol,
                "risk_level": risk_level,
            },
            "trend": {
                "direction": trend["direction"],
                "strength": trend["strength"],
                "confidence": "high" if trend["r_squared"] > 0.6 else "medium" if trend["r_squared"] > 0.3 else "low",
            },
            "prediction": {
                "next_price": prediction["predicted"],
                "range_low": prediction["low"],
                "range_high": prediction["high"],
                "confidence": prediction["confidence"],
            },
            "recommendation": StockAnalyzer._generate_recommendation(trend, prediction),
            "disclaimer": "本分析仅供学习参考，不构成投资建议。投资有风险，入市需谨慎。",
        }

    @staticmethod
    def _generate_recommendation(trend: Dict[str, Any], prediction: Dict[str, Any]) -> Dict[str, Any]:
        """生成建议（仅参考）"""
        direction_map = {
            "up": "上涨",
            "down": "下跌",
            "sideways": "横盘",
            "unknown": "未知",
        }
        strength_map = {
            "strong": "强劲",
            "moderate": "温和",
            "weak": "弱势",
            "noise": "噪音",
            "unknown": "未知",
        }

        # 综合判断
        if trend["direction"] == "up" and trend["strength"] in ("strong", "moderate"):
            action = "关注"
            reason = f"趋势{strength_map[trend['strength']]}向上"
        elif trend["direction"] == "down" and trend["strength"] in ("strong", "moderate"):
            action = "谨慎"
            reason = f"趋势{strength_map[trend['strength']]}向下"
        elif trend["direction"] == "sideways":
            action = "观望"
            reason = "趋势不明朗"
        else:
            action = "中性"
            reason = "趋势信号较弱"

        return {
            "action": action,
            "reason": reason,
            "suggestion": "建议结合更多信息综合判断，切勿盲目跟从。",
        }


# ============================================================
# 报告输出模块
# ============================================================
class ReportGenerator:
    """报告生成器"""

    @staticmethod
    def format_report(analysis: Dict[str, Any]) -> str:
        """生成文本报告"""
        overview = analysis["data_overview"]
        indicators = analysis["technical_indicators"]
        trend = analysis["trend"]
        prediction = analysis["prediction"]
        rec = analysis["recommendation"]

        lines = []
        lines.append("=" * 60)
        lines.append("股票分析报告")
        lines.append("=" * 60)
        lines.append(f"股票代码: {overview['stock_code']}")
        lines.append(f"分析区间: {overview['start_date']} 至 {overview['end_date']}")
        lines.append(f"数据点数: {overview['data_points']}")
        lines.append("")
        lines.append("【数据概览】")
        lines.append(f"  平均价格: {overview['avg_price']:.2f}")
        lines.append(f"  最高价格: {overview['max_price']:.2f}")
        lines.append(f"  最低价格: {overview['min_price']:.2f}")
        lines.append(f"  区间涨跌: {overview['total_change_percent']:+.2f}%")
        lines.append("")
        lines.append("【技术指标】")
        if indicators["ma5"] is not None:
            lines.append(f"  5日均线: {indicators['ma5']:.2f}")
        if indicators["ma20"] is not None:
            lines.append(f"  20日均线: {indicators['ma20']:.2f}")
        lines.append(f"  年化波动率: {indicators['annual_volatility']:.2%}")
        lines.append(f"  风险等级: {indicators['risk_level']}")
        lines.append("")
        lines.append("【趋势判断】")
        direction_cn = {"up": "上涨", "down": "下跌", "sideways": "横盘", "unknown": "未知"}
        strength_cn = {"strong": "强劲", "moderate": "温和", "weak": "弱势", "noise": "噪音", "unknown": "未知"}
        lines.append(f"  趋势方向: {direction_cn.get(trend['direction'], trend['direction'])}")
        lines.append(f"  趋势强度: {strength_cn.get(trend['strength'], trend['strength'])}")
        lines.append(f"  置信水平: {trend['confidence']}")
        lines.append("")
        lines.append("【预测参考】")
        if prediction["next_price"] is not None:
            lines.append(f"  预测下一价格: {prediction['next_price']:.2f}")
            lines.append(f"  预测区间: [{prediction['range_low']:.2f}, {prediction['range_high']:.2f}]")
        lines.append(f"  预测置信度: {prediction['confidence']}")
        lines.append("")
        lines.append("【参考建议】")
        lines.append(f"  建议: {rec['action']}")
        lines.append(f"  原因: {rec['reason']}")
        lines.append(f"  说明: {rec['suggestion']}")
        lines.append("")
        lines.append("-" * 60)
        lines.append(f"免责声明: {analysis['disclaimer']}")
        lines.append("=" * 60)

        return "\n".join(lines)


# ============================================================
# 自检模块
# ============================================================
def run_selftest() -> bool:
    """内置自检：使用硬编码样例数据验证核心逻辑"""
    print("=" * 60)
    print("自检模式：使用内置样例数据验证核心逻辑")
    print("=" * 60)

    # 生成模拟数据（确定性伪随机，保证可重复）
    # 使用线性上升 + 正弦波动 + 噪声
    import random

    random.seed(42)  # 固定种子确保可重复

    sample_data = []
    base_price = 100.0
    start_date = datetime(2024, 1, 1)

    for i in range(60):  # 60 个交易日
        date = (start_date + timedelta(days=i)).strftime("%Y-%m-%d")
        # 线性趋势 + 周期性波动 + 噪声
        trend = i * 0.3
        cycle = math.sin(i / 10.0) * 3.0
        noise = random.uniform(-1.5, 1.5)
        close = base_price + trend + cycle + noise

        sample_data.append(
            {
                "date": date,
                "close": close,
                "open": close - random.uniform(0, 1),
                "high": close + random.uniform(0, 1),
                "low": close - random.uniform(0, 1),
                "volume": random.uniform(10000, 50000),
            }
        )

    # 执行分析
    print("\n[1/4] 测试数据解析模块...")
    try:
        parsed = DataParser._normalize_rows(sample_data)
        assert len(parsed) == 60, f"数据解析失败: 期望60条, 实际{len(parsed)}条"
        assert all("date" in item and "close" in item for item in parsed), "解析结果缺少必要字段"
        print("  ✓ 数据解析模块正常")
    except Exception as e:
        print(f"  ✗ 数据解析模块异常: {e}")
        return False

    # 测试统计分析
    print("\n[2/4] 测试统计分析模块...")
    try:
        prices = [item["close"] for item in parsed]

        # 收益率计算
        returns = StockAnalyzer.calculate_returns(prices)
        assert len(returns) == 59, f"收益率计算错误: 期望59个, 实际{len(returns)}个"
        assert all(isinstance(r, float) for r in returns), "收益率类型错误"

        # 移动平均
        ma5 = StockAnalyzer.moving_average(prices, 5)
        assert ma5[0] is None and ma5[4] is not None, "移动平均计算错误"

        # 波动率
        vol = StockAnalyzer.volatility(returns)
        assert vol > 0, "波动率应为正数"

        # 趋势分析
        trend = StockAnalyzer.trend_strength(prices)
        assert trend["direction"] in ("up", "down", "sideways"), "趋势方向非法"
        assert trend["strength"] in ("strong", "moderate", "weak", "noise"), "趋势强度非法"

        # 预测
        prediction = StockAnalyzer.predict_next_price(prices)
        assert prediction["predicted"] is not None, "预测结果为空"
        assert prediction["low"] < prediction["predicted"] < prediction["high"], "预测区间不合理"

        print("  ✓ 统计分析模块正常")
        print(f"    趋势方向: {trend['direction']}, 强度: {trend['strength']}")
        print(f"    预测价格: {prediction['predicted']:.2f}, 区间: [{prediction['low']:.2f}, {prediction['high']:.2f}]")
    except Exception as e:
        print(f"  ✗ 统计分析模块异常: {e}")
        return False

    # 测试完整分析流程
    print("\n[3/4] 测试完整分析流程...")
    try:
        analysis = StockAnalyzer.analyze(parsed)
        assert "data_overview" in analysis, "缺少数据概览"
        assert "technical_indicators" in analysis, "缺少技术指标"
        assert "trend" in analysis, "缺少趋势分析"
        assert "prediction" in analysis, "缺少预测"
        assert "recommendation" in analysis, "缺少建议"

        # 验证关键字段
        overview = analysis["data_overview"]
        assert overview["data_points"] == 60, "数据点数错误"
        assert overview["avg_price"] > 0, "平均价格应为正"

        print("  ✓ 完整分析流程正常")
        print(f"    平均价格: {overview['avg_price']:.2f}")
        print(f"    区间涨跌: {overview['total_change_percent']:+.2f}%")
    except Exception as e:
        print(f"  ✗ 完整分析流程异常: {e}")
        return False

    # 测试报告生成
    print("\n[4/4] 测试报告生成...")
    try:
        analysis = StockAnalyzer.analyze(parsed)
        report = ReportGenerator.format_report(analysis)
        assert len(report) > 100, "报告内容过短"
        assert "股票分析报告" in report, "报告缺少标题"
        assert "免责声明" in report, "报告缺少免责声明"

        print("  ✓ 报告生成模块正常")
        print(f"    报告长度: {len(report)} 字符")
    except Exception as e:
        print(f"  ✗ 报告生成模块异常: {e}")
        return False

    print("\n" + "=" * 60)
    print("自检全部通过 ✓")
    print("=" * 60)
    return True


# ============================================================
# 主入口
# ============================================================
def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="股票预测与推荐分析工具（仅供学习参考，不构成投资建议）",
        epilog="示例: python main.py data.csv 或 python main.py --selftest",
    )
    parser.add_argument("file", nargs="?", help="输入数据文件路径（CSV/JSON/Excel）")
    parser.add_argument("--selftest", action="store_true", help="运行内置自检")
    parser.add_argument("--output", "-o", help="输出报告到文件")
    parser.add_argument("--code", help="股票代码")

    args = parser.parse_args()

    # 自检模式
    if args.selftest:
        success = run_selftest()
        sys.exit(0 if success else 1)

    # 正常模式
    if not args.file:
        parser.error("请提供输入文件路径（或使用 --selftest 运行自检）")

    try:
        # 解析数据
        print(f"正在读取文件: {args.file}")
        data = DataParser.parse_file(args.file)

        # 设置股票代码
        if args.code:
            for item in data:
                item["code"] = args.code

        # 执行分析
        print(f"数据加载成功，共 {len(data)} 条记录")
        print("正在分析...")
        analysis = StockAnalyzer.analyze(data)

        # 生成报告
        report = ReportGenerator.format_report(analysis)

        # 输出
        if args.output:
            output_path = args.output
            output_dir = os.path.dirname(os.path.abspath(output_path))
            if not os.path.exists(output_dir):
                raise ValueError(ERROR_CODES["E008"])
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(report)
            print(f"报告已保存至: {output_path}")
        else:
            print(report)

    except ValueError as e:
        print(f"错误: {e}", file=sys.stderr)
        # 根据错误信息映射错误码
        for code, msg in ERROR_CODES.items():
            if msg == str(e):
                print(f"错误码: {code}", file=sys.stderr)
                break
        sys.exit(1)
    except ImportError as e:
        print(f"错误: {e}", file=sys.stderr)
        print("请安装所需依赖后重试。", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"错误: 发生未知异常 - {e}", file=sys.stderr)
        print(f"错误码: E010", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
