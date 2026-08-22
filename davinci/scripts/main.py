#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
davinci - 数据可视化 智能解析 图表生成

功能：将用户数据文件/URL解析为结构化结果，支持批量与自定义格式输出。
仅依据功能规格独立实现（clean-room）。

作者：Ling Xiao
版本：1.0.8
许可证：MIT
"""

import argparse
import csv
import io
import json
import os
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse
import time
import urllib.request
import urllib.error
import shutil
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
import http.server
import threading
import socketserver

# G1 生产级重试退避
_max_retry = 3  # 最大重试次数
_retryable_status_codes = {500, 502, 503, 504}  # 仅5xx可重试，4xx永久失败

def _retry_request(fn, *args, **kwargs):
    """带重试退避的请求封装（G1 生产门禁）。
    
    仅对可重试错误（5xx、连接错误、超时）进行重试，
    4xx 永久失败直接抛出。
    """
    timeout = kwargs.pop('timeout', 10)  # 默认超时10秒
    for attempt in range(_max_retry):
        try:
            return fn(*args, timeout=timeout, **kwargs)
        except urllib.error.HTTPError as e:
            # 4xx 永久失败，直接抛出
            if e.code < 500:
                raise
            # 5xx 进行退避重试
            if attempt < _max_retry - 1:
                time.sleep(2 ** attempt)  # 指数退避
            else:
                raise
        except (urllib.error.URLError, TimeoutError, ConnectionError, OSError) as e:
            # 连接错误、超时等网络问题，进行退避重试
            if attempt < _max_retry - 1:
                time.sleep(2 ** attempt)  # 指数退避
            else:
                raise
        # 不再捕获所有异常，避免捕获KeyboardInterrupt和SystemExit

# 错误码定义
ERROR_CODES = {
    "E001": "参数错误：缺少必要参数或参数格式不正确",
    "E002": "文件不存在或无法访问",
    "E003": "文件格式不支持（仅支持 CSV/JSON/Excel/公开URL）",
    "E004": "文件大小超过50MB限制",
    "E005": "数据解析失败：无法识别数据结构",
    "E006": "URL 无法访问或需要登录",
    "E007": "批量处理失败：部分文件处理出错",
    "E008": "自定义格式模板错误",
    "E009": "输出写入失败",
    "E010": "内部处理异常",
}

# 能力常量
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
SUPPORTED_EXTENSIONS = {".csv", ".json", ".xlsx", ".xls", ".zip"}


def error_exit(code: str, message: Optional[str] = None) -> None:
    """输出错误码并退出程序"""
    msg = ERROR_CODES.get(code, "未知错误")
    if message:
        msg = f"{msg}：{message}"
    print(f"[错误 {code}] {msg}", file=sys.stderr)
    sys.exit(1)


def load_csv_data(file_path: str) -> List[Dict[str, Any]]:
    """从 CSV 文件加载数据（完整实现）"""
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            # 去除字段名两端的空白字符
            rows = []
            for row in reader:
                clean_row = {k.strip() if k else k: v for k, v in row.items()}
                rows.append(clean_row)
            if not rows:
                error_exit("E005", "CSV 文件为空或无有效数据")
            return rows
    except FileNotFoundError:
        error_exit("E002", f"文件不存在：{file_path}")
    except Exception as e:
        error_exit("E005", f"CSV 解析失败：{str(e)}")


def load_json_data(file_path: str) -> List[Dict[str, Any]]:
    """从 JSON 文件加载数据"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        # 支持多种 JSON 结构：列表、字典（取第一个列表值）
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        elif isinstance(data, dict):
            for value in data.values():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    return [item for item in value if isinstance(item, dict)]
            # 单条字典数据
            return [data]
        else:
            error_exit("E005", "JSON 顶层必须是对象或数组")
    except FileNotFoundError:
        error_exit("E002", f"文件不存在：{file_path}")
    except json.JSONDecodeError as e:
        error_exit("E005", f"JSON 解析失败：{str(e)}")


def load_excel_data(file_path: str) -> List[Dict[str, Any]]:
    """从 Excel 文件加载数据（使用 openpyxl 或降级方案）"""
    # 尝试使用 openpyxl（如果已安装）
    try:
        from openpyxl import load_workbook  # pip install openpyxl
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            error_exit("E005", "Excel 文件为空")
        headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if row and any(cell is not None for cell in row):
                result.append({headers[i]: row[i] for i in range(len(headers))})
        wb.close()
        return result
    except ImportError:
        # 降级：尝试用 zipfile 解析 xlsx（仅读取 sharedStrings 和 sheet1）
        try:
            with zipfile.ZipFile(file_path) as zf:
                # 读取共享字符串
                shared_strings = []
                if "xl/sharedStrings.xml" in zf.namelist():
                    tree = ET.parse(zf.open("xl/sharedStrings.xml"))
                    root = tree.getroot()
                    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                    for si in root.findall("m:si", ns):
                        text_parts = []
                        for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                            if t.text:
                                text_parts.append(t.text)
                        shared_strings.append("".join(text_parts))

                # 读取第一个 sheet
                sheet_file = None
                for name in zf.namelist():
                    if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                        sheet_file = name
                        break
                if not sheet_file:
                    error_exit("E005", "Excel 中未找到工作表")

                tree = ET.parse(zf.open(sheet_file))
                root = tree.getroot()
                ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

                # 解析行数据
                all_rows = []
                for row in root.findall(".//m:sheetData/m:row", ns):
                    row_data = []
                    for cell in row.findall("m:c", ns):
                        cell_type = cell.get("t", "")
                        value_elem = cell.find("m:v", ns)
                        if value_elem is None or value_elem.text is None:
                            row_data.append(None)
                            continue
                        if cell_type == "s":
                            idx = int(value_elem.text)
                            row_data.append(shared_strings[idx] if idx < len(shared_strings) else "")
                        else:
                            row_data.append(value_elem.text)
                    all_rows.append(row_data)

                if not all_rows:
                    error_exit("E005", "Excel 数据为空")
                headers = [str(h) if h else f"column_{i}" for i, h in enumerate(all_rows[0])]
                result = []
                for row_data in all_rows[1:]:
                    if any(c is not None for c in row_data):
                        result.append({headers[i]: row_data[i] for i in range(len(headers))})
                return result
        except Exception as e:
            error_exit("E005", f"Excel 解析失败：{str(e)}")


@lru_cache(maxsize=128)
def parse_url_data_cached(url: str) -> Tuple[List[Dict[str, Any]], str]:
    """带缓存的 URL 数据解析"""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        # 使用 _retry_request 封装，设置超时
        with _retry_request(urllib.request.urlopen, req) as resp:
            # 检查大小
            content_length = resp.headers.get("Content-Length")
            if content_length and int(content_length) > MAX_FILE_SIZE:
                error_exit("E004", f"URL 内容超过 {MAX_FILE_SIZE // (1024*1024)}MB")
            data = resp.read(MAX_FILE_SIZE + 1)
            if len(data) > MAX_FILE_SIZE:
                error_exit("E004", f"URL 内容超过 {MAX_FILE_SIZE // (1024*1024)}MB")
    except urllib.error.HTTPError as e:
        if e.code >= 400 and e.code < 500:
            error_exit("E006", f"URL 访问失败（HTTP {e.code}）：{url}")
        else:
            error_exit("E006", f"URL 访问失败（HTTP {e.code}）：{url}")
    except Exception as e:
        error_exit("E006", f"URL 访问失败：{str(e)}")

    # 根据 URL 后缀或内容猜测格式
    path = urlparse(url).path.lower()
    text = data.decode("utf-8", errors="ignore")

    try:
        if path.endswith(".json"):
            parsed = json.loads(text)
            return (parsed if isinstance(parsed, list) else [parsed]), "json"
        elif path.endswith(".csv"):
            reader = csv.DictReader(io.StringIO(text))
            return [row for row in reader], "csv"
        else:
            # 尝试 JSON
            try:
                parsed = json.loads(text)
                return (parsed if isinstance(parsed, list) else [parsed]), "json"
            except json.JSONDecodeError:
                # 尝试 CSV
                reader = csv.DictReader(io.StringIO(text))
                rows = [row for row in reader]
                if rows:
                    return rows, "csv"
                error_exit("E005", "URL 内容无法解析为 JSON 或 CSV")
    except Exception as e:
        error_exit("E005", f"URL 数据解析失败：{str(e)}")


def parse_url_data(url: str) -> List[Dict[str, Any]]:
    """从公开 URL 加载数据（带缓存）"""
    rows, _ = parse_url_data_cached(url)
    return rows


def load_zip_data(zip_path: str) -> List[Dict[str, Any]]:
    """从 zip 文件加载数据（解压后递归解析）"""
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            # 查找支持的文件
            supported_files = []
            for name in zf.namelist():
                ext = os.path.splitext(name)[1].lower()
                if ext in {".csv", ".json", ".xlsx", ".xls"}:
                    supported_files.append(name)
            
            if not supported_files:
                error_exit("E003", f"ZIP 文件中没有支持的格式文件（CSV/JSON/Excel）")
            
            # 解析所有支持的文件并合并结果
            all_rows = []
            for file_name in supported_files:
                with zf.open(file_name) as f:
                    # 创建临时文件
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_name)[1]) as tmp:
                        shutil.copyfileobj(f, tmp)
                        tmp_path = tmp.name
                    
                    try:
                        # 递归调用 load_data 解析临时文件
                        rows = load_data(tmp_path)
                        all_rows.extend(rows)
                    finally:
                        # 清理临时文件
                        os.unlink(tmp_path)
            
            if not all_rows:
                error_exit("E005", "ZIP 文件中没有有效数据")
            return all_rows
    except zipfile.BadZipFile:
        error_exit("E005", f"ZIP 文件损坏：{zip_path}")
    except Exception as e:
        error_exit("E005", f"ZIP 解析失败：{str(e)}")


def load_data(input_path: str) -> List[Dict[str, Any]]:
    """根据输入路径加载数据（支持文件或 URL）"""
    # 检查是否为 URL
    if input_path.startswith(("http://", "https://")):
        return parse_url_data(input_path)

    # 检查文件是否存在
    if not os.path.exists(input_path):
        error_exit("E002", f"文件不存在：{input_path}")

    # 检查文件大小
    file_size = os.path.getsize(input_path)
    if file_size > MAX_FILE_SIZE:
        error_exit("E004", f"文件大小 {file_size / (1024*1024):.1f}MB 超过 50MB 限制")

    # 根据扩展名加载
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".csv":
        return load_csv_data(input_path)
    elif ext == ".json":
        return load_json_data(input_path)
    elif ext in (".xlsx", ".xls"):
        return load_excel_data(input_path)
    elif ext == ".zip":
        return load_zip_data(input_path)
    else:
        error_exit("E003", f"不支持的格式：{ext}")


def analyze_data(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """分析数据结构，返回字段信息和置信度"""
    if not rows:
        error_exit("E005", "无数据可分析")

    # 收集所有字段
    all_fields = set()
    for row in rows:
        all_fields.update(row.keys())

    field_info = {}
    for field in all_fields:
        values = [row.get(field) for row in rows if row.get(field) is not None]
        non_null_count = len(values)
        total_count = len(rows)
        null_count = total_count - non_null_count

        # 判断数据类型
        field_type = "string"
        numeric_count = 0
        for v in values:
            if isinstance(v, (int, float)):
                numeric_count += 1
            elif isinstance(v, str):
                try:
                    float(v)
                    numeric_count += 1
                except (ValueError, TypeError):
                    pass

        if numeric_count > total_count * 0.8:
            field_type = "numeric"
        elif values and all(isinstance(v, str) for v in values):
            # 检查是否为时间字段（简单启发式）
            date_count = 0
            for v in values:
                try:
                    datetime.fromisoformat(str(v).replace("Z", "+00:00"))
                    date_count += 1
                except ValueError:
                    pass
            if date_count > total_count * 0.8:
                field_type = "datetime"
            else:
                field_type = "string"

        # 置信度计算：非空比例越高，置信度越高
        if null_count == 0:
            confidence = "高"
        elif null_count <= total_count * 0.2:
            confidence = "中"
        else:
            confidence = "低"

        field_info[field] = {
            "type": field_type,
            "non_null": non_null_count,
            "null": null_count,
            "confidence": confidence,
            "sample_values": values[:3],
        }

    return {
        "total_rows": len(rows),
        "field_count": len(all_fields),
        "fields": field_info,
    }


def format_template_output(rows: List[Dict[str, Any]], template: str) -> str:
    """自定义格式模板解析函数，支持 {field} 占位符"""
    try:
        output_lines = []
        for row in rows:
            line = template
            for key, value in row.items():
                placeholder = "{" + key + "}"
                if placeholder in line:
                    line = line.replace(placeholder, str(value))
            output_lines.append(line)
        return "\n".join(output_lines)
    except Exception as e:
        error_exit("E008", f"模板解析失败：{str(e)}")


def format_output(rows: List[Dict[str, Any]], analysis: Dict[str, Any],
                  output_format: str = "json", custom_fields: Optional[List[str]] = None,
                  template: Optional[str] = None) -> str:
    """按指定格式输出结果"""
    # 应用自定义字段过滤
    if custom_fields:
        filtered_rows = []
        for row in rows:
            new_row = {}
            for field in custom_fields:
                if field in row:
                    new_row[field] = row[field]
            filtered_rows.append(new_row)
        rows = filtered_rows

    if output_format == "json":
        return json.dumps({"analysis": analysis, "data": rows}, ensure_ascii=False, indent=2)
    elif output_format == "csv":
        if not rows:
            return ""
        output = io.StringIO()
        fieldnames = list
