#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
django-dashing 技能实现脚本

本脚本根据功能规格独立实现（clean-room 风格），提供：
- 数据源接入解析（CSV/JSON/Excel/数据库连接串）
- 图表配置推荐生成（折线/柱状/饼图）
- Django dashboard 视图代码生成
- 模块化网格布局配置生成
- 数据刷新策略配置生成

支持命令行参数 --selftest 进行离线自检（不依赖外部文件/网络/工作目录）。
"""

import argparse
import json
import os
import sqlite3
import sys
import tempfile
import time
import urllib.request
import urllib.error
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

# 尝试导入 pandas/openpyxl 用于 Excel 解析
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 错误码定义
ERROR_CODES = {
    "E001": "参数错误：缺少必要的输入参数",
    "E002": "文件读取失败：文件不存在或无法读取",
    "E003": "数据解析失败：格式不支持或数据损坏",
    "E004": "图表类型不匹配：无法为给定数据推荐图表",
    "E005": "Django 项目路径无效：路径不存在或不是目录",
    "E006": "布局配置生成失败：模块参数无效",
    "E007": "刷新策略配置失败：间隔参数无效",
    "E008": "JSON 序列化失败：数据无法序列化",
    "E009": "自检失败：核心逻辑验证未通过",
    "E010": "未知错误：发生未预期的异常",
}


class DashboardError(Exception):
    """仪表盘自定义异常，携带错误码。"""

    def __init__(self, code: str, message: str = ""):
        self.code = code
        self.message = message or ERROR_CODES.get(code, "未知错误")
        super().__init__(f"[{code}] {self.message}")


# ---------------------------------------------------------------------------
# 数据源接入模块
# ---------------------------------------------------------------------------

def parse_data_source(source: str, data: Optional[str] = None) -> Dict[str, Any]:
    """
    解析数据源，返回标准化数据字典。

    支持：
    - CSV 文本（逗号分隔）
    - JSON 文本
    - Excel 文件（需 pandas/openpyxl）
    - 数据库连接串（sqlite3/psycopg2）

    参数:
        source: 数据来源标识（文件路径、连接串或数据格式名）
        data: 可选的数据内容（当 source 为格式名时使用）

    返回:
        标准化数据字典，格式: {"type": ..., "fields": [...], "rows": [...]}
    """
    if not source:
        raise DashboardError("E001", "数据源未指定")

    # 处理文件路径
    if os.path.isfile(source):
        try:
            ext = os.path.splitext(source)[1].lower()
            if ext in (".xlsx", ".xls"):
                return _parse_excel(source)
            with open(source, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            raise DashboardError("E002", f"读取文件失败: {e}")

        if ext in (".csv", ".txt"):
            return _parse_csv(content)
        elif ext == ".json":
            return _parse_json(content)
        else:
            raise DashboardError("E003", f"不支持的文件格式: {ext}")

    # 处理数据库连接串
    if "://" in source and not source.startswith("http"):
        return _parse_database(source)

    # 处理直接数据内容
    if data is not None:
        if source.lower() in ("csv", "text", "txt"):
            return _parse_csv(data)
        elif source.lower() == "json":
            return _parse_json(data)
        else:
            raise DashboardError("E003", f"不支持的数据格式: {source}")

    raise DashboardError("E001", "无法识别数据源类型")


def _parse_csv(content: str) -> Dict[str, Any]:
    """解析 CSV 文本为标准化数据字典。"""
    lines = [line.strip() for line in content.strip().splitlines() if line.strip()]
    if len(lines) < 2:
        raise DashboardError("E003", "CSV 数据至少需要表头和一行数据")

    fields = [f.strip() for f in lines[0].split(",")]
    rows = []
    for line in lines[1:]:
        values = [v.strip() for v in line.split(",")]
        if len(values) != len(fields):
            raise DashboardError("E003", "CSV 行字段数量与表头不一致")
        rows.append(dict(zip(fields, values)))

    return {"type": "csv", "fields": fields, "rows": rows}


def _parse_json(content: str) -> Dict[str, Any]:
    """解析 JSON 文本为标准化数据字典。"""
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as e:
        raise DashboardError("E003", f"JSON 解析失败: {e}")

    if isinstance(parsed, list):
        if not parsed:
            return {"type": "json", "fields": [], "rows": []}
        fields = list(parsed[0].keys()) if isinstance(parsed[0], dict) else []
        rows = parsed if all(isinstance(r, dict) for r in parsed) else []
        return {"type": "json", "fields": fields, "rows": rows}
    elif isinstance(parsed, dict):
        # 支持 {data: [...]} 或 {field: [...]} 形式
        if "data" in parsed and isinstance(parsed["data"], list):
            return _parse_json(json.dumps(parsed["data"]))
        return {"type": "json", "fields": list(parsed.keys()), "rows": [parsed]}
    else:
        raise DashboardError("E003", "JSON 数据必须是对象或数组")


def _parse_excel(filepath: str) -> Dict[str, Any]:
    """解析 Excel 文件为标准化数据字典。"""
    if not HAS_PANDAS and not HAS_OPENPYXL:
        raise DashboardError("E003", "Excel 解析需要 pandas 或 openpyxl 库，请先安装")

    try:
        if HAS_PANDAS:
            df = pd.read_excel(filepath)
            fields = list(df.columns)
            rows = df.to_dict('records')
            # 转换所有值为字符串以保持一致性
            rows = [{k: str(v) for k, v in row.items()} for row in rows]
        else:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                rows.append({headers[i]: str(v) if v is not None else "" for i, v in enumerate(row)})
            fields = headers

        if not fields:
            raise DashboardError("E003", "Excel 文件没有表头")
        return {"type": "excel", "fields": fields, "rows": rows}
    except DashboardError:
        raise
    except Exception as e:
        raise DashboardError("E003", f"Excel 解析失败: {e}")


def _parse_database(connection_string: str) -> Dict[str, Any]:
    """
    解析数据库连接串并执行查询。

    支持 SQLite (sqlite:///path) 和 PostgreSQL (postgresql://user:pass@host/db)
    """
    scheme = connection_string.split("://")[0].lower()
    
    if scheme == "sqlite":
        # SQLite 连接串格式: sqlite:///path/to/db.sqlite3
        db_path = connection_string.replace("sqlite:///", "")
        if not os.path.isfile(db_path):
            raise DashboardError("E003", f"SQLite 数据库文件不存在: {db_path}")
        
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            # 获取所有表名
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            if not tables:
                conn.close()
                raise DashboardError("E003", "数据库中没有表")
            
            # 使用第一个表进行查询
            table_name = tables[0]
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 100")
            columns = [desc[0] for desc in cursor.description]
            rows_data = cursor.fetchall()
            conn.close()
            
            rows = []
            for row in rows_data:
                rows.append({col: str(val) for col, val in zip(columns, row)})
            
            return {"type": "database", "scheme": scheme, "fields": columns, "rows": rows}
        except sqlite3.Error as e:
            raise DashboardError("E003", f"SQLite 查询失败: {e}")
    
    elif scheme in ("postgresql", "postgres"):
        # PostgreSQL 需要 psycopg2
        try:
            import psycopg2
        except ImportError:
            raise DashboardError("E003", "PostgreSQL 需要 psycopg2 库，请先安装")
        
        try:
            conn = psycopg2.connect(connection_string.replace("postgres://", "postgresql://"))
            cursor = conn.cursor()
            # 获取所有表
            cursor.execute("SELECT tablename FROM pg_tables WHERE schemaname = 'public'")
            tables = [row[0] for row in cursor.fetchall()]
            if not tables:
                conn.close()
                raise DashboardError("E003", "数据库中没有表")
            
            table_name = tables[0]
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 100")
            columns = [desc[0] for desc in cursor.description]
            rows_data = cursor.fetchall()
            conn.close()
            
            rows = []
            for row in rows_data:
                rows.append({col: str(val) for col, val in zip(columns, row)})
            
            return {"type": "database", "scheme": scheme, "fields": columns, "rows": rows}
        except Exception as e:
            raise DashboardError("E003", f"PostgreSQL 查询失败: {e}")
    
    else:
        raise DashboardError("E003", f"不支持的数据库类型: {scheme}")


# ---------------------------------------------------------------------------
# 图表配置推荐模块
# ---------------------------------------------------------------------------

def recommend_chart(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    根据数据特征推荐图表类型及 ECharts 配置。

    规则：
    - 时间序列数据（字段名含 date/time/日期/时间/月/年）→ 折线图
    - 分类数据（字段数 >= 2 且首字段为类别）→ 柱状图
    - 单字段数值分布 → 饼图

    参数:
        data: 标准化数据字典（来自 parse_data_source）

    返回:
        ECharts 配置 JSON 字典
    """
    if not data or "fields" not in data:
        raise DashboardError("E004", "数据字典格式无效")

    fields = data.get("fields", [])
    rows = data.get("rows", [])

    if not fields or not rows:
        raise DashboardError("E004", "数据为空，无法推荐图表")

    # 检测时间序列 - 扩展关键词列表
    time_keywords = [
        "date", "time", "datetime", "timestamp",
        "日期", "时间", "年", "月", "日",
        "year", "month", "day", "hour", "minute", "second",
        "week", "quarter", "季度", "周"
    ]
    
    # 检查是否有时间字段
    has_time = False
    time_field = None
    for f in fields:
        f_lower = f.lower()
        if any(k in f_lower for k in time_keywords):
            has_time = True
            time_field = f
            break

    # 检测数值字段
    numeric_fields = []
    for f in fields:
        try:
            sample = rows[0].get(f, "0")
            if isinstance(sample, (int, float)):
                numeric_fields.append(f)
            else:
                # 尝试转换为数值
                float(sample)
                numeric_fields.append(f)
        except (ValueError, TypeError):
            continue

    if not numeric_fields:
        raise DashboardError("E004", "未找到可数值化的字段")

    # 推荐图表类型
    if has_time and time_field:
        chart_type = "line"
        x_field = time_field
    elif len(fields) >= 2:
        chart_type = "bar"
        x_field = fields[0]
    else:
        chart_type = "pie"
        x_field = fields[0]

    y_field = numeric_fields[0]

    # 构建 ECharts 配置
    config = {
        "chart_type": chart_type,
        "title": {"text": f"{y_field} 分析", "left": "center"},
        "tooltip": {"trigger": "axis" if chart_type != "pie" else "item"},
        "legend": {"data": [y_field], "bottom": 10},
        "grid": {"left": 50, "right": 30, "top": 60, "bottom": 60},
        "xAxis": {"type": "category", "data": [str(r.get(x_field, "")) for r in rows]},
        "yAxis": {"type": "value", "name": y_field},
        "series": [{
            "name": y_field,
            "type": chart_type,
            "data": [float(r.get(y_field, 0) or 0) for r in rows],
            "smooth": chart_type == "line",
        }],
    }

    if chart_type == "pie":
        config["series"][0]["radius"] = "60%"
        config["series"][0]["data"] = [
            {"name": str(r.get(x_field, "")), "value": float(r.get(y_field, 0) or 0)}
            for r in rows
        ]

    return config


# ---------------------------------------------------------------------------
# Django 项目集成模块
# ---------------------------------------------------------------------------

def generate_django_view(project_path: str, chart_config: Dict[str, Any]) -> str:
    """
    生成可嵌入 Django 项目的 dashboard 视图代码。

    参数:
        project_path: Django 项目路径（用于验证）
        chart_config: 图表配置（来自 recommend_chart）

    返回:
        Django views.py 代码字符串
    """
    if not os.path.isdir(project_path):
        raise DashboardError("E005", f"项目路径不存在: {project_path}")

    chart_json = json.dumps(chart_config, ensure_ascii=False, indent=2)

    code = f'''"""
dashboard 视图模块 - 由 django-dashing 技能生成
"""
from django.shortcuts import render
from django.http import JsonResponse
import json

# 图表配置（预生成）
CHART_CONFIG = {chart_json}


def dashboard(request):
    """数据看板主页：渲染图表配置到模板。"""
    context = {{
        "chart_config_json": json.dumps(CHART_CONFIG, ensure_ascii=False),
        "page_title": "数据看板",
    }}
    return render(request, "dashboard/dashboard.html", context)


def dashboard_data(request):
    """数据接口：返回最新图表数据（用于前端刷新）。"""
    return JsonResponse(CHART_CONFIG)


def dashboard_config(request):
    """配置接口：返回当前布局与刷新策略。"""
    layout = {{
        "grid": [12, 8],
        "modules": [
            {{"id": "chart-1", "x": 0, "y": 0, "w": 6, "h": 4, "chart": "main"}},
            {{"id": "chart-2", "x": 6, "y": 0, "w": 6, "h": 4, "chart": "secondary"}},
        ],
    }}
    refresh = {{"strategy": "polling", "interval_seconds": 30}}
    return JsonResponse({{"layout": layout, "refresh": refresh}})
'''
    return code


# ---------------------------------------------------------------------------
# 模块化布局模块
# ---------------------------------------------------------------------------

def generate_layout(module_count: int, columns: int = 12, row_height: int = 100) -> Dict[str, Any]:
    """
    生成可拖拽的网格布局配置。

    参数:
        module_count: 模块数量（>= 1）
        columns: 网格列数（默认 12）
        row_height: 行高像素（默认 100）

    返回:
        layout.json 字典
    """
    if module_count < 1:
        raise DashboardError("E006", "模块数量必须 >= 1")
    if columns < 1:
        raise DashboardError("E006", "列数必须 >= 1")
    if row_height <= 0:
        raise DashboardError("E006", "行高必须为正数")

    modules = []
    per_row = max(1, columns // 2)  # 每行最多 2 个模块（简单布局）

    for i in range(module_count):
        row = i // per_row
        col = i % per_row
        module = {
            "id": f"module-{i+1}",
            "x": col * (columns // per_row),
            "y": row * 2,  # 每个模块高度为 2 行
            "w": columns // per_row,
            "h": 2,
            "content": f"模块 {i+1} 内容区域",
        }
        modules.append(module)

    return {
        "grid": {"columns": columns, "row_height": row_height, "margin": [10, 10]},
        "modules": modules,
        "drag_enabled": True,
        "resize_enabled": True,
    }


# ---------------------------------------------------------------------------
# 数据刷新策略模块
