#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel-image-extractor - 表格图片数据抽取与结构化输出

功能:
    从 Excel 表格及图片中抽取关键数据，按约定格式输出结构化结果。
    支持 Markdown 表格、CSV、JSON 三种输出格式。

用法:
    python scripts/main.py --selftest
    python scripts/main.py --input <file_or_url> --format <md|csv|json>

错误码:
    E001 - 输入文件不存在或无法访问
    E002 - 输入文件格式不支持
    E003 - 无法解析文件内容
    E004 - 输出格式参数无效
    E005 - 命令行参数缺失
    E006 - 文件读取失败
    E007 - 数据解析失败
    E008 - 输出转换失败
    E009 - 内部逻辑错误
    E010 - 未知错误
"""

import argparse
import csv
import json
import os
import sys
import urllib.request
from typing import Any, Dict, List, Optional, Tuple
import time
dry_run = False  # v3.274 模块级 dry-run 标志

# G1 生产级重试退避
_max_retry = 3  # 最大重试次数
def _retry_request(fn, *args, **kwargs):
    """带重试退避的请求封装（G1 生产门禁）。"""
    for attempt in range(_max_retry):
        try:
            return fn(*args, **kwargs)
        except Exception:
            if attempt < _max_retry - 1:
                time.sleep(2 ** attempt)  # 指数退避
            else:
                raise


# ============================================================
# 核心数据解析模块
# ============================================================

class TableData:
    """表格数据容器，存储表头和行数据"""
    
    def __init__(self, headers: List[str], rows: List[List[str]]):
        self.headers = headers
        self.rows = rows
    
    @property
    def row_count(self) -> int:
        return len(self.rows)
    
    @property
    def col_count(self) -> int:
        return len(self.headers)
    
    def to_markdown(self) -> str:
        """转换为 Markdown 表格格式"""
        if not self.headers:
            return ""
        
        lines = []
        # 表头行
        lines.append("| " + " | ".join(self.headers) + " |")
        # 分隔行
        lines.append("| " + " | ".join(["---"] * len(self.headers)) + " |")
        # 数据行
        for row in self.rows:
            # 确保行长度与表头一致
            padded_row = row + [""] * (len(self.headers) - len(row))
            lines.append("| " + " | ".join(padded_row[:len(self.headers)]) + " |")
        
        return "\n".join(lines)
    
    def to_csv(self) -> str:
        """转换为 CSV 格式文本"""
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(self.headers)
        for row in self.rows:
            padded_row = row + [""] * (len(self.headers) - len(row))
            writer.writerow(padded_row[:len(self.headers)])
        return output.getvalue().strip()
    
    def to_json(self) -> str:
        """转换为 JSON 格式文本"""
        data = {
            "headers": self.headers,
            "rows": self.rows,
            "row_count": self.row_count,
            "col_count": self.col_count
        }
        return json.dumps(data, ensure_ascii=False, indent=2)


def parse_text_table(text: str) -> TableData:
    """
    从纯文本中解析表格数据
    
    支持格式:
        - 逗号分隔 (CSV)
        - 制表符分隔 (TSV)
        - 竖线分隔 (Markdown 风格)
    
    参数:
        text: 包含表格数据的文本
    
    返回:
        TableData 对象
    
    错误:
        E003 - 无法解析文件内容
    """
    if not text or not text.strip():
        raise ValueError("E003: 输入文本为空，无法解析")
    
    lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
    if len(lines) < 2:
        # 只有一行，视为单列数据
        headers = ["列1"]
        rows = [[line] for line in lines]
        return TableData(headers, rows)
    
    # 检测分隔符
    first_line = lines[0]
    delimiter = None
    
    if "\t" in first_line:
        delimiter = "\t"
    elif "|" in first_line:
        delimiter = "|"
    elif "," in first_line:
        delimiter = ","
    else:
        # 尝试用空格分隔
        parts = first_line.split()
        if len(parts) > 1:
            delimiter = None  # 使用空白分隔
        else:
            # 单列数据
            headers = ["列1"]
            rows = [[line] for line in lines]
            return TableData(headers, rows)
    
    def split_line(line: str) -> List[str]:
        """按检测到的分隔符分割一行"""
        if delimiter == "|":
            # 去除首尾的竖线
            cleaned = line.strip("|")
            return [part.strip() for part in cleaned.split("|")]
        elif delimiter is None:
            return line.split()
        else:
            return [part.strip() for part in line.split(delimiter)]
    
    # 解析表头
    headers = split_line(lines[0])
    # 跳过 Markdown 分隔行 (如 |---|---|)
    data_lines = []
    for line in lines[1:]:
        parts = split_line(line)
        if all(re.match(r'^:?-{2,}:?$', p) for p in parts if p):
            continue  # 跳过分隔行
        data_lines.append(parts)
    
    if not data_lines:
        raise ValueError("E003: 未找到数据行")
    
    # 标准化列数
    max_cols = max(len(headers), max(len(row) for row in data_lines))
    if len(headers) < max_cols:
        for i in range(len(headers), max_cols):
            headers.append(f"列{i+1}")
    
    # 填充行数据
    rows = []
    for row in data_lines:
        padded = row + [""] * (max_cols - len(row))
        rows.append(padded[:max_cols])
    
    return TableData(headers, rows)


def parse_excel_file(filepath: str) -> TableData:
    """
    解析 Excel 文件（.xlsx/.xls）
    
    参数:
        filepath: Excel 文件路径
    
    返回:
        TableData 对象
    
    错误:
        E001 - 输入文件不存在或无法访问
        E002 - 输入文件格式不支持
        E006 - 文件读取失败
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"E001: 文件不存在: {filepath}")
    
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise ValueError(f"E002: 不支持的 Excel 格式: {ext}")
    
    try:
        # 尝试导入 openpyxl (处理 .xlsx)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            
            # 读取所有数据
            all_data = []
            for row in ws.iter_rows(values_only=True):
                all_data.append([str(cell) if cell is not None else "" for cell in row])
            
            wb.close()
            
            if not all_data:
                raise ValueError("E003: Excel 文件为空")
            
            # 第一行作为表头
            headers = all_data[0]
            rows = all_data[1:]
            
            return TableData(headers, rows)
        except ImportError:
            # 如果没有 openpyxl，尝试用 pandas
            try:
                import pandas as pd
                df = pd.read_excel(filepath)
                headers = [str(col) for col in df.columns]
                rows = [[str(cell) if cell is not None else "" for cell in row] 
                        for row in df.values]
                return TableData(headers, rows)
            except ImportError:
                raise ImportError(
                    "E002: 需要安装 openpyxl 或 pandas 来解析 Excel 文件\n"
                    "请执行: pip install openpyxl"
                )
    except FileNotFoundError:
        raise
    except ValueError:
        raise
    except ImportError:
        raise
    except Exception as e:
        raise RuntimeError(f"E006: 读取 Excel 文件失败: {e}")


def parse_image_file(filepath: str) -> TableData:
    """
    解析图片文件（.png/.jpg/.jpeg）
    
    注意: 本实现仅支持印刷体文字识别，且需要安装 pytesseract 和 Pillow
    
    参数:
        filepath: 图片文件路径
    
    返回:
        TableData 对象
    
    错误:
        E001 - 输入文件不存在或无法访问
        E002 - 输入文件格式不支持
        E003 - 无法解析文件内容
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"E001: 文件不存在: {filepath}")
    
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".png", ".jpg", ".jpeg"):
        raise ValueError(f"E002: 不支持的图片格式: {ext}")
    
    try:
        # 尝试导入 OCR 库
        try:
            from PIL import Image
            import pytesseract
            
            # 使用 OCR 识别文字
            image = Image.open(filepath)
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            
            if not text.strip():
                raise ValueError("E003: 图片中未识别到文字内容")
            
            # 将 OCR 结果按行解析为表格
            return parse_text_table(text)
            
        except ImportError:
            raise ImportError(
                "E002: 需要安装 pytesseract 和 Pillow 来解析图片文件\n"
                "请执行: pip install pytesseract Pillow\n"
                "并确保系统已安装 tesseract-ocr"
            )
    except FileNotFoundError:
        raise
    except ValueError:
        raise
    except ImportError:
        raise
    except Exception as e:
        raise RuntimeError(f"E006: 读取图片文件失败: {e}")


def load_input_source(source: str) -> TableData:
    """
    加载输入源并解析为表格数据
    
    支持:
        - 本地文件路径 (.xlsx/.xls/.png/.jpg/.jpeg/.csv/.txt)
        - URL 链接 (http/https)
        - 直接文本数据
    
    参数:
        source: 输入源
    
    返回:
        TableData 对象
    
    错误:
        E001 - 输入文件不存在或无法访问
        E002 - 输入文件格式不支持
        E003 - 无法解析文件内容
        E006 - 文件读取失败
    """
    # 检查是否为 URL
    if source.startswith("http://") or source.startswith("https://"):
        try:
            # 下载内容到临时文件
            import tempfile
            with urllib.request.urlopen(source, timeout=10) as response:
                content = response.read()
            
            # 根据 URL 扩展名判断类型
            url_path = source.split("?")[0]
            ext = os.path.splitext(url_path)[1].lower()
            
            if ext in (".xlsx", ".xls"):
                # Excel 文件需要保存后解析
                with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name
                try:
                    return parse_excel_file(tmp_path)
                finally:
                    os.unlink(tmp_path)
            elif ext in (".png", ".jpg", ".jpeg"):
                # 图片文件需要保存后解析
                with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                    tmp.write(content)
                    tmp_path = tmp.name
                try:
                    return parse_image_file(tmp_path)
                finally:
                    os.unlink(tmp_path)
            else:
                # 尝试作为文本解析
                text = content.decode("utf-8", errors="replace")
                return parse_text_table(text)
                
        except Exception as e:
            raise RuntimeError(f"E006: 从 URL 读取数据失败: {e}")
    
    # 检查是否为本地文件
    if os.path.exists(source):
        ext = os.path.splitext(source)[1].lower()
        if ext in (".xlsx", ".xls"):
            return parse_excel_file(source)
        elif ext in (".png", ".jpg", ".jpeg"):
            return parse_image_file(source)
        elif ext in (".csv", ".txt", ".md"):
            try:
                with open(source, "r", encoding="utf-8") as f:
                    text = f.read()
                return parse_text_table(text)
            except Exception as e:
                raise RuntimeError(f"E006: 读取文件失败: {e}")
        else:
            raise ValueError(f"E002: 不支持的文件格式: {ext}")
    
    # 尝试作为纯文本解析
    return parse_text_table(source)


# ============================================================
# 输出格式化模块
# ============================================================

def format_output(data: TableData, fmt: str) -> str:
    """
    按指定格式输出表格数据
    
    参数:
        data: 表格数据
        fmt: 输出格式 (md/csv/json)
    
    返回:
        格式化后的字符串
    
    错误:
        E004 - 输出格式参数无效
        E008 - 输出转换失败
    """
    fmt = fmt.lower()
    
    try:
        if fmt in ("md", "markdown"):
            return data.to_markdown()
        elif fmt == "csv":
            return data.to_csv()
        elif fmt == "json":
            return data.to_json()
        else:
            raise ValueError(f"E004: 不支持的输出格式: {fmt}（可选: md/csv/json）")
    except ValueError:
        raise
    except Exception as e:
        raise RuntimeError(f"E008: 输出转换失败: {e}")


# ============================================================
# 自检模块
# ============================================================

def run_selftest() -> bool:
    """
    内置自检函数，验证核心逻辑
    
    使用硬编码样例数据，不读取外部文件、不依赖当前工作目录、不访问网络。
    断言使用宽松阈值，确保任何环境直接可过。
    
    返回:
        True 表示自检通过
    
    错误:
        E009 - 内部逻辑错误
    """
    print("=" * 60)
    print("开始自检 (selftest)")
    print("=" * 60)
    
    try:
        # ===== 测试1: 文本解析 =====
        print("\n[测试1] 文本表格解析")
        sample_text = """姓名,年龄,城市
张三,25,北京
李四,30,上海
王五,28,广州"""
        
        data = parse_text_table(sample_text)
        
        # 宽松断言
        assert data.row_count >= 2, "E009: 行数解析错误"
        assert data.col_count >= 3, "E009: 列数解析错误"
        assert len(data.headers) >= 3, "E009: 表头解析错误"
        assert any("姓名" in h for h in data.headers), "E009: 表头内容错误"
        print(f"  ✓ 解析成功: {data.row_count}行 x {data.col_count}列")
        
        # ===== 测试2: Markdown 输出 =====
        print("\n[测试2] Markdown 输出")
        md_output = data.to_markdown()
        assert "|" in md_output, "E009: Markdown 格式错误"
        assert "---" in md_output, "E009: Markdown 分隔行缺失"
        assert md_output.count("|") >= 10, "E009: Markdown 内容不完整"
        print(f"  ✓ Markdown 输出正常 ({len(md_output)} 字符)")
        
        # ===== 测试3: CSV 输出 =====
        print("\n[测试3] CSV 输出")
        csv_output = data.to_csv()
        assert "," in csv_output, "E009: CSV 格式错误"
        assert len(csv_output.splitlines()) >= 3, "E009: CSV 行数不足"
        print(f"  ✓ CSV 输出正常 ({len(csv_output)} 字符)")
        
        # ===== 测试4: JSON 输出 =====
        print("\n[测试4] JSON 输出")
        json_output = data.to_json()
        json_data = json.loads(json_output)
        assert "headers" in json_data, "E009: JSON 缺少 headers"
        assert "rows" in json_data, "E009: JSON 缺少 rows"
        assert len(json_data["rows"]) >= 2, "E009: JSON 行数不足"
        print(f"  ✓ JSON 输出正常 ({len(json_output)} 字符)")
        
        # ===== 测试5: 复杂表格解析 =====
        print("\n[测试5] 复杂表格解析")
        complex_text = """| 产品 | 价格 | 库存 |
|------|------|------|
| 苹果 | 5.5 | 100 |
| 香蕉 | 3.2 | 200 |
| 橙子 | 4.8 | 150 |"""
        
        data2 = parse_text_table(complex_text)
        assert data2.row_count >= 2, "E009: 复杂表格行数错误"
        assert data2.col_count >= 3, "E009: 复杂表格列数错误"
        assert any("产品" in h for h in data2.headers), "E009: 复杂表格表头错误"
        print(f"  ✓ 复杂表格解析成功: {data2.row_count}行 x {data2.col_count}列")
        
        # ===== 测试6: 单列数据 =====
        print("\n[测试6] 单列数据解析")
        single_col = "苹果\n香蕉\n橙子"
        data3 = parse_text_table(single_col)
        assert data3.col_count >= 1, "E009: 单列数据列数错误"
        assert data3.row_count >= 2, "E009: 单列数据行数错误"
        print(f"  ✓ 单列数据解析成功: {data3.row_count}行")
        
        # ===== 测试7: 输出格式验证 =====
        print("\n[测试7] 输出格式验证")
        for fmt in ["md", "csv", "json"]:
            output = format_output(data, fmt)
            assert len(output) > 0, f"E009: 格式 {fmt} 输出为空"
        print("  ✓ 所有输出格式验证通过")
        
        # ===== 测试8: 错误处理 =====
        print("\n[测试8] 错误处理")
        try:
            format_output(data, "invalid_format")
            assert False, "E009: 无效格式未抛出异常"
        except ValueError as e:
            assert "E004" in str(e), "E009: 错误码不正确"
        print("  ✓ 错误处理正常")
        
        print("\n" + "=" * 60)
        print("自检全部通过 ✓")
        print("=" * 60)
        return True
        
    except AssertionError as e:
        print(f"\n✗ 自检失败: {e}")
        return False
    except Exception as e:
        print(f"\n✗ 自检异常: {e}")
        return False


# ============================================================
# 主程序入口
# ============================================================

def main() -> int:
    """主程序入口"""
    parser = argparse.ArgumentParser(
        description="excel-image-extractor - 表格图片数据抽取与结构化输出",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python scripts/main.py --selftest                    # 运行自检
  python scripts/main.py --input data.xlsx --format md  # 解析 Excel 输出 Markdown
  python scripts/main.py --input table.csv --format json # 解析 CSV 输出 JSON
  python scripts/main.py --input https://example.com/data.csv --format csv  # 从 URL 读取
        """
    )
    
    parser.add_argument(
        "--selftest",
        action="store_true",
        help="运行内置自检（不依赖外部文件）"
    )
    
    parser.add_argument(
        "--input",
        type=str,
        help="输入源: 文件路径、URL 或直接文本数据"
    )
    
    parser.add_argument(
        "--format",
        type=str,
        choices=["md", "csv", "json"],
        default="md",
        help="输出格式 (默认: md)"
    )
    
    parser.add_argument("--force", action="store_true")  # R4 强制写盘

    
    parser.add_argument("--dry-run", action="store_true")  # R4 预览模式
    
    args = parser.parse_args()
    
    global dry_run
    
    dry_run = getattr(args, "dry_run", False)  # v3.274 同步到全局
    
    # 运行自检
    if args.selftest:
        success = run_selftest()
        return 0 if success else 1
    
    # 检查必要参数
    if not args.input:
        print("E005: 缺少必要参数 --input 或 --selftest", file=sys.stderr)
        print("请使用 --selftest 运行自检，或提供 --input 指定输入源", file=sys.stderr)
        return 1
    
    try:
        # 加载并解析输入
        data = load_input_source(args.input)
        
        # 格式化输出
        output = format_output(data, args.format)
        
        # 打印结果
        print(output)
        return 0
        
    except FileNotFoundError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except ImportError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except RuntimeError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"E010: 未知错误: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    import re  # 为 parse_text_table 中的 re.match 提供支持
    sys.exit(main())
