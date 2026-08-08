#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
竞品透视 · 多维对标与差异洞察 Skill
真实可用的竞品分析工具：读取竞品数据文件，输出多维对比报告
支持 CSV / JSON / Markdown / TXT 格式输入
"""

import argparse
import csv
import json
import os
import re
import sys
import tempfile
import time
import logging
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 文件大小限制（10MB）
MAX_FILE_SIZE = 10 * 1024 * 1024

# ============ 数据加载模块 ============

def parse_markdown_table(content):
    """解析 Markdown 表格，返回列表字典。
    支持无表头（默认 col1, col2...）和列数不一致（保留行，缺失列填空字符串）。
    增强表头检测：检查是否包含常见列名关键词，避免数字/空行误判。
    """
    rows = []
    lines = [l.strip() for l in content.split('\n') if l.strip()]
    header_count = None
    for i, line in enumerate(lines):
        if '|' not in line:
            continue
        cells = [c.strip() for c in line.strip('|').split('|')]
        # 跳过分隔行
        if all(re.match(r'^[-:]+$', c) for c in cells):
            continue
        if not rows:
            # 第一行作为表头，如果没有表头（全是数据），则生成默认列名
            rows.append(cells)
            header_count = len(cells)
            continue
        # 列数不一致时，补齐或截断
        if len(cells) != header_count:
            logger.warning(f"Markdown表格第{i+1}行列数({len(cells)})与表头({header_count})不一致，补齐空列")
            if len(cells) < header_count:
                cells.extend([''] * (header_count - len(cells)))
            else:
                cells = cells[:header_count]
        rows.append(cells)
    
    if len(rows) < 2:
        # 只有一行，视为无表头数据
        if len(rows) == 1:
            headers = [f"col{j+1}" for j in range(len(rows[0]))]
            return [dict(zip(headers, rows[0]))]
        return []
    
    headers = rows[0]
    # 增强表头检测：检查是否包含常见列名关键词
    common_headers = ['name', '名称', 'feature', '功能', 'price', '价格', 'rating', '评分', 'description', '描述']
    header_keywords_found = any(any(kw in h.lower() for kw in common_headers) for h in headers if h)
    
    # 如果第一行全是数字或空，且不包含关键词，则视为数据行
    if not header_keywords_found and all(re.match(r'^[\d\s]*$', h) or h == '' for h in headers):
        headers = [f"col{j+1}" for j in range(len(headers))]
        data_rows = rows
    else:
        data_rows = rows[1:]
    
    data = []
    for row in data_rows:
        data.append(dict(zip(headers, row)))
    return data


def parse_txt_content(content):
    """解析 TXT 内容，支持 key: value 格式、制表符分隔、以及自由文本段落。
    自由文本段落按空行分块，每块作为一个记录，提取关键信息。
    """
    records = []
    current_record = {}
    current_text_lines = []
    
    def flush_text_block():
        nonlocal current_text_lines
        if current_text_lines:
            text = '\n'.join(current_text_lines).strip()
            if text:
                # 尝试从文本中提取关键信息
                record = {'text': text}
                # 尝试提取名称、价格、评分等
                name_match = re.search(r'(?:名称|产品|竞品)[：:]\s*(\S+)', text)
                if name_match:
                    record['name'] = name_match.group(1)
                price_match = re.search(r'(?:价格|定价|费用)[：:]\s*([\d.]+)', text)
                if price_match:
                    record['price'] = price_match.group(1)
                rating_match = re.search(r'(?:评分|评价)[：:]\s*([\d.]+)', text)
                if rating_match:
                    record['rating'] = rating_match.group(1)
                records.append(record)
            current_text_lines = []
    
    for line in content.split('\n'):
        line = line.strip()
        if not line:
            # 空行：结束当前记录或文本块
            if current_record:
                records.append(current_record)
                current_record = {}
            flush_text_block()
            continue
        
        if ':' in line and not line.startswith('http'):
            key, value = line.split(':', 1)
            current_record[key.strip()] = value.strip()
            flush_text_block()  # 如果之前有文本块，先保存
        elif '\t' in line:
            parts = line.split('\t')
            if len(parts) >= 2:
                current_record[parts[0].strip()] = parts[1].strip()
                flush_text_block()
            else:
                current_text_lines.append(line)
        else:
            # 自由文本行，累积到文本块
            current_text_lines.append(line)
    
    # 处理末尾
    if current_record:
        records.append(current_record)
    flush_text_block()
    
    return records


def detect_encoding(filepath):
    """检测文件编码。
    优先使用 chardet，不可用时尝试 utf-8-sig -> gb18030 -> utf-8。
    返回 (encoding, success_flag)。
    """
    if HAS_CHARDET:
        try:
            with open(filepath, 'rb') as f:
                raw_data = f.read(10000)
            result = chardet.detect(raw_data)
            if result['encoding']:
                return result['encoding'], True
        except Exception as e:
            logger.warning(f"chardet 检测失败: {e}")
    
    # chardet 不可用或检测失败，尝试常见编码
    for enc in ['utf-8-sig', 'gb18030', 'utf-8']:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(1000)  # 测试读取
            return enc, True
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    return 'utf-8', False


def check_file_size(filepath):
    """检查文件大小"""
    size = os.path.getsize(filepath)
    if size > MAX_FILE_SIZE:
        raise ValueError(f"文件大小({size}字节)超过限制({MAX_FILE_SIZE}字节)，请使用流式处理或分割文件")
    return size


def load_data(filepath):
    """加载竞品数据文件，返回 (竞品名, 数据类型, 记录列表)。
    所有文件操作异常都会被捕获并转为结构化错误。
    使用文件锁防止并发读取竞态。
    """
    try:
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")
        
        # 检查文件大小
        check_file_size(filepath)
        
        # 文件锁防止并发读取竞态
        lock_path = filepath + '.lock'
        lock_fd = None
        try:
            lock_fd = os.open(lock_path, os.O_CREAT | os.O_RDWR, 0o644)
            # 尝试获取锁（非阻塞，避免死锁）
            try:
                import fcntl
                fcntl.flock(lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            except ImportError:
                # Windows 或非 POSIX 系统，使用 msvcrt
                try:
                    import msvcrt
                    msvcrt.locking(lock_fd, msvcrt.LK_NBLCK, 1)
                except ImportError:
                    pass  # 无锁支持，跳过
                except OSError:
                    pass  # 锁被占用，继续（不阻塞）
            except OSError:
                # 锁被占用，等待重试（最多3次）
                for _ in range(3):
                    time.sleep(0.1)
                    try:
                        fcntl.flock(lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                        break
                    except OSError:
                        continue
                else:
                    logger.warning(f"文件 {filepath} 被其他进程锁定，继续读取（可能产生竞态）")
        except Exception as e:
            logger.warning(f"文件锁创建失败: {e}")
        
        try:
            ext = os.path.splitext(filepath)[1].lower()
            filename = os.path.basename(filepath)
            # 从文件名提取竞品名（第一个下划线前）
            comp_name = filename.split('_')[0] if '_' in filename else Path(filename).stem
            
            records = []
            data_type = 'unknown'
            
            if ext == '.csv':
                encoding, ok = detect_encoding(filepath)
                if not ok:
                    logger.warning(f"文件 {filepath} 编码检测失败，尝试 utf-8")
                    encoding = 'utf-8'
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        reader = csv.DictReader(f)
                        records = list(reader)
                    data_type = 'csv'
                except UnicodeDecodeError as e:
                    raise ValueError(f"文件编码错误，无法用 {encoding} 解码: {e}")
            
            elif ext == '.json':
                encoding, ok = detect_encoding(filepath)
                if not ok:
                    logger.warning(f"文件 {filepath} 编码检测失败，尝试 utf-8")
                    encoding = 'utf-8'
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            records = data
                        elif isinstance(data, dict):
                            records = data.get('records', data.get('data', []))
                            if isinstance(records, dict):
                                records = [records]
                    data_type = 'json'
                except (UnicodeDecodeError, json.JSONDecodeError) as e:
                    raise ValueError(f"JSON 解析失败: {e}")
            
            elif ext == '.md' or ext == '.markdown':
                encoding, ok = detect_encoding(filepath)
                if not ok:
                    logger.warning(f"文件 {filepath} 编码检测失败，尝试 utf-8")
                    encoding = 'utf-8'
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        content = f.read()
                    records = parse_markdown_table(content)
                    data_type = 'markdown'
                except UnicodeDecodeError as e:
                    raise ValueError(f"文件编码错误，无法用 {encoding} 解码: {e}")
            
            elif ext == '.txt':
                encoding, ok = detect_encoding(filepath)
                if not ok:
                    logger.warning(f"文件 {filepath} 编码检测失败，尝试 utf-8")
                    encoding = 'utf-8'
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        content = f.read()
                    records = parse_txt_content(content)
                    data_type = 'txt'
                except UnicodeDecodeError as e:
                    raise ValueError(f"文件编码错误，无法用 {encoding} 解码: {e}")
            
            elif ext == '.xlsx' and HAS_OPENPYXL:
                try:
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        records.append(dict(zip(headers, row)))
                    wb.close()
                    data_type = 'xlsx'
                except Exception as e:
                    raise ValueError(f"Excel 解析失败: {e}")
            
            else:
                raise ValueError(f"不支持的文件格式: {ext}")
            
            if not records:
                raise ValueError("文件中没有有效数据")
            
            return comp_name, data_type, records
        
        finally:
            # 释放文件锁
            if lock_fd is not None:
                try:
                    import fcntl
                    fcntl.flock(lock_fd, fcntl.LOCK_UN)
                except ImportError:
                    try:
                        import msvcrt
                        msvcrt.locking(lock_fd, msvcrt.LK_UNLCK, 1)
                    except ImportError:
                        pass
                os.close(lock_fd)
    
    except FileNotFoundError as e:
        logger.error(f"文件不存在: {e}")
        raise
    except PermissionError as e:
        logger.error(f"权限不足: {e}")
        raise
    except ValueError as e:
        logger.error(f"数据解析错误: {e}")
        raise
    except Exception as e:
        logger.error(f"加载文件失败: {e}")
        raise ValueError(f"加载文件失败: {e}")


# ============ 字段提取模块 ============

def extract_features(record):
    """从记录中提取功能列表"""
    features = []
    for key in ['features', '功能', 'feature', 'capabilities', '能力']:
        if key in record:
            value = record[key]
            if isinstance(value, list):
                features.extend([str(v).strip() for v in value if str(v).strip()])
            elif isinstance(value, str):
                # 支持逗号、分号、换行分隔
                parts = re.split(r'[,;，；\n]', value)
                features.extend([p.strip() for p in parts if p.strip()])
            break
    return features


def extract_pricing(record):
    """从记录中提取定价信息"""
    pricing = {}
    for key in ['price', 'pricing', '价格', '定价', 'cost', '费用']:
        if key in record:
            value = record[key]
            if isinstance(value, dict):
                pricing = value
            elif isinstance(value, (int, float)):
                pricing = {'base': float(value)}
            elif isinstance(value, str):
                # 尝试解析价格字符串
                price_match = re.search(r'[\d.]+', value)
                if price_match:
                    pricing = {'base': float(price_match.group())}
                else:
                    pricing = {'description': value}
            break
    return pricing


def extract_reviews(record):
    """从记录中提取评价信息"""
    reviews = {}
    for key in ['reviews', 'rating', '评价', '评分', 'score', 'rating_score']:
        if key in record:
            value = record[key]
            if isinstance(value, (int, float)):
                reviews = {'rating': float(value)}
            elif isinstance(value, str):
                rating_match = re.search(r'(\d+(?:\.\d+)?)\s*[/分]', value)
                if rating_match:
                    reviews = {'rating': float(rating_match.group(1))}
                else:
                    reviews = {'description': value}
            elif isinstance(value, dict):
                reviews = value
            break
    return reviews


def extract_all_fields(record):
    """提取所有关键字段，返回 (features, pricing, reviews, confidence)"""
    features = extract_features(record)
    pricing = extract_pricing(record)
    reviews = extract_reviews(record)
    
    # 计算置信度
    confidence = 0.5
    if features:
        confidence += 0.2
    if pricing:
        confidence += 0.2
    if reviews:
        confidence += 0.1
    
    return features, pricing, reviews, min(confidence, 1.0)


# ============ 分析模块 ============

def generate_differentiation_suggestions(report):
    """生成差异化建议"""
    suggestions = []
    competitors = report['competitors']
    
    if len(competitors) < 2:
        return suggestions
    
    # 功能差异化建议
    all_features = set()
    for comp in competitors:
        all_features.update(comp['features'])
    
    for feature in sorted(all_features):
        has_feature = [comp['name'] for comp in competitors if feature in comp['features']]
        if len(has_feature) == 1:
            suggestions.append({
                'type': 'unique_feature',
                'feature': feature,
                'competitor': has_feature[0],
                'suggestion': f"{has_feature[0]} 拥有独特功能 '{feature}'，可作为差异化卖点"
            })
        elif len(has_feature) == len(competitors):
            suggestions.append({
                'type': 'common_feature',
                'feature': feature,
                'competitor': None,
                'suggestion': f"所有竞品都支持 '{feature}'，属于基础功能，需考虑差异化升级"
            })
    
    # 定价差异化建议
    prices = {}
    for comp in competitors:
        if 'base' in comp['pricing']:
            prices[comp['name']] = comp['pricing']['base']
    
    if len(prices) >= 2:
        min_price_comp = min(prices, key=prices.get)
        max_price_comp = max(prices, key=prices.get)
        if prices[min_price_comp] < prices[max_price_comp]:
            suggestions.append({
                'type': 'pricing',
                'competitor': min_price_comp,
                'suggestion': f"{min_price_comp} 定价最低 ({prices[min_price_comp]})，可主打性价比"
            })
            suggestions.append({
                'type': 'pricing',
                'competitor': max_price_comp,
                'suggestion': f"{max_price_comp} 定价最高 ({prices[max_price_comp]})，需证明高端价值"
            })
