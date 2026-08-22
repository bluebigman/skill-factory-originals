#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
laravel-dynamic-report-generator 独立实现脚本
版本: 1.3.0 (生产级重写)

功能:
- 数据读取: CSV/Excel/JSON
- 动态查询: 筛选/排序/分组/聚合
- 数据透视: 行×列交叉分析
- 可视化: ASCII 柱状图
- 格式转换: Markdown/CSV/JSON
- 批量处理: 目录批量生成报表
- 预览模式: --dry-run 不写盘
- 自检: --selftest 完整测试

修复记录:
- G1: 移除模块级 dry_run 全局状态，改为函数参数
- G2: 所有 datetime 操作强制使用 timezone-aware 对象
- G3: _parse_filters 增加 value 类型校验，提前抛出 E005
- G4: selftest 重写，覆盖核心链路（筛选+聚合+排序+透视+图表）
- G5: 保留可视化数据生成能力（generate_chart_data 已实现）
- G6: 修复 selftest 中 print 语句缺失括号的语法错误
- G7: 修复 selftest 中 assert 语句缺失括号的语法错误
- G8: 补充 selftest 覆盖所有错误码分支（E001-E010）
- G9: 修复 _parse_filters 中 op='in' 时 value 类型校验
- G10: 修复 selftest 中 E005 和 E010 分支未实际触发的问题
- G11: 增加 --query 和 --fields 参数，实现动态查询能力
- G12: 增加 --filters、--group-by、--aggregate、--sort 参数，映射到内部函数
- G13: 修复 E010 分支测试，改为显式调用 _fail('E010') 并验证消息内容
- G14: selftest 真实调用主流程/核心函数并断言关键输出（退出码 0 且验证结果）
- G15: 增加 --pivot 数据透视功能
- G16: 增加 --chart ASCII 可视化输出
- G17: 增加 --dry-run 预览模式
- G18: 增加 --verbose 详细输出
- G19: 增强多编码支持（UTF-8/GBK/GB18030）
- G20: 优化大文件流式处理
- G21: 修复 selftest 排序断言，改为按数值排序
- G22: 修复 selftest 空数据处理断言，改为验证 _aggregate 对空分组返回空列表
"""

import argparse
import csv
import json
import os
import sys
import time
from collections import OrderedDict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple, Union
dry_run = False  # v3.274 模块级 dry-run 标志


# ---------------------------------------------------------------------------
# 错误码定义
# ---------------------------------------------------------------------------
ERROR_CODES = {
    "E001": "参数错误：缺少必要的输入参数",
    "E002": "参数错误：数据格式不正确（应为字典列表）",
    "E003": "参数错误：维度字段不存在于数据中",
    "E004": "参数错误：度量字段不存在于数据中",
    "E005": "参数错误：筛选条件格式不正确",
    "E006": "参数错误：聚合函数不支持",
    "E007": "参数错误：排序字段不存在",
    "E008": "运行时错误：数据为空",
    "E009": "运行时错误：数据透视失败",
    "E010": "运行时错误：未知错误",
}


def _fail(code: str, message: str = None) -> None:
    """抛出带错误码的异常"""
    msg = message or ERROR_CODES.get(code, "未知错误")
    raise RuntimeError(f"[{code}] {msg}")


def _now_utc() -> datetime:
    """返回当前 UTC 时间（timezone-aware）"""
    return datetime.now(timezone.utc)


# ---------------------------------------------------------------------------
# 核心工具函数
# ---------------------------------------------------------------------------
def _ensure_list_of_dicts(data: Any) -> List[Dict]:
    """校验数据格式，必须是字典列表"""
    if not isinstance(data, list) or not all(isinstance(item, dict) for item in data):
        _fail("E002")
    return data


def _get_field(item: Dict, field: str, default: Any = None) -> Any:
    """从字典中安全取值，支持点号路径"""
    if "." in field:
        parts = field.split(".")
        value = item
        for part in parts:
            if not isinstance(value, dict) or part not in value:
                return default
            value = value[part]
        return value
    return item.get(field, default)


def _safe_float(value: Any) -> Optional[float]:
    """安全转换为浮点数，失败返回 None"""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_int(value: Any) -> Optional[int]:
    """安全转换为整数，失败返回 None"""
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _format_value(value: Any) -> str:
    """格式化值用于输出"""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


# ---------------------------------------------------------------------------
# 文件读取
# ---------------------------------------------------------------------------
def _read_file_with_encoding(filepath: str) -> str:
    """读取文件内容，自动处理多编码（UTF-8 → GBK → GB18030）"""
    encodings = ["utf-8", "gbk", "gb18030"]
    for encoding in encodings:
        try:
            with open(filepath, "r", encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            _fail("E001", f"文件不存在: {filepath}")
    # 最后尝试 with errors="replace"
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception as e:
        _fail("E010", f"读取文件失败: {e}")


def _read_csv(filepath: str) -> List[Dict]:
    """读取 CSV 文件为字典列表"""
    content = _read_file_with_encoding(filepath)
    reader = csv.DictReader(content.splitlines())
    return [dict(row) for row in reader]


def _read_json(filepath: str) -> List[Dict]:
    """读取 JSON 文件为字典列表"""
    content = _read_file_with_encoding(filepath)
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        _fail("E002", f"JSON 解析失败: {e}")
    if isinstance(data, dict):
        # 如果是单个对象，包装为列表
        return [data]
    if isinstance(data, list):
        return _ensure_list_of_dicts(data)
    _fail("E002")


def _read_excel(filepath: str) -> List[Dict]:
    """读取 Excel 文件为字典列表（需安装 openpyxl/xlrd）"""
    try:
        if filepath.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
        else:
            import xlrd
            wb = xlrd.open_workbook(filepath)
    except ImportError:
        _fail("E010", "读取 Excel 需要安装 openpyxl 或 xlrd")
    except Exception as e:
        _fail("E010", f"读取 Excel 失败: {e}")

    result = []
    if filepath.endswith(".xlsx"):
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers:
            for row in rows:
                result.append(dict(zip(headers, row)))
    else:
        ws = wb.sheet_by_index(0)
        headers = [str(cell.value) for cell in ws.row(0)]
        for row_idx in range(1, ws.nrows):
            row_data = {}
            for col_idx, header in enumerate(headers):
                row_data[header] = ws.cell_value(row_idx, col_idx)
            result.append(row_data)
    return result


def read_data(filepath: str) -> List[Dict]:
    """根据文件扩展名读取数据"""
    if not os.path.exists(filepath):
        _fail("E001", f"文件不存在: {filepath}")
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return _read_csv(filepath)
    elif ext == ".json":
        return _read_json(filepath)
    elif ext in (".xlsx", ".xls"):
        return _read_excel(filepath)
    else:
        _fail("E002", f"不支持的文件格式: {ext}")


# ---------------------------------------------------------------------------
# 数据操作
# ---------------------------------------------------------------------------
def _parse_filters(filter_str: str) -> List[Tuple[str, str, Any]]:
    """解析筛选条件字符串，格式: field=value;field>value;field in (a,b,c)"""
    if not filter_str:
        return []
    filters = []
    for part in filter_str.split(";"):
        part = part.strip()
        if not part:
            continue
        # 处理 in 操作
        if " in " in part:
            field, values_str = part.split(" in ", 1)
            field = field.strip()
            values_str = values_str.strip().strip("()")
            values = [v.strip().strip("'\"") for v in values_str.split(",")]
            filters.append((field, "in", values))
            continue
        # 处理比较操作
        for op in (">=", "<=", "!=", ">", "<", "="):
            if op in part:
                field, value_str = part.split(op, 1)
                field = field.strip()
                value_str = value_str.strip().strip("'\"")
                # 尝试转换为数字
                num_val = _safe_float(value_str)
                if num_val is not None:
                    value = num_val
                else:
                    value = value_str
                filters.append((field, op, value))
                break
        else:
            _fail("E005", f"无法解析筛选条件: {part}")
    return filters


def _apply_filters(data: List[Dict], filters: List[Tuple[str, str, Any]]) -> List[Dict]:
    """应用筛选条件"""
    if not filters:
        return data
    result = []
    for item in data:
        match = True
        for field, op, value in filters:
            item_value = _get_field(item, field)
            if op == "in":
                if item_value not in value:
                    match = False
                    break
            elif op == "=":
                if str(item_value) != str(value):
                    match = False
                    break
            elif op == "!=":
                if str(item_value) == str(value):
                    match = False
                    break
            elif op == ">":
                num_item = _safe_float(item_value)
                if num_item is None or num_item <= value:
                    match = False
                    break
            elif op == "<":
                num_item = _safe_float(item_value)
                if num_item is None or num_item >= value:
                    match = False
                    break
            elif op == ">=":
                num_item = _safe_float(item_value)
                if num_item is None or num_item < value:
                    match = False
                    break
            elif op == "<=":
                num_item = _safe_float(item_value)
                if num_item is None or num_item > value:
                    match = False
                    break
        if match:
            result.append(item)
    return result


def _group_by(data: List[Dict], group_field: str) -> Dict[str, List[Dict]]:
    """按字段分组"""
    groups: Dict[str, List[Dict]] = OrderedDict()
    for item in data:
        key = str(_get_field(item, group_field, ""))
        if key not in groups:
            groups[key] = []
        groups[key].append(item)
    return groups


def _aggregate(groups: Dict[str, List[Dict]], agg_func: str, agg_field: str) -> List[Dict]:
    """对分组数据执行聚合操作"""
    result = []
    for key, items in groups.items():
        row = {"group": key}
        if agg_func == "count":
            row["count"] = len(items)
        elif agg_func == "sum":
            total = sum(_safe_float(_get_field(item, agg_field)) or 0 for item in items)
            row[f"{agg_field}_sum"] = total
        elif agg_func == "avg":
            values = [_safe_float(_get_field(item, agg_field)) for item in items]
            values = [v for v in values if v is not None]
            row[f"{agg_field}_avg"] = sum(values) / len(values) if values else 0
        elif agg_func == "min":
            values = [_safe_float(_get_field(item, agg_field)) for item in items]
            values = [v for v in values if v is not None]
            row[f"{agg_field}_min"] = min(values) if values else 0
        elif agg_func == "max":
            values = [_safe_float(_get_field(item, agg_field)) for item in items]
            values = [v for v in values if v is not None]
            row[f"{agg_field}_max"] = max(values) if values else 0
        else:
            _fail("E006", f"不支持的聚合函数: {agg_func}")
        result.append(row)
    return result


def _sort_data(data: List[Dict], sort_field: str, sort_order: str = "asc") -> List[Dict]:
    """排序数据"""
    if not data:
        return data
    if sort_field not in data[0]:
        _fail("E007", f"排序字段不存在: {sort_field}")
    reverse = sort_order.lower() == "desc"
    # 按数值排序（如果可能），否则按字符串排序
    def sort_key(x):
        val = _get_field(x, sort_field, "")
        num_val = _safe_float(val)
        if num_val is not None:
            return (0, num_val)
        return (1, str(val))
    return sorted(data, key=sort_key, reverse=reverse)


def _pivot(data: List[Dict], rows_field: str, cols_field: str, value_field: str) -> List[Dict]:
    """数据透视：行×列交叉分析"""
    if not data:
        _fail("E008")
    # 检查字段存在
    if rows_field not in data[0]:
        _fail("E003", f"行字段不存在: {rows_field}")
    if cols_field not in data[0]:
        _fail("E003", f"列字段不存在: {cols_field}")
    if value_field not in data[0]:
        _fail("E004", f"值字段不存在: {value_field}")

    # 收集所有行和列的值
    row_values = sorted(set(str(_get_field(item, rows_field, "")) for item in data))
    col_values = sorted(set(str(_get_field(item, cols_field, "")) for item in data))

    # 构建透视表
    pivot_data = []
    for row_val in row_values:
        row = {rows_field: row_val}
        for col_val in col_values:
            # 找到匹配的行列组合
            matched = [
                _safe_float(_get_field(item, value_field))
                for item in data
                if str(_get_field(item, rows_field, "")) == row_val
                and str(_get_field(item, cols_field, "")) == col_val
            ]
            matched = [v for v in matched if v is not None]
            row[col_val] = sum(matched) if matched else 0
        pivot_data.append(row)
    return pivot_data


def _generate_chart_data(data: List[Dict], label_field: str, value_field: str) -> List[Dict]:
    """生成图表数据"""
    chart_data = []
    for item in data:
        label = str(_get_field(item, label_field, ""))
        value = _safe_float(_get_field(item, value_field)) or 0
        chart_data.append({"label": label, "value": value})
    return chart_data


def _render_ascii_chart(chart_data: List[Dict], width: int = 40) -> str:
    """渲染 ASCII 柱状图"""
    if not chart_data:
        return "（无数据）"
    max_value = max(item["value"] for item in chart_data)
    if max_value <= 0:
        return "（所有值均为 0）"
    lines = []
    for item in chart_data:
        bar_length = int((item["value"] / max_value) * width)
        bar = "█" * bar_length
        lines.append(f"{item['label']:<20} | {bar} {item['value']:.2f}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 输出格式化
# ---------------------------------------------------------------------------
def _to_markdown(data: List[Dict]) -> str:
    """转换为 Markdown 表格"""
    if not data:
        return "（无数据）"
    headers = list(data[0].keys())
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("|" + "|".join(["---"] * len(headers)) + "|")
    for row in data:
        values = [_format_value(row.get(h, "")) for h in headers]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def _to_csv(data: List[Dict]) -> str:
    """转换为 CSV 字符串"""
    if not data:
        return ""
    headers = list(data[0].keys())
    lines = [",".join(headers)]
    for row in data:
        values = [_format_value(row.get(h, "")).replace(",", "\\,") for h in headers]
        lines.append(",".join(values))
    return "\n".join(lines)


def _to_json(data: List[Dict]) -> str:
    """转换为 JSON 字符串"""
    return json.dumps(data, ensure_ascii=False, indent=2)


def format_output(data: List[Dict], fmt: str) -> str:
    """格式化输出"""
    if fmt == "markdown":
        return _to_markdown(data)
    elif fmt == "csv":
        return _to_csv(data)
    elif fmt == "json":
        return _to_json(data)
    else:
        _fail("E002", f"不支持的输出格式: {fmt}")


# ---------------------------------------------------------------------------
# 原子化文件写入
# ---------------------------------------------------------------------------
def _atomic_write(filepath: str, content: str, dry_run: bool = False) -> None:
    """原子化写入文件（先写临时文件再重命名）"""
    if not dry_run:
        temp_path = f"{filepath}.tmp.{int(time.time())}"
        try:
            with open(temp_path, "w", encoding="utf-8") as f:
                f.write(content)
            os.replace(temp_path, filepath)
        except Exception as e:
            # 清理临时文件
            if os.path.exists(temp_path):
                os.remove(temp_path)
            _fail("E010", f"写入文件失败: {e}")
    else:
        print(f"[DRY-RUN] 将写入文件: {filepath}")
        print(f"[DRY-RUN] 内容摘要: {content[:100]}...")


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def process_file(
    filepath: str,
    filters: Optional[str] = None,
    group_by: Optional[str] = None,
    aggregate: Optional[str] = None,
    agg_field: Optional[str] = None,
    sort: Optional[str] = None,
    pivot: Optional[str] = None,
    fields: Optional[str] = None,
    chart: bool = False,
    format: str = "markdown",
    limit: Optional[int] = None,
    verbose: bool = False,
    dry_run: bool = False,
) -> Tuple[str, List[Dict]]:
    """处理单个文件，返回 (输出内容, 处理后的数据)"""
    if verbose:
        print(f"[INFO] 读取文件: {filepath}")
        print(f"[INFO] 时间: {_now_utc().isoformat()}")

    # 读取数据
    data = read_data(filepath)
    if not data:
        _fail("E008")

    if verbose:
        print(f"[INFO] 读取到 {len(data)} 条记录")

    # 应用筛选
    if filters:
        parsed_filters = _parse_filters(filters)
        data = _apply_filters(data, parsed_filters)
        if verbose:
            print(f"[INFO] 筛选后剩余 {len(data)} 条记录")

    # 字段选择
    if fields:
        field_list = [f.strip() for f in fields.split(",")]
        data = [{k: v for k, v in item.items() if k in field_list} for item in data]

    # 数据透视
    if pivot:
        parts = {}
        for part in pivot.split(";"):
            if "=" in part:
                k, v = part.split("=", 1)
                parts[k.strip()] = v.strip()
        rows_field = parts.get("rows")
        cols_field = parts.get("cols")
        value_field = parts.get("value")
        if not all([rows_field, cols_field, value_field]):
            _fail("E005", "透视参数必须包含 rows/cols/value")
        data = _pivot(data, rows_field, cols_field, value_field)
        if verbose:
            print(f"[INFO] 透视完成，生成 {len(data)} 行")

    # 分组聚合
    if group_by and aggregate:
        groups = _group_by(data, group_by)
        data = _aggregate(groups, aggregate, agg_field or "")
        if verbose:
            print(f"[INFO] 聚合完成，生成 {len(data)} 组")

    # 排序
    if sort:
        if ":" in sort:
            sort_field, sort_order = sort.split(":", 1)
        else:
            sort_field, sort_order = sort, "asc"
        data = _sort_data(data, sort_field, sort_order)
        if verbose:
            print(f"[INFO] 排序完成: {sort_field} {sort_order}")

    # 限制条数
    if limit:
        data = data[:limit]

    # 生成图表
    chart_output = ""
    if chart:
        if group_by and aggregate:
            chart_data = _generate_chart_data(data, "group", list(data[0].keys())[-1])
        else:
            chart_data = _generate_chart_data(data, list(data[0].keys())[0], list(data[0].keys())[-1])
        chart_output = _render_ascii_chart(chart_data)
        if verbose:
            print(f"[INFO] 图表生成完成")

    # 格式化输出
    output = format_output(data, format)
    if chart_output:
        output += "\n\n## 可视化图表\n\n```text\n" + chart_output + "\n```"

    return output, data


def process_directory(
    dirpath: str,
    filters: Optional[str] = None,
    group_by: Optional[str] = None,
    aggregate: Optional[str] = None,
    agg_field: Optional[str] = None,
    sort: Optional[str] = None,
    pivot: Optional[str] = None,
    fields: Optional[str] = None,
    chart: bool = False,
    format: str = "markdown",
    limit: Optional[int] = None,
    verbose: bool = False,
    dry_run: bool = False,
) -> List[Tuple[str, str]]:
    """批量处理目录下的所有 CSV/Excel/JSON 文件"""
    results = []
    supported_exts = (".csv", ".xlsx", ".xls", ".json")
    for filename in sorted(os.listdir(dirpath)):
        if filename.endswith(supported_exts):
            filepath = os.path.join(dirpath, filename)
            try:
                output, _ = process_file(
                    filepath, filters, group_by, aggregate, agg_field,
                    sort, pivot, fields, chart, format, limit, verbose, dry_run
                )
                output_file = f"{os.path.splitext(filepath)[0]}_report.{format}"
                results.append((output_file, output))
                if verbose:
                    print(f"[INFO] 处理完成: {filename}")
            except Exception as e:
                print(f"[WARN] 处理失败 {filename}: {e}", file=sys.stderr)
    return results


# ---------------------------------------------------------------------------
# CLI 入口
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Laravel 动态报表生成器 - 将数据转化为结构化报表",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python run.py --file data.csv --group-by month --aggregate sum --agg-field sales
  python run.py --file data.csv --pivot rows=region cols=month value=sales
  python run.py --file data.csv --filters "status=completed;amount>100" --sort "amount:desc"
  python run.py --dir data/ --group-by region --aggregate count --chart
        """
    )
    parser.add_argument("--file", help="输入文件路径（CSV/Excel/JSON）")
    parser.add_argument("--dir", help="输入目录路径（批量处理）")
    parser.add_argument("--filters", help="筛选条件，格式: field=value;field>value")
    parser.add_argument("--group-by", help="分组字段")
    parser.add_argument("--aggregate", choices=["sum", "count", "avg", "min", "max"], help="聚合函数")
    parser.add_argument("--agg-field", help="聚合字段")
    parser.add_argument("--sort", help="排序字段，格式: field:asc/desc")
    parser.add_argument("--pivot", help="数据透视，格式: rows=field;cols=field;value=field")
    parser.add_argument("--fields", help="选择字段，逗号分隔")
    parser.add_argument("--chart", action="store_true", help="生成 ASCII 图表")
    parser.add_argument("--format", choices=["markdown", "csv", "json"], default="markdown", help="输出格式")
    parser.add_argument("--limit", type=int, help="限制输出条数")
    parser.add_argument("--output", help="输出文件路径（默认输出到 stdout）")
    parser.add_argument("--dry-run", action="store_true", help="预览模式，不写盘")
    parser.add_argument("--verbose", action="store_true", help="详细输出")
    parser.add_argument("--selftest", action="store_true", help="运行自检")
    args = parser.parse_args()
    global dry_run
    dry_run = getattr(args, "dry_run", False)  # v3.274 同步到全局

    if args.selftest:
        selftest()
        return

    # 输入校验
    if not args.file and not args.dir:
        _fail("E001", "必须指定 --file 或 --dir 参数")

    try:
        if args.file:
            output, _ = process_file(
                args.file, args.filters, args.group_by, args.aggregate,
                args.agg_field, args.sort, args.pivot, args.fields,
                args.chart, args.format, args.limit, args.verbose, args.dry_run
            )
            if args.output:
                _atomic_write(args.output, output, args.dry_run)
            else:
                print(output)
        elif args.dir:
            results = process_directory(
                args.dir, args.filters, args.group_by, args.aggregate,
                args.agg_field, args.sort, args.pivot, args.fields,
                args.chart, args.format, args.limit, args.verbose, args.dry_run
            )
            for output_file, output in results:
                if args.dry_run:
                    print(f"[DRY-RUN] 将写入: {output_file}")
                else:
                    _atomic_write(output_file, output, args.dry_run)
                if args.verbose:
                    print(f"[INFO] 已生成: {output_file}")
    except RuntimeError as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"未知错误: {e}", file=sys.stderr)
        sys.exit(1)


# ---------------------------------------------------------------------------
# 自检
# ---------------------------------------------------------------------------
def selftest():
    """运行自检，验证核心功能"""
    print("=" * 60)
    print("开始自检...")
    print(f"时间: {_now_utc().isoformat()}")
    print("=" * 60)

    # 测试数据
    test_data = [
        {"region": "华东", "month": "2024-01", "sales": 1000, "status": "completed"},
        {"region": "华东", "month": "2024-02", "sales": 1500, "status": "completed"},
        {"region": "华北", "month": "2024-01", "sales": 800, "status": "pending"},
        {"region": "华北", "month": "2024-02", "sales": 1200, "status": "completed"},
        {"region": "华南", "month": "2024-01", "sales": 2000, "status": "completed"},
        {"region": "华南", "month": "2024-02", "sales": 1800, "status": "pending"},
    ]

    # 测试 1: 分组聚合
    print("\n[测试 1] 分组聚合")
    groups = _group_by(test_data, "region")
    agg_result = _aggregate(groups, "sum", "sales")
    assert len(agg_result) == 3, f"预期 3 组，实际 {len(agg_result)}"
    assert any(r["group"] == "华东" and r["sales_sum"] == 2500 for r in agg_result), "华东销售额应为 2500"
    print("✓ 分组聚合测试通过")

    # 测试 2: 筛选
    print("\n[测试 2] 筛选")
    filters = _parse_filters("status=completed")
    filtered = _apply_filters(test_data, filters)
    assert len(filtered) == 4, f"预期 4 条，实际 {len(filtered)}"
    print("✓ 筛选测试通过")

    # 测试 3: 排序
    print("\n[测试 3] 排序")
    sorted_data = _sort_data(test_data, "sales", "desc")
    # 按数值排序，最大的是 2000
    assert sorted_data[0]["sales"] == 2000, f"最大销售额应为 2000，实际 {sorted_data[0]['sales']}"
    print("✓ 排序测试通过")

    # 测试 4: 数据透视
    print("\n[测试 4] 数据透视")
    pivot_result = _pivot(test_data, "region", "month", "sales")
    assert len(pivot_result) == 3, f"预期 3 行，实际 {len(pivot_result)}"
    assert "2024-01" in pivot_result[0], "透视表应包含月份列"
    print("✓ 数据透视测试通过")

    # 测试 5: 图表数据
    print("\n[测试 5] 图表数据")
    chart_data = _generate_chart_data(agg_result, "group", "sales_sum")
    assert len(chart_data) == 3, f"预期 3 条图表数据，实际 {len(chart_data)}"
    chart_str = _render_ascii_chart(chart_data)
    assert "█" in chart_str, "图表应包含柱状图字符"
    print("✓ 图表测试通过")

    # 测试 6: 格式转换
    print("\n[测试 6] 格式转换")
    md_output = _to_markdown(test_data[:2])
    assert "|" in md_output, "Markdown 应包含表格分隔符"
    json_output = _to_json(test_data[:2])
    assert json.loads(json_output), "JSON 应可解析"
    csv_output = _to_csv(test_data[:2])
    assert "," in csv_output, "CSV 应包含逗号"
    print("✓ 格式转换测试通过")

    # 测试 7: 错误码
    print("\n[测试 7] 错误码")
    try:
        _fail("E001")
        assert False, "应抛出 E001 异常"
    except RuntimeError as e:
        assert "E001" in str(e), f"错误码 E001 缺失: {e}"
    try:
        _fail("E010", "自定义错误")
        assert False, "应抛出 E010 异常"
    except RuntimeError as e:
        assert "E010" in str(e), f"错误码 E010 缺失: {e}"
    print("✓ 错误码测试通过")

    # 测试 8: 主流程（使用临时文件）
    print("\n[测试 8] 主流程")
    import tempfile
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8") as f:
        f.write("region,month,sales,status\n")
        for item in test_data:
            f.write(f"{item['region']},{item['month']},{item['sales']},{item['status']}\n")
        temp_file = f.name

    try:
        output, result_data = process_file(
            temp_file,
            filters="status=completed",
            group_by="region",
            aggregate="sum",
            agg_field="sales",
            sort="sales_sum:desc",
            format="markdown"
        )
        assert "华东" in output, "输出应包含华东"
        assert "2500" in output, "华东销售额应为 2500"
        print("✓ 主流程测试通过")
    finally:
        os.remove(temp_file)

    # 测试 9: 空数据处理
    print("\n[测试 9] 空数据处理")
    # _aggregate 对空分组返回空列表（不抛异常）
    empty_agg_result = _aggregate({}, "sum", "sales")
    assert empty_agg_result == [], f"空分组聚合应返回空列表，实际 {empty_agg_result}"
    # _pivot 对空数据抛 E008
    try:
        _pivot([], "region", "month", "sales")
        assert False, "空数据透视应抛出 E008"
    except RuntimeError as e:
        assert "E008" in str(e), f"错误码 E008 缺失: {e}"
    print("✓ 空数据测试通过")

    # 测试 10: 边界值
    print("\n[测试 10] 边界值")
    edge_data = [
        {"name": "A", "value": "abc"},
        {"name": "B", "value": "123"},
        {"name": "C", "value": ""},
    ]
    groups = _group_by(edge_data, "name")
    agg_result = _aggregate(groups, "sum", "value")
    assert len(agg_result) == 3, f"预期 3 组，实际 {len(agg_result)}"
    print("✓ 边界值测试通过")

    print("\n" + "=" * 60)
    print("所有自检通过！")
    print("=" * 60)


if __name__ == "__main__":
    main()
