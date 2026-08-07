#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CanvasXpress 数据分析与可视化 Skill - 独立实现脚本

功能概述：
    1. 数据文件解析（CSV / TSV / JSON / Excel）
    2. URL 数据抓取与解析
    3. 图表类型推荐（基于数据维度与字段类型）
    4. 审计追踪生成（记录数据加载、转换、绘图步骤）
    5. 批量图表输出（多组数据生成多个 HTML 文件或合并报告）

本脚本为 clean-room 实现，仅依据功能规格独立编写。
依赖：标准库 + openpyxl（仅 Excel 解析需要，可选）
    # pip install openpyxl

用法示例：
    python scripts/main.py --selftest                     # 离线自检
    python scripts/main.py --file data.csv --chart auto   # 解析文件并推荐图表
    python scripts/main.py --url https://... --chart bar  # 抓取 URL 数据并绘图
    python scripts/main.py --batch data1.csv data2.csv    # 批量输出图表

错误码：
    E001: 参数错误或缺少必要参数
    E002: 文件不存在或无法读取
    E003: 文件格式不支持或解析失败
    E004: URL 访问失败或数据格式错误
    E005: 图表类型不支持
    E006: 数据为空或缺少必要字段
    E007: 输出目录无法创建或写入失败
    E008: Excel 解析需要 openpyxl 库
    E009: 内部逻辑错误（不应发生）
    E010: 自检失败
"""

import argparse
import csv
import json
import os
import sys
import tempfile
import urllib.request
import urllib.error
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

# 尝试导入 openpyxl（仅 Excel 解析需要）
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# 审计追踪模块
# ============================================================

class AuditLogger:
    """审计追踪记录器：记录每次数据操作的步骤与时间戳"""

    def __init__(self):
        self._entries = []

    def log(self, action: str, detail: str = "", level: str = "INFO"):
        """记录一条审计日志"""
        entry = OrderedDict([
            ("timestamp", datetime.now().isoformat(timespec="seconds")),
            ("level", level),
            ("action", action),
            ("detail", detail),
        ])
        self._entries.append(entry)
        return entry

    def get_entries(self):
        """返回所有审计记录（列表副本）"""
        return list(self._entries)

    def to_text(self) -> str:
        """将审计记录格式化为纯文本"""
        lines = ["=== CanvasXpress 审计追踪 ==="]
        for i, e in enumerate(self._entries, 1):
            lines.append(
                f"[{i:03d}] {e['timestamp']} | {e['level']:5s} | "
                f"{e['action']} | {e['detail']}"
            )
        return "\n".join(lines)

    def to_json(self) -> str:
        """将审计记录序列化为 JSON 字符串"""
        return json.dumps(self._entries, ensure_ascii=False, indent=2)


# ============================================================
# 数据解析模块
# ============================================================

class DataParser:
    """数据文件解析器：支持 CSV / TSV / JSON / Excel"""

    @staticmethod
    def _detect_delimiter(first_line: str) -> str:
        """根据首行内容猜测分隔符（逗号或制表符）"""
        if "\t" in first_line and "," not in first_line:
            return "\t"
        return ","

    @staticmethod
    def _guess_type(values):
        """尝试将字符串列表转换为合适的数据类型"""
        # 尝试转浮点数
        try:
            floats = [float(v) for v in values if v != ""]
            if len(floats) == len(values):
                return "numeric"
        except (ValueError, TypeError):
            pass
        # 尝试转整数
        try:
            ints = [int(v) for v in values if v != ""]
            if len(ints) == len(values):
                return "integer"
        except (ValueError, TypeError):
            pass
        # 否则视为字符串
        return "string"

    @staticmethod
    def _convert_values(values):
        """将字符串列表转换为合适类型的 Python 对象"""
        converted = []
        for v in values:
            v = v.strip()
            if v == "":
                converted.append(None)
                continue
            # 尝试数字
            try:
                converted.append(int(v))
                continue
            except ValueError:
                pass
            try:
                converted.append(float(v))
                continue
            except ValueError:
                pass
            # 尝试布尔
            if v.lower() in ("true", "false"):
                converted.append(v.lower() == "true")
                continue
            # 保持字符串
            converted.append(v)
        return converted

    def parse_file(self, filepath: str, logger: AuditLogger = None):
        """解析数据文件，返回 OrderedDict 格式的数据"""
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._parse_csv(path, logger)
        elif suffix == ".tsv":
            return self._parse_csv(path, logger, delimiter="\t")
        elif suffix == ".json":
            return self._parse_json(path, logger)
        elif suffix in (".xlsx", ".xls"):
            return self._parse_excel(path, logger)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

    def _parse_csv(self, path: Path, logger: AuditLogger = None, delimiter=None):
        """解析 CSV/TSV 文件"""
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                content = f.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            with open(path, "r", encoding="gbk", errors="replace") as f:
                content = f.read()

        lines = [l for l in content.splitlines() if l.strip()]
        if not lines:
            raise ValueError("文件内容为空")

        if delimiter is None:
            delimiter = self._detect_delimiter(lines[0])

        # 使用 csv 模块解析
        reader = csv.reader(lines, delimiter=delimiter)
        rows = [row for row in reader if any(cell.strip() for cell in row)]

        if len(rows) < 2:
            raise ValueError("数据至少需要表头和一行数据")

        headers = [h.strip() for h in rows[0]]
        data_rows = rows[1:]

        # 构建列式数据
        result = OrderedDict()
        for col_idx, header in enumerate(headers):
            column_values = []
            for row in data_rows:
                if col_idx < len(row):
                    column_values.append(row[col_idx].strip())
                else:
                    column_values.append("")
            result[header] = self._convert_values(column_values)

        # 记录类型信息
        types = {k: self._guess_type(v) for k, v in result.items()}
        if logger:
            logger.log("数据解析", f"CSV 文件 {path.name}，{len(result)} 列，{len(data_rows)} 行", "INFO")
            logger.log("字段类型识别", json.dumps(types, ensure_ascii=False), "DEBUG")

        return result

    def _parse_json(self, path: Path, logger: AuditLogger = None):
        """解析 JSON 文件"""
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except json.JSONDecodeError as e:
            raise ValueError(f"JSON 解析失败: {e}")

        # 支持两种格式：列式 {col: [values]} 或 行式 [{col: val}, ...]
        if isinstance(data, dict):
            # 列式
            result = OrderedDict()
            for k, v in data.items():
                if isinstance(v, list):
                    result[k] = v
                else:
                    result[k] = [v]
            if logger:
                logger.log("数据解析", f"JSON 文件 {path.name}（列式），{len(result)} 列", "INFO")
            return result
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            # 行式转列式
            result = OrderedDict()
            for key in data[0].keys():
                result[key] = [row.get(key) for row in data]
            if logger:
                logger.log("数据解析", f"JSON 文件 {path.name}（行式转列式），{len(result)} 列", "INFO")
            return result
        else:
            raise ValueError("JSON 数据格式不支持：需为列式对象或行式数组")

    def _parse_excel(self, path: Path, logger: AuditLogger = None):
        """解析 Excel 文件（需要 openpyxl）"""
        if not HAS_OPENPYXL:
            raise ImportError("解析 Excel 需要 openpyxl 库，请执行: pip install openpyxl")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel 文件为空")

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result = OrderedDict()
        for col_idx, header in enumerate(headers):
            values = []
            for row in rows[1:]:
                if col_idx < len(row):
                    values.append(row[col_idx])
                else:
                    values.append(None)
            result[header] = values

        if logger:
            logger.log("数据解析", f"Excel 文件 {path.name}，{len(result)} 列，{len(rows)-1} 行", "INFO")
        return result

    def parse_url(self, url: str, logger: AuditLogger = None):
        """从 URL 获取数据并解析"""
        # 处理 file:// URL（用于测试）
        if url.startswith("file://"):
            filepath = url.replace("file://", "")
            if os.name == "nt" and filepath.startswith("/"):
                filepath = filepath[1:]  # 移除 Windows 路径前的 /
            return self.parse_file(filepath, logger)

        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                content = resp.read().decode("utf-8-sig", errors="replace")
        except urllib.error.URLError as e:
            raise ConnectionError(f"URL 访问失败: {e}")

        # 根据内容猜测格式
        content = content.strip()
        if content.startswith("{") or content.startswith("["):
            # JSON
            try:
                data = json.loads(content)
                if isinstance(data, dict):
                    result = OrderedDict()
                    for k, v in data.items():
                        result[k] = v if isinstance(v, list) else [v]
                elif isinstance(data, list) and data and isinstance(data[0], dict):
                    result = OrderedDict()
                    for key in data[0].keys():
                        result[key] = [row.get(key) for row in data]
                else:
                    raise ValueError("URL JSON 格式不支持")
                if logger:
                    logger.log("URL 数据获取", f"从 {url} 获取 JSON 数据，{len(result)} 列", "INFO")
                return result
            except json.JSONDecodeError:
                raise ValueError("URL 内容 JSON 解析失败")
        else:
            # 尝试 CSV/TSV
            lines = [l for l in content.splitlines() if l.strip()]
            if not lines:
                raise ValueError("URL 内容为空")
            delimiter = self._detect_delimiter(lines[0])
            reader = csv.reader(lines, delimiter=delimiter)
            rows = [row for row in reader if any(cell.strip() for cell in row)]
            if len(rows) < 2:
                raise ValueError("URL 数据至少需要表头和一行数据")
            headers = [h.strip() for h in rows[0]]
            result = OrderedDict()
            for col_idx, header in enumerate(headers):
                values = []
                for row in rows[1:]:
                    values.append(row[col_idx].strip() if col_idx < len(row) else "")
                result[header] = self._convert_values(values)
            if logger:
                logger.log("URL 数据获取", f"从 {url} 获取 CSV 数据，{len(result)} 列", "INFO")
            return result


# ============================================================
# 图表推荐模块
# ============================================================

class ChartRecommender:
    """根据数据特征推荐合适的图表类型"""

    # 支持的图表类型
    SUPPORTED_CHARTS = [
        "scatter", "bar", "heatmap", "boxplot", "line",
        "pie", "histogram", "bubble"
    ]

    def recommend(self, data: OrderedDict, logger: AuditLogger = None):
        """根据数据维度、字段类型、行数推荐图表"""
        n_cols = len(data)
        n_rows = len(next(iter(data.values()), []))

        # 收集字段类型信息
        numeric_cols = []
        string_cols = []
        for name, values in data.items():
            # 检查是否为数值型
            numeric_count = sum(1 for v in values if isinstance(v, (int, float)))
            if numeric_count >= max(1, len(values) * 0.7):
                numeric_cols.append(name)
            else:
                string_cols.append(name)

        recommendations = []

        # 规则1: 单一数值列 -> 柱状图或直方图
        if n_cols == 1 and numeric_cols:
            recommendations.append("histogram")
            recommendations.append("bar")

        # 规则2: 一个字符串列 + 一个数值列 -> 柱状图
        if len(string_cols) >= 1 and len(numeric_cols) >= 1 and n_cols >= 2:
            recommendations.append("bar")

        # 规则3: 两个数值列 -> 散点图
        if len(numeric_cols) >= 2 and n_cols >= 2:
            recommendations.append("scatter")

        # 规则4: 多个数值列（>=3） -> 热力图
        if len(numeric_cols) >= 3:
            recommendations.append("heatmap")

        # 规则5: 有分组字段（字符串列）且数值列>=2 -> 箱线图
        if len(string_cols) >= 1 and len(numeric_cols) >= 2:
            recommendations.append("boxplot")

        # 规则6: 时间序列（首列为日期或有序数值） -> 折线图
        first_col = next(iter(data.keys()), "")
        first_vals = next(iter(data.values()), [])
        if first_vals and all(isinstance(v, (int, float)) for v in first_vals):
            # 检查是否像时间序列（数值递增）
            if len(first_vals) >= 3:
                diffs = [b - a for a, b in zip(first_vals[:-1], first_vals[1:])]
                if all(d >= 0 for d in diffs):
                    recommendations.append("line")

        # 规则7: 单个分类列 -> 饼图
        if n_cols == 1 and string_cols and n_rows <= 10:
            recommendations.append("pie")

        # 去重并保持顺序
        seen = set()
        result = []
        for r in recommendations:
            if r not in seen and r in self.SUPPORTED_CHARTS:
                seen.add(r)
                result.append(r)

        if not result:
            result = ["bar"]  # 默认推荐

        if logger:
            logger.log("图表推荐", f"推荐图表: {result}（数据 {n_cols} 列 × {n_rows} 行）", "INFO")

        return result[0], result  # 返回首选和全部推荐


# ============================================================
# 图表生成模块
# ============================================================

class ChartGenerator:
    """生成交互式 HTML 图表（基于 CanvasXpress 风格的静态 HTML）"""

    @staticmethod
    def _safe_json(data):
        """将数据安全地转为 JSON 字符串"""
        return json.dumps(data, ensure_ascii=False, default=str)

    def generate_html(self, data: OrderedDict, chart_type: str, title: str = "CanvasXpress 图表"):
        """生成单个图表的 HTML 内容"""
        if chart_type not in ChartRecommender.SUPPORTED_CHARTS:
            raise ValueError(f"不支持的图表类型: {chart_type}")

        # 准备数据（转置为 CanvasXpress 格式）
        columns = list(data.keys())
        rows = []
        n_rows = max(len(v) for v in data.values()) if data else 0
        for i in range(n_rows):
            row = []
            for col in columns:
                vals = data[col]
                row.append(vals[i] if i < len(vals) else None)
            rows.append(row)

        # 构建 HTML
        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>{title}</title>
<script src="https://cdn.canvasxpress.org/js/canvasXpress.min.js"></script>
<style>
  body {{ font-family: 'Microsoft YaHei', sans-serif; margin: 20px; }}
  .chart-container {{ width: 100%; max-width: 900px; margin: 0 auto; }}
  h1 {{ text-align: center; color: #333; }}
  .audit-info {{ margin-top: 20px; padding: 10px; background: #f5f5f5; border-radius: 5px; }}
</style>
</head>
<body>
<h1>{title}</h1>
<div class="chart-container" id="chart"></div>
<div class="audit-info" id="audit"></div>
<script>
  var data = {self._safe_json(rows)};
  var columns = {self._safe_json(columns)};
  var chartType = "{chart_type}";

  // 构建 CanvasXpress 数据对象
  var cxData = {{ y: {{ vars: columns, data: data }} }};
  var cxConfig = {{
    graphType: chartType,
    title: "{title}",
    xAxis: columns[0] || "",
    yAxis: columns[1] || "",
    theme: "CanvasXpress",
    width: 800,
    height: 500
  }};

  // 创建图表实例
  var chart = new CanvasXpress("chart", cxData, cxConfig);

  // 记录审计信息
  document.getElementById("audit").innerHTML =
    "<b>审计追踪:</b> 图表类型=" + chartType +
    ", 数据列=" + columns.length +
    ", 数据行=" + data.length +
    ", 生成时间=" + new Date().toLocaleString();
</script>
</body>
</html>"""
        return html

    def generate_batch(self, datasets: list, chart_types: list, output_dir: str,
                       logger: AuditLogger = None):
        """批量生成图表，输出为多个 HTML 文件"""
        out_path = Path(output_dir)
        out_path.mkdir(parents=True, exist_ok=True)

        generated_files = []
        for idx, (data, chart_type) in enumerate(zip(datasets, chart_types)):
            title = f"图表_{idx + 1}"
            html = self.generate_html(data, chart_type, title)
            filepath = out_path / f"chart_{idx + 1}_{chart_type}.html"
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(html)
            generated_files.append(str(filepath))
            if logger:
                logger.log("图表生成", f"已生成 {filepath}", "INFO")

        # 生成合并报告
        report_path = out_path / "report.html"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(self._generate_report(generated_files, logger))
        generated_files.append(str(report_path))

        if logger:
            logger.log("批量输出", f"共生成 {len(generated_files)} 个文件到 {output_dir}", "INFO")

        return generated_files

    def _generate_report(self, chart_files: list, logger: AuditLogger = None):
        """生成合并报告 HTML"""
        items = "".join(
            f'<li><a href="{Path(f).name}" target="_blank">{Path(f).name}</a></li>'
            for f in chart_files
        )
        audit_text = ""
        if logger:
            audit_text = logger.to_text().replace("\n", "<br>")
        return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>CanvasXpress 批量报告</title></head>
<body>
<h1>CanvasXpress 图表报告</h1>
<p>生成时间: {datetime.now().isoformat(timespec="seconds")}</p>
<h2>图表列表</h2>
<ul>{items}</ul>
<h2>审计追踪</h2>
<pre>{audit_text}</pre>
</body>
</html>"""


# ============================================================
# 主程序模块
# ============================================================

class CanvasXpressSkill:
    """CanvasXpress Skill 主类，整合各模块"""

    def __init__(self):
        self.logger = AuditLogger()
        self.parser = DataParser()
        self.recommender = ChartRecommender()
        self.generator = ChartGenerator()

    def process_file(self, filepath: str, chart_type: str = None):
        """处理本地数据文件"""
        self.logger.log("启动", f"处理文件: {filepath}", "INFO")
        data = self.parser.parse_file(filepath, self.logger)

        if chart_type is None or chart_type == "auto":
            chart_type, all_recs = self.recommender.recommend(data, self.logger)
            self.logger.log("图表选择", f"自动推荐: {chart_type}（全部: {all_recs}）", "INFO")
        else:
            if chart_type not in ChartRecommender.SUPPORTED_CHARTS:
                raise ValueError(f"不支持的图表类型: {chart_type}")

        html = self.generator.generate_html(data, chart_type)
        self.logger.log("完成", f"图表生成成功，类型={chart_type}", "INFO")
        return html, data, self.logger

    def process_url(self, url: str, chart_type: str = None):
        """处理 URL 数据"""
        self.logger.log("启动", f"处理 URL: {url}", "INFO")
        data = self.parser.parse_url(url, self.logger)

        if chart_type is None or chart_type == "auto":
            chart_type, all_recs = self.recommender.recommend(data, self.logger)
            self.logger.log("图表选择", f"自动推荐: {chart_type}", "INFO")
        else:
            if chart_type not in ChartRecommender.SUPPORTED_CHARTS:
                raise ValueError(f"不支持的图表类型: {chart_type}")

        html = self.generator.generate_html(data, chart_type)
        self.logger.log("完成", f"URL 图表生成成功，类型={chart_type}", "INFO")
        return html, data, self.logger

    def process_batch(self, files: list, output_dir: str):
        """批量处理多个数据文件"""
        self.logger.log("启动", f"批量处理 {len(files)} 个文件", "INFO")
        datasets = []
        chart_types = []
        for f in files:
            data = self.parser.parse_file(f, self.logger)
            chart_type, _ = self.recommender.recommend(data, self.logger)
            datasets.append(data)
            chart_types.append(chart_type)

        output = self.generator.generate_batch(datasets, chart_types, output_dir, self.logger)
        self.logger.log("完成", f"批量处理完成，输出 {len(output)} 个文件", "INFO")
        return output, self.logger


# ============================================================
# 自检模块
# ============================================================

def run_selftest():
    """离线自检核心逻辑，使用内置硬编码数据，不依赖外部环境"""
    try:
        # 创建临时目录（用于批量输出测试）
        temp_dir = tempfile.mkdtemp(prefix="canvasxpress_selftest_")

        # ---- 测试 1: 数据解析（CSV） ----
        skill = CanvasXpressSkill()
        csv_content = "name,age,score\nAlice,25,85.5\nBob,30,92.0\nCarol,35,78.5\n"
        csv_path = os.path.join(temp_dir, "test_data.csv")
        with open(csv_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        data = skill.parser.parse_file(csv_path, skill.logger)
        assert "name" in data, "CSV 解析失败: 缺少 name 列"
        assert "age" in data, "CSV 解析失败: 缺少 age 列"
        assert len(data["name"]) == 3, "CSV 解析失败: 行数不对"
        assert all(isinstance(v, (int, float)) for v in data["age"]), "CSV 解析失败: age 应转为数值"
        print("[PASS] 数据解析 (CSV)")

        # ---- 测试 2: 图表推荐 ----
        chart_type, all_recs = skill.recommender.recommend(data, skill.logger)
        assert chart_type in ChartRecommender.SUPPORTED_CHARTS, "图表推荐失败: 返回了不支持的图表"
        assert len(all_recs) >= 1, "图表推荐失败: 推荐列表为空"
        print(f"[PASS] 图表推荐 (首选={chart_type}, 全部={all_recs})")

        # ---- 测试 3: HTML 生成 ----
        html = skill.generator.generate_html(data, chart_type, "自检图表")
        assert "<!DOCTYPE html>" in html, "HTML 生成失败: 缺少 DOCTYPE"
        assert "CanvasXpress" in html or "canvasXpress" in html, "HTML 生成失败: 缺少 CanvasXpress"
        assert len(html) > 500, "HTML 生成失败: 内容过短"
        print(f"[PASS] HTML 图表生成 (长度={len(html)} 字符)")

        # ---- 测试 4: JSON 解析 ----
        json_data = {"x": [1, 2, 3, 4], "y": [10, 20, 30, 40], "label": ["a", "b", "c", "d"]}
        json_path = os.path.join(temp_dir, "test_data.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f)

        data2 = skill.parser.parse_file(json_path, skill.logger)
        assert len(data2["x"]) == 4, "JSON 解析失败: x 列长度不对"
        assert data2["label"][0] == "a", "JSON 解析失败: label 值不对"
        print("[PASS] 数据解析 (JSON)")

        # ---- 测试 5: 批量生成 ----
        output_files = skill.process_batch([csv_path, json_path], temp_dir)
        assert len(output_files) >= 2, "批量生成失败: 输出文件不足"
        for f in output_files:
            assert os.path.exists(f), f"批量生成失败: 文件不存在 {f}"
        print(f"[PASS] 批量图表输出 ({len(output_files)} 个文件)")

        # ---- 测试 6: 审计追踪 ----
        audit_text = skill.logger.to_text()
        assert "审计追踪" in audit_text, "审计追踪失败: 缺少标题"
        assert len(skill.logger.get_entries()) >= 5, "审计追踪失败: 记录条数不足"
        print(f"[PASS] 审计追踪 ({len(skill.logger.get_entries())} 条记录)")

        # ---- 测试 7: 空数据错误处理 ----
        try:
            skill.parser.parse_file(os.path.join(temp_dir, "nonexistent.csv"), skill.logger)
            assert False, "应抛出文件不存在错误"
        except FileNotFoundError:
            print("[PASS] 错误处理 (文件不存在)")

        # ---- 测试 8: 不支持的图表类型 ----
        try:
            skill.generator.generate_html(data, "invalid_chart", "测试")
            assert False, "应抛出图表类型错误"
        except ValueError:
            print("[PASS] 错误处理 (不支持的图表类型)")

        # ---- 测试 9: URL 解析（使用本地文件模拟，不访问网络） ----
        # 通过 file:// URL 测试 URL 解析逻辑
        local_url = "file://" + csv_path
        data3 = skill.parser.parse_url(local_url, skill.logger)
        assert len(data3["name"]) == 3, "URL 解析失败"
        print("[PASS] URL 数据获取 (file:// 模拟)")

        # ---- 测试 10: 综合流程 ----
        skill2 = CanvasXpressSkill()
        data4 = skill2.parser.parse_file(csv_path, skill2.logger)
        rec, _ = skill2.recommender.recommend(data4, skill2.logger)
        html2 = skill2.generator.generate_html(data4, rec, "综合测试")
        assert len(html2) > 0, "综合流程失败"
        print("[PASS] 综合流程")

        # 清理临时文件
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)

        print("\n=== 全部自检通过 ===")
        return 0

    except AssertionError as e:
        print(f"\n[FAIL] 自检失败: {e}", file=sys.stderr)
        return 10  # E010
    except Exception as e:
        print(f"\n[FAIL] 自检异常: {type(e).__name__}: {e}", file=sys.stderr)
        return 10  # E010


# ============================================================
# 命令行入口
# ============================================================

def main():
    """命令行主入口"""
    parser = argparse.ArgumentParser(
        description="CanvasXpress 数据分析与可视化 Skill",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s --selftest                     # 运行自检
  %(prog)s --file data.csv                # 解析 CSV 并推荐图表
  %(prog)s --file data.csv --chart bar    # 指定图表类型
  %(prog)s --url https://example.com/data.csv
  %(prog)s --batch f1.csv f2.json --output ./out
        """
    )
    parser.add_argument("--selftest", action="store_true", help="运行离线自检")
    parser.add_argument("--file", type=str, help="输入数据文件路径 (CSV/TSV/JSON/Excel)")
    parser.add_argument("--url", type=str, help="数据 URL 地址")
    parser.add_argument("--chart", type=str, default="auto",
                        help="图表类型: auto/scatter/bar/heatmap/boxplot/line/pie/histogram/bubble")
    parser.add_argument("--batch", nargs="+", help="批量处理多个数据文件")
    parser.add_argument("--output", type=str, default="./output",
                        help="输出目录 (批量模式使用)")

    args = parser.parse_args()

    # 自检模式
    if args.selftest:
        sys.exit(run_selftest())

    # 参数检查
    if not args.file and not args.url and not args.batch:
        print("错误: 必须提供 --file, --url 或 --batch 参数 (E001)", file=sys.stderr)
        print("提示: 使用 --selftest 运行自检", file=sys.stderr)
        sys.exit(1)

    try:
        skill = CanvasXpressSkill()

        if args.batch:
            # 批量模式
            output_files, logger = skill.process_batch(args.batch, args.output)
            print(f"批量处理完成，输出文件:")
            for f in output_files:
                print(f"  {f}")
            print(f"\n审计日志:\n{logger.to_text()}")

        elif args.file:
            # 单文件模式
            html, data, logger = skill.process_file(args.file, args.chart)
            # 输出 HTML 到文件
            out_path = Path(args.output)
            out_path.mkdir(parents=True, exist_ok=True)
            html_file = out_path / "chart.html"
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"图表已生成: {html_file}")
            print(f"数据列数: {len(data)}")
            print(f"数据行数: {len(next(iter(data.values()), []))}")
            print(f"\n审计日志:\n{logger.to_text()}")

        elif args.url:
            # URL 模式
            html, data, logger = skill.process_url(args.url, args.chart)
            out_path = Path(args.output)
            out_path.mkdir(parents=True, exist_ok=True)
            html_file = out_path / "chart.html"
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(html)
            print(f"图表已生成: {html_file}")
            print(f"数据列数: {len(data)}")
            print(f"\n审计日志:\n{logger.to_text()}")

    except FileNotFoundError as e:
        print(f"错误: {e} (E002)", file=sys.stderr)
        sys.exit(2)
    except ValueError as e:
        print(f"错误: {e} (E003)", file=sys.stderr)
        sys.exit(3)
    except ConnectionError as e:
        print(f"错误: {e} (E004)", file=sys.stderr)
        sys.exit(4)
    except ImportError as e:
        print(f"错误: {e} (E008)", file=sys.stderr)
        sys.exit(8)
    except Exception as e:
        print(f"未预期错误: {type(e).__name__}: {e} (E009)", file=sys.stderr)
        sys.exit(9)


if __name__ == "__main__":
    main()
