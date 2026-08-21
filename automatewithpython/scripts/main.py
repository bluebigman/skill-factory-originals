#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
automatewithpython - 办公自动化 Python 脚本生成器

根据用户描述生成可执行的 Python 自动化脚本。
本实现为 clean-room 重写，仅依据功能规格独立开发。
支持 --selftest 离线自检，不依赖外部文件与网络。
"""

import argparse
import csv
import json
import os
import re
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple
dry_run = False  # v3.274 模块级 dry-run 标志


# =============================================================================
# 错误码定义
# =============================================================================
ERROR_CODES = {
    "E001": "参数错误：缺少必要的命令行参数",
    "E002": "输入文件不存在或无法读取",
    "E003": "输出目录不存在或无法写入",
    "E004": "任务描述为空或格式不正确",
    "E005": "不支持的文件类型（仅支持 CSV/Excel/TXT）",
    "E006": "生成的脚本语法检查失败",
    "E007": "自检数据初始化失败",
    "E008": "自检断言失败：核心逻辑异常",
    "E009": "运行时异常：未预期的错误",
    "E010": "环境检查失败：Python 版本过低",
}


# =============================================================================
# 数据模型
# =============================================================================
@dataclass
class TaskSpec:
    """任务规格描述"""
    description: str
    input_files: List[str] = field(default_factory=list)
    output_dir: str = "."
    task_type: str = "unknown"  # rename / excel / csv / log / organize / web


@dataclass
class GeneratedScript:
    """生成的脚本结果"""
    code: str
    dependencies: List[str] = field(default_factory=list)
    usage_notes: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# =============================================================================
# 核心逻辑：任务解析
# =============================================================================
class TaskParser:
    """解析用户任务描述，识别任务类型和关键参数"""

    # 任务类型识别关键词
    TASK_KEYWORDS = {
        "rename": ["重命名", "改名", "rename", "批量改名"],
        "excel": ["excel", "单元格", "公式", "表格", "xlsx", "xls"],
        "csv": ["csv", "清洗", "去重", "筛选", "排序"],
        "log": ["日志", "log", "分析日志", "提取日志"],
        "organize": ["整理", "分类", "归档", "移动文件"],
        "web": ["网页", "抓取", "爬虫", "下载", "http", "url"],
    }

    def parse(self, description: str) -> TaskSpec:
        """解析任务描述"""
        if not description or not description.strip():
            raise ValueError("E004")

        desc_lower = description.lower()
        task_type = "unknown"
        max_score = 0

        for ttype, keywords in self.TASK_KEYWORDS.items():
            score = sum(1 for kw in keywords if kw.lower() in desc_lower)
            if score > max_score:
                max_score = score
                task_type = ttype

        return TaskSpec(
            description=description.strip(),
            task_type=task_type,
        )


# =============================================================================
# 核心逻辑：脚本生成器
# =============================================================================
class ScriptGenerator:
    """根据任务规格生成 Python 脚本"""

    def __init__(self):
        self.parser = TaskParser()

    def generate(self, task: TaskSpec) -> GeneratedScript:
        """生成脚本"""
        if task.task_type == "rename":
            return self._generate_rename_script(task)
        elif task.task_type == "excel":
            return self._generate_excel_script(task)
        elif task.task_type == "csv":
            return self._generate_csv_script(task)
        elif task.task_type == "log":
            return self._generate_log_script(task)
        elif task.task_type == "organize":
            return self._generate_organize_script(task)
        elif task.task_type == "web":
            return self._generate_web_script(task)
        else:
            return self._generate_generic_script(task)

    def _generate_rename_script(self, task: TaskSpec) -> GeneratedScript:
        """生成文件重命名脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""文件批量重命名脚本 (由 automatewithpython 生成)"""

import os
import sys
from pathlib import Path

def rename_files(directory: str, pattern: str = "file_", extension: str = None):
    """批量重命名文件
    
    Args:
        directory: 目标目录
        pattern: 新文件名前缀
        extension: 只处理指定扩展名（如 .txt），None 表示处理所有文件
    """
    target_dir = Path(directory)
    if not target_dir.exists():
        print(f"错误：目录不存在 - {directory}")
        return False
    
    files = list(target_dir.iterdir())
    renamed_count = 0
    
    for idx, file_path in enumerate(files, start=1):
        if file_path.is_file():
            if extension and file_path.suffix != extension:
                continue
            
            new_name = f"{pattern}{idx:03d}{file_path.suffix}"
            new_path = target_dir / new_name
            
            if new_path.exists():
                print(f"跳过：目标已存在 - {new_name}")
                continue
            
            try:
                file_path.rename(new_path)
                print(f"重命名: {file_path.name} -> {new_name}")
                renamed_count += 1
            except Exception as e:
                print(f"失败: {file_path.name} - {e}")
    
    print(f"\\n完成！共重命名 {renamed_count} 个文件")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python rename_script.py <目录> [前缀] [扩展名]")
        sys.exit(1)
    
    directory = sys.argv[1]
    pattern = sys.argv[2] if len(sys.argv) > 2 else "file_"
    extension = sys.argv[3] if len(sys.argv) > 3 else None
    
    if not rename_files(directory, pattern, extension):
        sys.exit(1)
'''
        return GeneratedScript(
            code=code,
            dependencies=[],
            usage_notes=[
                "用法: python rename_script.py <目录> [前缀] [扩展名]",
                "示例: python rename_script.py ./docs report_ .txt",
                "注意: 脚本会按文件顺序自动编号",
            ],
            warnings=["重命名操作不可逆，请提前备份文件"],
        )

    def _generate_excel_script(self, task: TaskSpec) -> GeneratedScript:
        """生成 Excel 处理脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 批量处理脚本 (由 automatewithpython 生成)"""

import sys
from pathlib import Path

try:
    import openpyxl  # pip install openpyxl
except ImportError:
    print("请安装依赖: pip install openpyxl")
    sys.exit(1)

def process_excel(file_path: str, sheet_name: str = None):
    """处理 Excel 文件
    
    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称，None 表示第一个工作表
    """
    path = Path(file_path)
    if not path.exists():
        print(f"错误：文件不存在 - {file_path}")
        return False
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        print(f"处理文件: {path.name}")
        print(f"工作表: {ws.title}")
        print(f"数据范围: {ws.dimensions}")
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
        
        # 示例：打印前 5 行数据
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
            print(row)
        
        # 在这里添加你的处理逻辑
        
        output_path = path.with_stem(path.stem + "_processed")
        wb.save(output_path)
        print(f"\\n已保存到: {output_path}")
        return True
        
    except Exception as e:
        print(f"处理失败: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python excel_script.py <Excel文件> [工作表名]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not process_excel(file_path, sheet_name):
        sys.exit(1)
'''
        return GeneratedScript(
            code=code,
            dependencies=["openpyxl"],
            usage_notes=[
                "用法: python excel_script.py <Excel文件> [工作表名]",
                "需要安装: pip install openpyxl",
            ],
            warnings=["大型 Excel 文件（>50MB）建议分块处理"],
        )

    def _generate_csv_script(self, task: TaskSpec) -> GeneratedScript:
        """生成 CSV 处理脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CSV 数据清洗脚本 (由 automatewithpython 生成)"""

import csv
import sys
from pathlib import Path

def clean_csv(input_file: str, output_file: str = None, delimiter: str = ","):
    """CSV 数据清洗
    
    Args:
        input_file: 输入 CSV 文件
        output_file: 输出文件，None 自动生成
        delimiter: 分隔符
    """
    in_path = Path(input_file)
    if not in_path.exists():
        print(f"错误：文件不存在 - {input_file}")
        return False
    
    out_path = Path(output_file) if output_file else in_path.with_stem(in_path.stem + "_cleaned")
    
    try:
        with open(in_path, "r", encoding="utf-8") as fin:
            reader = csv.DictReader(fin, delimiter=delimiter)
            fieldnames = reader.fieldnames
            
            rows = []
            seen = set()  # 用于去重
            
            for row in reader:
                # 去除空白字符
                cleaned = {k: (v.strip() if v else v) for k, v in row.items()}
                
                # 跳过空行
                if all(v is None or v == "" for v in cleaned.values()):
                    continue
                
                # 基于第一列去重
                key = cleaned.get(fieldnames[0], "")
                if key in seen:
                    continue
                seen.add(key)
                rows.append(cleaned)
        
        with open(out_path, "w", encoding="utf-8", newline="") as fout:
            writer = csv.DictWriter(fout, fieldnames=fieldnames, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(rows)
        
        print(f"清洗完成: {len(rows)} 行数据")
        print(f"输出文件: {out_path}")
        return True
        
    except Exception as e:
        print(f"处理失败: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python csv_script.py <CSV文件> [输出文件] [分隔符]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    delimiter = sys.argv[3] if len(sys.argv) > 3 else ","
    
    if not clean_csv(input_file, output_file, delimiter):
        sys.exit(1)
'''
        return GeneratedScript(
            code=code,
            dependencies=[],
            usage_notes=[
                "用法: python csv_script.py <CSV文件> [输出文件] [分隔符]",
                "功能: 去除空白、去重、跳过空行",
            ],
            warnings=["默认按第一列去重，请确认业务逻辑"],
        )

    def _generate_log_script(self, task: TaskSpec) -> GeneratedScript:
        """生成日志分析脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""日志分析脚本 (由 automatewithpython 生成)"""

import re
import sys
from collections import Counter
from pathlib import Path

def analyze_log(log_file: str, pattern: str = r"ERROR|WARN|INFO"):
    """分析日志文件
    
    Args:
        log_file: 日志文件路径
        pattern: 匹配模式（正则表达式）
    """
    log_path = Path(log_file)
    if not log_path.exists():
        print(f"错误：文件不存在 - {log_file}")
        return False
    
    try:
        with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        
        regex = re.compile(pattern)
        matches = Counter()
        matched_lines = []
        
        for line_num, line in enumerate(lines, start=1):
            found = regex.findall(line)
            if found:
                for match in found:
                    matches[match] += 1
                matched_lines.append((line_num, line.strip()))
        
        print(f"日志文件: {log_path.name}")
        print(f"总行数: {len(lines)}")
        print(f"匹配行数: {len(matched_lines)}")
        print("\\n统计结果:")
        
        for key, count in matches.most_common():
            print(f"  {key}: {count} 次")
        
        print("\\n匹配行预览 (前10条):")
        for line_num, line in matched_lines[:10]:
            print(f"  [{line_num}] {line[:100]}")
        
        return True
        
    except Exception as e:
        print(f"分析失败: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python log_script.py <日志文件> [匹配模式]")
        sys.exit(1)
    
    log_file = sys.argv[1]
    pattern = sys.argv[2] if len(sys.argv) > 2 else r"ERROR|WARN|INFO"
    
    if not analyze_log(log_file, pattern):
        sys.exit(1)
'''
        return GeneratedScript(
            code=code,
            dependencies=[],
            usage_notes=[
                "用法: python log_script.py <日志文件> [匹配模式]",
                "功能: 统计日志级别、提取匹配行",
            ],
            warnings=["大日志文件建议使用分块读取"],
        )

    def _generate_organize_script(self, task: TaskSpec) -> GeneratedScript:
        """生成文件整理脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""文件整理脚本 (由 automatewithpython 生成)"""

import shutil
import sys
from pathlib import Path

# 按扩展名分类的映射
CATEGORY_MAP = {
    "图片": [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg"],
    "文档": [".doc", ".docx", ".pdf", ".txt", ".md", ".xls", ".xlsx", ".ppt", ".pptx"],
    "音频": [".mp3", ".wav", ".flac", ".aac", ".ogg"],
    "视频": [".mp4", ".avi", ".mkv", ".mov", ".wmv"],
    "压缩包": [".zip", ".rar", ".7z", ".tar", ".gz"],
    "代码": [".py", ".js", ".java", ".c", ".cpp", ".html", ".css"],
}

def organize_files(directory: str):
    """按类型整理文件
    
    Args:
        directory: 目标目录
    """
    target_dir = Path(directory)
    if not target_dir.exists():
        print(f"错误：目录不存在 - {directory}")
        return False
    
    # 创建分类子目录
    for category in CATEGORY_MAP:
        (target_dir / category).mkdir(exist_ok=True)
    
    moved_count = 0
    skipped_count = 0
    
    for file_path in target_dir.iterdir():
        if not file_path.is_file():
            continue
        
        # 跳过脚本自身
        if file_path.name == __file__:
            continue
        
        # 查找分类
        category = "其他"
        for cat, extensions in CATEGORY_MAP.items():
            if file_path.suffix.lower() in extensions:
                category = cat
                break
        
        # 移动到对应目录
        dest_dir = target_dir / category
        dest_path = dest_dir / file_path.name
        
        try:
            if dest_path.exists():
                # 重命名避免冲突
                dest_path = dest_dir / f"{file_path.stem}_dup{file_path.suffix}"
            
            shutil.move(str(file_path), str(dest_path))
            print(f"移动: {file_path.name} -> {category}/")
            moved_count += 1
        except Exception as e:
            print(f"失败: {file_path.name} - {e}")
            skipped_count += 1
    
    print(f"\\n整理完成！移动 {moved_count} 个文件，跳过 {skipped_count} 个")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python organize_script.py <目录>")
        sys.exit(1)
    
    if not organize_files(sys.argv[1]):
        sys.exit(1)
'''
        return GeneratedScript(
            code=code,
            dependencies=[],
            usage_notes=[
                "用法: python organize_script.py <目录>",
                "功能: 按扩展名自动分类到子目录",
            ],
            warnings=["同名文件会自动添加 _dup 后缀避免覆盖"],
        )

    def _generate_web_script(self, task: TaskSpec) -> GeneratedScript:
        """生成网页抓取脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""网页数据抓取脚本 (由 automatewithpython 生成)"""

import sys
from urllib.request import urlopen
from urllib.parse import urljoin
from html.parser import HTMLParser

class LinkParser(HTMLParser):
    """提取网页链接的解析器"""
    def __init__(self):
        super().__init__()
        self.links = []
    
    def handle_starttag(self, tag, attrs):
        if tag == "a":
            for attr, value in attrs:
                if attr == "href":
                    self.links.append(value)

def fetch_web_data(url: str, output_file: str = None):
    """抓取网页数据
    
    Args:
        url: 目标 URL
        output_file: 输出文件，None 则打印到控制台
    """
    try:
        # 设置 User-Agent 避免被拒绝
        req = __import__("urllib.request", fromlist=["Request"])
        request = req.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        
        with urlopen(request, timeout=10) as response:
            html = response.read().decode("utf-8", errors="ignore")
        
        # 解析链接
        parser = LinkParser()
        parser.feed(html)
        
        # 转为绝对链接
        absolute_links = [urljoin(url, link) for link in parser.links]
        unique_links = list(dict.fromkeys(absolute_links))  # 去重保序
        
        print(f"抓取 URL: {url}")
        print(f"找到 {len(unique_links)} 个唯一链接")
        
        result = "\\n".join(unique_links)
        
        if output_file:
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(result)
            print(f"已保存到: {output_file}")
        else:
            print("\\n链接列表:")
            for link in unique_links[:20]:  # 只显示前20条
                print(f"  {link}")
        
        return True
        
    except Exception as e:
        print(f"抓取失败: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python web_script.py <URL> [输出文件]")
        sys.exit(1)
    
    url = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not fetch_web_data(url, output_file):
        sys.exit(1)
'''
        return GeneratedScript(
            code=code,
            dependencies=[],
            usage_notes=[
                "用法: python web_script.py <URL> [输出文件]",
                "功能: 抓取网页并提取所有链接",
            ],
            warnings=["请遵守目标网站的 robots.txt 和使用条款"],
        )

    def _generate_generic_script(self, task: TaskSpec) -> GeneratedScript:
        """生成通用脚本"""
        code = '''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""通用自动化脚本 (由 automatewithpython 生成)"""

import sys
from pathlib import Path

def main():
    """主函数"""
    print("这是由 automatewithpython 生成的通用脚本")
    print(f"任务描述: {task_description}")
    return True

if __name__ == "__main__":
    task_description = "未指定具体任务"
    if len(sys.argv) > 1:
        task_description = " ".join(sys.argv[1:])
    sys.exit(0 if main() else 1)
'''
        return GeneratedScript(
            code=code,
            dependencies=[],
            usage_notes=[
                "通用模板，请根据实际需求修改",
            ],
            warnings=["此脚本为模板，需要手动补充业务逻辑"],
        )


# =============================================================================
# 核心逻辑：脚本验证
# =============================================================================
class ScriptValidator:
    """验证生成的脚本语法正确性"""

    def validate(self, script: GeneratedScript) -> bool:
        """检查脚本语法"""
        try:
            compile(script.code, "<generated>", "exec")
            return True
        except SyntaxError:
            return False


# =============================================================================
# 核心逻辑：主控制器
# =============================================================================
class AutomationGenerator:
    """主控制器：协调任务解析、脚本生成和验证"""

    def __init__(self):
        self.parser = TaskParser()
        self.generator = ScriptGenerator()
        self.validator = ScriptValidator()

    def process(self, description: str, input_files: List[str] = None, output_dir: str = ".") -> GeneratedScript:
        """处理任务描述，生成脚本"""
        # 解析任务
        task = self.parser.parse(description)
        task.input_files = input_files or []
        task.output_dir = output_dir

        # 验证输入文件
        for file_path in task.input_files:
            if not Path(file_path).exists():
                raise FileNotFoundError(f"E002: {file_path}")

        # 验证输出目录
        out_dir = Path(output_dir)
        if not out_dir.exists():
            try:
                out_dir.mkdir(parents=True, exist_ok=True)
            except OSError:
                raise PermissionError("E003")

        # 生成脚本
        script = self.generator.generate(task)

        # 验证语法
        if not self.validator.validate(script):
            raise SyntaxError("E006")

        return script

    def save_script(self, script: GeneratedScript, output_dir: str = ".") -> Path:
        """保存生成的脚本到文件"""
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        # 生成文件名
        timestamp = __import__("time").strftime("%Y%m%d_%H%M%S")
        filename = f"automation_{timestamp}.py"
        file_path = out_dir / filename

        # 写入文件
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(script.code)

        return file_path


# =============================================================================
# 自检模块
# =============================================================================
class SelfTest:
    """离线自检功能，使用内置硬编码样例数据"""

    @staticmethod
    def run() -> bool:
        """执行自检"""
        try:
            print("=" * 60)
            print("automatewithpython 自检开始")
            print("=" * 60)

            # 测试 1: 任务解析器
            print("\n[1/5] 测试任务解析器...")
            parser = TaskParser()
            test_cases = [
                ("批量重命名文件夹中的图片文件", "rename"),
                ("处理 Excel 表格中的销售数据", "excel"),
                ("清洗 CSV 文件中的重复数据", "csv"),
                ("分析服务器日志文件", "log"),
                ("整理下载文件夹中的文件", "organize"),
                ("抓取网页上的文章链接", "web"),
            ]
            for desc, expected_type in test_cases:
                task = parser.parse(desc)
                assert task.task_type == expected_type, f"任务类型不匹配: {desc} -> {task.task_type}"
            print("  ✓ 任务类型识别正确")

            # 测试 2: 脚本生成器
            print("[2/5] 测试脚本生成器...")
            generator = ScriptGenerator()
            for desc, _ in test_cases:
                task = parser.parse(desc)
                script = generator.generate(task)
                assert script.code and len(script.code) > 100, f"脚本内容过短: {desc}"
                assert isinstance(script.dependencies, list), "依赖列表类型错误"
                assert isinstance(script.usage_notes, list), "使用说明类型错误"
            print("  ✓ 脚本生成成功")

            # 测试 3: 脚本语法验证
            print("[3/5] 测试脚本语法验证...")
            validator = ScriptValidator()
            for desc, _ in test_cases:
                task = parser.parse(desc)
                script = generator.generate(task)
                assert validator.validate(script), f"语法验证失败: {desc}"
            print("  ✓ 所有脚本语法正确")

            # 测试 4: 完整流程
            print("[4/5] 测试完整处理流程...")
            automation = AutomationGenerator()
            
            # 创建临时目录用于测试
            with tempfile.TemporaryDirectory() as tmpdir:
                # 生成脚本
                script = automation.process(
                    "批量重命名测试文件",
                    output_dir=tmpdir
                )
                assert script.code, "生成的脚本为空"
                
                # 保存脚本
                saved_path = automation.save_script(script, tmpdir)
                assert saved_path.exists(), "脚本文件未保存"
                assert saved_path.suffix == ".py", "脚本扩展名错误"
                
                # 验证文件内容
                with open(saved_path, "r", encoding="utf-8") as f:
                    content = f.read()
                assert len(content) > 100, "保存的脚本内容过短"
            print("  ✓ 完整流程正常")

            # 测试 5: 错误处理
            print("[5/5] 测试错误处理...")
            automation = AutomationGenerator()
            
            # 空描述
            try:
                automation.process("")
                assert False, "空描述应该抛出异常"
            except ValueError as e:
                assert "E004" in str(e), f"错误码不正确: {e}"
            
            # 不存在的文件
            try:
                automation.process("测试任务", input_files=["/nonexistent/file.txt"])
                assert False, "不存在的文件应该抛出异常"
            except FileNotFoundError as e:
                assert "E002" in str(e), f"错误码不正确: {e}"
            
            # 无法写入的目录 - 使用一个文件路径作为目录（确保无法创建）
            try:
                # 创建一个临时文件，然后尝试将其作为输出目录
                with tempfile.NamedTemporaryFile(delete=False) as tf:
                    temp_file_path = tf.name
                
                # 尝试在文件路径下创建目录（应该失败）
                automation.process("测试任务", output_dir=os.path.join(temp_file_path, "subdir"))
                assert False, "无法写入的目录应该抛出异常"
            except (PermissionError, OSError, NotADirectoryError):
                pass  # 预期行为
            finally:
                # 清理临时文件
                try:
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                except Exception as e:
                    print(f"[WARN] 降级处理: {e}", file=sys.stderr)  # R2 降级输出
            
            print("  ✓ 错误处理正常")

            print("\n" + "=" * 60)
            print("自检通过！所有测试均成功。")
            print("=" * 60)
            return True

        except AssertionError as e:
            print(f"\n✗ 自检失败: {e}")
            print(f"错误码: E008")
            return False
        except Exception as e:
            print(f"\n✗ 自检异常: {e}")
            print(f"错误码: E009")
            return False


# =============================================================================
# 命令行入口
# =============================================================================
def main():
    """主入口函数"""
    # 检查 Python 版本
    if sys.version_info < (3, 8):
        print(f"错误: 需要 Python 3.8+，当前版本 {sys.version}")
        print("错误码: E010")
        return 1

    parser = argparse.ArgumentParser(
        description="办公自动化 Python 脚本生成器",
        epilog="示例: python main.py --task '批量重命名文件' --input ./files --output ./scripts"
    )
    parser.add_argument(
        "--task", "-t",
        type=str,
        help="任务描述文本（必填）"
    )
    parser.add_argument(
        "--input", "-i",
        nargs="*",
        default=[],
        help="输入文件列表"
    )
    parser.add_argument(
        "--output", "-o",
        default=".",
        help="输出目录"
    )
    parser.add_argument(
        "--selftest",
        action="store_true",
        help="运行离线自检（不依赖外部文件）"
    )
    parser.add_argument(
        "--save",
        action="store_true",
        help="将生成的脚本保存到文件"
    )

    parser.add_argument("--force", action="store_true")  # R4 强制写盘


    parser.add_argument("--dry-run", action="store_true")  # R4 预览模式

    args = parser.parse_args()

    global dry_run

    dry_run = getattr(args, "dry_run", False)  # v3.274 同步到全局

    # 自检模式
    if args.selftest:
        success = SelfTest.run()
        return 0 if success else 1

    # 检查必填参数
    if not args.task:
        print(f"错误: {ERROR_CODES['E001']}")
        print("请使用 --task 参数提供任务描述，或使用 --selftest 运行自检")
        return 1

    try:
        # 创建生成器
        automation = AutomationGenerator()

        # 处理任务
        print(f"任务描述: {args.task}")
        print(f"输入文件: {args.input if args.input else '无'}")
        print(f"输出目录: {args.output}")

        script = automation.process(
            description=args.task,
            input_files=args.input,
            output_dir=args.output
        )

        # 输出结果
        print(f"\n生成脚本成功！")
        print(f"依赖库: {', '.join(script.dependencies) if script.dependencies else '无'}")
        print(f"\n使用说明:")
        for note in script.usage_notes:
            print(f"  • {note}")

        if script.warnings:
            print(f"\n注意事项:")
            for warning in script.warnings:
                print(f"  ⚠ {warning}")

        # 保存脚本
        if args.save:
            saved_path = automation.save_script(script, args.output)
            print(f"\n脚本已保存到: {saved_path}")
        else:
            print(f"\n脚本内容预览:")
            print("-" * 60)
            print(script.code[:2000] + ("\n..." if len(script.code) > 2000 else ""))
            print("-" * 60)
            print("提示: 使用 --save 参数保存完整脚本")

        return 0

    except FileNotFoundError as e:
        print(f"错误: {e}")
        return 1
    except PermissionError as e:
        print(f"错误: {e}")
        return 1
    except ValueError as e:
        print(f"错误: {e}")
        return 1
    except SyntaxError as e:
        print(f"错误: {e}")
        return 1
    except Exception as e:
        print(f"错误: {ERROR_CODES['E009']} - {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
