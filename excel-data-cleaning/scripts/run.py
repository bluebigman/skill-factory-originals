#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel Data Cleaning Skill - run.py
Main entry point for the excel-data-cleaning skill.
Supports CSV and XLSX files with streaming processing for large files.
"""

import argparse
import csv
import json
import os
import sys
import tempfile
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
dry_run = False  # v3.274 模块级 dry-run 标志

# Try to import openpyxl for XLSX support
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False
    print("Warning: openpyxl not installed. XLSX support disabled.", file=sys.stderr)


def load_spec(config_path: str) -> Dict[str, Any]:
    """Load the cleaning specification from a JSON config file."""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Config file not found: {config_path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in config file {config_path}: {e}")


def match_trigger(filename: str, trigger_patterns: List[str]) -> bool:
    """Check if a filename matches any of the trigger patterns."""
    if not trigger_patterns:
        return True
    name = Path(filename).name.lower()
    for pattern in trigger_patterns:
        if pattern.lower() in name:
            return True
    return False


def parse_date_value(value: str, date_format: str) -> str:
    """Parse a date string and reformat it."""
    if not value:
        return value
    
    # Try multiple common date formats
    formats = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%m/%d/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y%m%d',
        '%b %d, %Y',
        '%B %d, %Y',
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(value.strip(), fmt)
            return dt.strftime(date_format)
        except (ValueError, TypeError):
            continue
    
    # If value is already in target format, return as-is
    try:
        datetime.strptime(value.strip(), date_format)
        return value
    except (ValueError, TypeError):
        pass
    
    return value  # Return original if no format matches


def clean_csv_file_streaming(input_path: str, output_path: str, spec: Dict[str, Any]) -> Tuple[int, int, int]:
    """
    Clean a CSV file according to the spec using streaming processing.
    Returns (processed_count, skipped_count, error_count).
    """
    try:
        delimiter = spec.get('delimiter', ',')
        encoding = spec.get('encoding', 'utf-8')
        remove_duplicates = spec.get('remove_duplicates', False)
        drop_columns = spec.get('drop_columns', [])
        rename_columns = spec.get('rename_columns', {})
        fillna_value = spec.get('fillna_value', '')
        strip_whitespace = spec.get('strip_whitespace', False)
        date_columns = spec.get('date_columns', [])
        date_format = spec.get('date_format', '%Y-%m-%d')

        processed_count = 0
        skipped_count = 0
        seen = set()

        # Create temp file for atomic write
        temp_fd, temp_path = tempfile.mkstemp(dir=os.path.dirname(output_path) or '.', suffix='.tmp')
        
        try:
            with open(input_path, 'r', encoding=encoding, newline='') as infile, \
                 os.fdopen(temp_fd, 'w', encoding=encoding, newline='') as outfile:
                
                reader = csv.DictReader(infile, delimiter=delimiter)
                fieldnames = reader.fieldnames
                
                # Apply column operations to fieldnames
                final_fieldnames = []
                for col in fieldnames:
                    if col in drop_columns:
                        continue
                    new_name = rename_columns.get(col, col)
                    if new_name not in final_fieldnames:
                        final_fieldnames.append(new_name)
                
                writer = csv.DictWriter(outfile, fieldnames=final_fieldnames, delimiter=delimiter)
                writer.writeheader()
                
                for row in reader:
                    # Skip empty rows
                    if not row or all(v is None or v == '' for v in row.values()):
                        skipped_count += 1
                        continue
                    
                    # Strip whitespace if requested
                    if strip_whitespace:
                        row = {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
                    
                    # Remove duplicates if requested
                    if remove_duplicates:
                        row_key = tuple(sorted(row.items()))
                        if row_key in seen:
                            skipped_count += 1
                            continue
                        seen.add(row_key)
                    
                    # Drop columns
                    for col in drop_columns:
                        if col in row:
                            del row[col]
                    
                    # Rename columns
                    for old_name, new_name in rename_columns.items():
                        if old_name in row:
                            row[new_name] = row.pop(old_name)
                    
                    # Fill NA values
                    for col in fieldnames:
                        if col in row and (row[col] is None or row[col] == ''):
                            row[col] = fillna_value
                    
                    # Format date columns
                    for col in date_columns:
                        if col in row and row[col]:
                            row[col] = parse_date_value(row[col], date_format)
                    
                    # Write only fields that are in final_fieldnames
                    filtered_row = {k: v for k, v in row.items() if k in final_fieldnames}
                    writer.writerow(filtered_row)
                    processed_count += 1
            
            os.replace(temp_path, output_path)
            return processed_count, skipped_count, 0
            
        except Exception as e:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            raise ValueError(f"Failed to process CSV: {e}")
            
    except Exception as e:
        return 0, 0, 1


def clean_xlsx_file(input_path: str, output_path: str, spec: Dict[str, Any]) -> Tuple[int, int, int]:
    """
    Clean an XLSX file according to the spec.
    Returns (processed_count, skipped_count, error_count).
    """
    if not XLSX_SUPPORT:
        return 0, 0, 1
    
    try:
        remove_duplicates = spec.get('remove_duplicates', False)
        drop_columns = spec.get('drop_columns', [])
        rename_columns = spec.get('rename_columns', {})
        fillna_value = spec.get('fillna_value', '')
        strip_whitespace = spec.get('strip_whitespace', False)
        date_columns = spec.get('date_columns', [])
        date_format = spec.get('date_format', '%Y-%m-%d')

        # Load workbook
        wb = load_workbook(input_path, read_only=True)
        ws = wb.active
        
        # Get headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Apply column operations to headers
        final_headers = []
        for col in headers:
            if col in drop_columns:
                continue
            new_name = rename_columns.get(col, col)
            if new_name not in final_headers:
                final_headers.append(new_name)
        
        # Create output workbook
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append(final_headers)
        
        processed_count = 0
        skipped_count = 0
        seen = set()
        
        # Process rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            
            # Skip empty rows
            if not row_dict or all(v is None or v == '' for v in row_dict.values()):
                skipped_count += 1
                continue
            
            # Strip whitespace if requested
            if strip_whitespace:
                row_dict = {k: (v.strip() if isinstance(v, str) else v) for k, v in row_dict.items()}
            
            # Remove duplicates if requested
            if remove_duplicates:
                row_key = tuple(sorted(row_dict.items()))
                if row_key in seen:
                    skipped_count += 1
                    continue
                seen.add(row_key)
            
            # Drop columns
            for col in drop_columns:
                if col in row_dict:
                    del row_dict[col]
            
            # Rename columns
            for old_name, new_name in rename_columns.items():
                if old_name in row_dict:
                    row_dict[new_name] = row_dict.pop(old_name)
            
            # Fill NA values
            for col in headers:
                if col in row_dict and (row_dict[col] is None or row_dict[col] == ''):
                    row_dict[col] = fillna_value
            
            # Format date columns
            for col in date_columns:
                if col in row_dict and row_dict[col]:
                    if isinstance(row_dict[col], datetime):
                        row_dict[col] = row_dict[col].strftime(date_format)
                    elif isinstance(row_dict[col], str):
                        row_dict[col] = parse_date_value(row_dict[col], date_format)
            
            # Write only fields that are in final_headers
            filtered_row = [row_dict.get(col, '') for col in final_headers]
            out_ws.append(filtered_row)
            processed_count += 1
        
        wb.close()
        
        # Save output
        out_wb.save(output_path)
        out_wb.close()
        
        return processed_count, skipped_count, 0
        
    except Exception as e:
        return 0, 0, 1


def process_file(input_path: str, output_path: str, spec: Dict[str, Any]) -> Dict[str, Any]:
    """Process a single file according to the spec."""
    try:
        # Check if file exists
        if not os.path.exists(input_path):
            return {'file': input_path, 'status': 'failed', 'error': 'File not found'}

        # Check file extension
        ext = Path(input_path).suffix.lower()
        if ext not in ['.csv', '.txt', '.xlsx', '.xls']:
            return {'file': input_path, 'status': 'failed', 'error': f'Unsupported file type: {ext}'}

        # Process based on file type
        if ext in ['.csv', '.txt']:
            success, skipped, failed = clean_csv_file_streaming(input_path, output_path, spec)
        elif ext in ['.xlsx', '.xls']:
            if not XLSX_SUPPORT:
                return {'file': input_path, 'status': 'failed', 'error': 'XLSX support requires openpyxl package'}
            success, skipped, failed = clean_xlsx_file(input_path, output_path, spec)
        else:
            return {'file': input_path, 'status': 'failed', 'error': 'Unsupported file type'}
        
        if failed > 0:
            return {'file': input_path, 'status': 'failed', 'error': 'Processing failed'}
        return {'file': input_path, 'status': 'success', 'processed': success, 'skipped': skipped}

    except Exception as e:
        return {'file': input_path, 'status': 'failed', 'error': str(e)}


def run_selftest() -> int:
    """Run self-test to verify core functionality."""
    print("Running self-test...")
    
    test_dir = tempfile.mkdtemp()
    
    try:
        # Test 1: CSV cleaning with all features
        test_input = os.path.join(test_dir, 'test_input.csv')
        test_output = os.path.join(test_dir, 'test_output.csv')
        
        with open(test_input, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['name', 'age', 'email', 'join_date'])
            writer.writerow([' Alice ', '30', 'alice@example.com', '2023-01-15'])
            writer.writerow(['Bob', '25', 'bob@example.com', '2023/02/20'])
            writer.writerow(['Alice', '30', 'alice@example.com', '2023-01-15'])  # duplicate
            writer.writerow(['', '40', '', '2023-03-10'])  # empty name and email
        
        spec = {
            'delimiter': ',',
            'encoding': 'utf-8',
            'remove_duplicates': True,
            'strip_whitespace': True,
            'fillna_value': 'N/A',
            'date_columns': ['join_date'],
            'date_format': '%d/%m/%Y',
            'drop_columns': ['age'],
            'rename_columns': {'email': 'contact'}
        }
        
        success, skipped, failed = clean_csv_file_streaming(test_input, test_output, spec)
        assert success == 3, f"Expected 3 rows processed, got {success}"
        assert skipped == 1, f"Expected 1 row skipped, got {skipped}"
        assert failed == 0, f"Expected 0 failures, got {failed}"
        
        # Verify output
        with open(test_output, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            assert len(rows) == 3, f"Expected 3 rows in output, got {len(rows)}"
            assert 'age' not in rows[0], "age column should be dropped"
            assert 'contact' in rows[0], "email should be renamed to contact"
            assert rows[0]['name'] == 'Alice', "Whitespace should be stripped"
            assert rows[0]['join_date'] == '15/01/2023', f"Date should be reformatted, got {rows[0]['join_date']}"
            assert rows[2]['name'] == 'N/A', "Empty name should be filled with N/A"
        
        # Test 2: XLSX support (if available)
        if XLSX_SUPPORT:
            test_xlsx_input = os.path.join(test_dir, 'test_input.xlsx')
            test_xlsx_output = os.path.join(test_dir, 'test_output.xlsx')
            
            wb = Workbook()
            ws = wb.active
            ws.append(['name', 'age', 'join_date'])
            ws.append(['Alice', 30, datetime(2023, 1, 15)])
            ws.append(['Bob', 25, datetime(2023, 2, 20)])
            ws.append(['Alice', 30, datetime(2023, 1, 15)])  # duplicate
            wb.save(test_xlsx_input)
            wb.close()
            
            xlsx_spec = {
                'remove_duplicates': True,
                'date_columns': ['join_date'],
                'date_format': '%Y-%m-%d'
            }
            
            success, skipped, failed = clean_xlsx_file(test_xlsx_input, test_xlsx_output, xlsx_spec)
            assert success == 2, f"Expected 2 rows processed, got {success}"
            assert skipped == 1, f"Expected 1 row skipped, got {skipped}"
            assert failed == 0, f"Expected 0 failures, got {failed}"
            
            # Verify XLSX output
            wb_out = load_workbook(test_xlsx_output)
            ws_out = wb_out.active
            rows_out = list(ws_out.iter_rows(values_only=True))
            assert len(rows_out) == 3, f"Expected 3 rows (header + 2 data), got {len(rows_out)}"
            assert rows_out[1][0] == 'Alice', "First data row should be Alice"
            assert rows_out[1][2] == '2023-01-15', f"Date should be formatted, got {rows_out[1][2]}"
            wb_out.close()
        
        # Test 3: match_trigger
        assert match_trigger('test.csv', ['test']) == True
        assert match_trigger('other.csv', ['test']) == False
        assert match_trigger('any.csv', []) == True
        
        # Test 4: load_spec with invalid file
        try:
            load_spec('/nonexistent/path/config.json')
            assert False, "Should have raised FileNotFoundError"
        except FileNotFoundError:
            pass
        
        # Test 5: process_file with nonexistent file
        result = process_file('/nonexistent/file.csv', '/tmp/out.csv', spec)
        assert result['status'] == 'failed', "Should have failed for nonexistent file"
        
        # Test 6: process_file with unsupported extension
        result = process_file('test.pdf', 'test_out.pdf', spec)
        assert result['status'] == 'failed', "Should have failed for unsupported extension"
        
        # Test 7: process_file with CSV
        result = process_file(test_input, os.path.join(test_dir, 'process_out.csv'), spec)
        assert result['status'] == 'success', f"Should have succeeded, got {result}"
        assert result['processed'] == 3, f"Expected 3 processed, got {result['processed']}"
        
        print("Self-test passed!")
        return 0
        
    except AssertionError as e:
        print(f"Self-test failed: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Self-test error: {e}", file=sys.stderr)
        return 1
    finally:
        import shutil
        shutil.rmtree(test_dir, ignore_errors=True)


def main() -> int:
    parser = argparse.ArgumentParser(description='Excel Data Cleaning Skill')
    parser.add_argument('--input', '-i', help='Input file or directory')
    parser
