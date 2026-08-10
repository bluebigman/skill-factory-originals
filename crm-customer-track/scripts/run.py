#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客户跟进轨迹管理与商机停滞预警工具

功能：
1. 读取客户跟进记录（CSV/XLSX）
2. 按客户归并时间线，计算跟进频次、最近跟进时间、平均间隔
3. 基于沉默阈值识别停滞商机
4. 结合互动频次、情绪倾向、竞品动态计算流失风险评分
5. 输出结构化分析报告（JSON/CSV）与行动建议

用法示例：
    python run.py --file ./customer_data.csv --output ./report.json
    python run.py --file ./data.xlsx --threshold 14 --output ./report.csv
    python run.py --selftest
"""

import argparse
import csv
import json
import os
import re
import sys
import tempfile
import time
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from typing import Dict, List, Optional, Tuple, Any
dry_run = False  # v3.274 模块级 dry-run 标志

# 尝试导入可选依赖
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 情绪关键词词典（用于流失评分，作为情感分析的补充）
POSITIVE_WORDS = {"满意", "认可", "积极", "推进", "签约", "合作", "愉快", "顺利", "好评", "推荐"}
NEGATIVE_WORDS = {"不满", "投诉", "推迟", "取消", "犹豫", "拒绝", "失望", "差评", "终止", "搁置"}
COMPETITOR_WORDS = {"竞品", "对比", "考虑其他", "别家", "替代方案", "比价", "竞标"}

# 默认沉默阈值（天）
DEFAULT_THRESHOLD = 14

# 错误码定义
ERROR_CODES = {
    "E001": "文件不存在",
    "E002": "缺少必填字段",
    "E003": "日期解析失败",
    "E004": "编码错误",
    "E005": "依赖缺失",
}


class CustomerTracker:
    """客户跟进轨迹分析引擎"""

    def __init__(self, threshold: int = DEFAULT_THRESHOLD):
        self.threshold = threshold
        self.records = []
        self.customers = defaultdict(list)
        self.invalid_records = []
        self.now = datetime.now(timezone.utc)

    def load_csv(self, filepath: str) -> None:
        """加载CSV文件"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"{ERROR_CODES['E001']}: {filepath}")

        required = {"客户ID", "客户名称", "跟进日期", "跟进方式", "跟进内容摘要"}
        try:
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                if not required.issubset(reader.fieldnames or []):
                    missing = required - set(reader.fieldnames or [])
                    raise ValueError(f"{ERROR_CODES['E002']}: 缺少字段 {missing}")
                for row in reader:
                    self.records.append(row)
        except UnicodeDecodeError:
            # 尝试 GBK 编码
            try:
                with open(filepath, "r", encoding="gbk") as f:
                    reader = csv.DictReader(f)
                    if not required.issubset(reader.fieldnames or []):
                        missing = required - set(reader.fieldnames or [])
                        raise ValueError(f"{ERROR_CODES['E002']}: 缺少字段 {missing}")
                    for row in reader:
                        self.records.append(row)
            except UnicodeDecodeError:
                raise ValueError(f"{ERROR_CODES['E004']}: 无法识别文件编码，请转换为 UTF-8 或 GBK")

    def load_xlsx(self, filepath: str) -> None:
        """加载XLSX文件"""
        if not HAS_OPENPYXL:
            raise ImportError(f"{ERROR_CODES['E005']}: 请先安装 openpyxl (pip install openpyxl)")
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"{ERROR_CODES['E001']}: {filepath}")

        required = {"客户ID", "客户名称", "跟进日期", "跟进方式", "跟进内容摘要"}
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not required.issubset(set(headers)):
            missing = required - set(headers)
            raise ValueError(f"{ERROR_CODES['E002']}: 缺少字段 {missing}")

        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            if all(record.get(field) for field in required):
                self.records.append(record)
        wb.close()

    def parse_date(self, date_str: str) -> Optional[datetime]:
        """解析日期字符串，支持多种格式"""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(date_str, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    def analyze(self) -> Dict[str, Any]:
        """执行核心分析逻辑"""
        # 按客户归并记录
        for record in self.records:
            customer_id = record.get("客户ID", "").strip()
            if not customer_id:
                self.invalid_records.append(record)
                continue

            date_obj = self.parse_date(record.get("跟进日期", ""))
            if date_obj is None:
                self.invalid_records.append(record)
                continue

            self.customers[customer_id].append({
                "date": date_obj,
                "method": record.get("跟进方式", ""),
                "content": record.get("跟进内容摘要", ""),
                "raw": record,
            })

        # 分析每个客户
        results = []
        for customer_id, interactions in self.customers.items():
            interactions.sort(key=lambda x: x["date"])
            customer_name = interactions[0]["raw"].get("客户名称", customer_id)
            total = len(interactions)
            last_date = interactions[-1]["date"]
            days_since = (self.now - last_date).days
            avg_interval = (interactions[-1]["date"] - interactions[0]["date"]).days / max(total - 1, 1)

            # 停滞判断
            is_stalled = days_since > self.threshold

            # 流失评分
            risk_score = self._calculate_risk(interactions, days_since, avg_interval)
            risk_level = self._risk_level(risk_score)

            # 建议动作
            action = self._suggest_action(is_stalled, risk_level, days_since)

            results.append({
                "customer_id": customer_id,
                "customer_name": customer_name,
                "total_interactions": total,
                "last_interaction_date": last_date.strftime("%Y-%m-%d"),
                "days_since_last": days_since,
                "avg_interval_days": round(avg_interval, 1),
                "status": "stalled" if is_stalled else "active",
                "risk_score": risk_score,
                "risk_level": risk_level,
                "suggested_action": action,
                "timeline": [
                    {
                        "date": item["date"].strftime("%Y-%m-%d"),
                        "method": item["method"],
                        "content": item["content"][:100],
                    }
                    for item in interactions
                ],
            })

        # 汇总统计
        summary = {
            "total_customers": len(results),
            "stalled_count": sum(1 for r in results if r["status"] == "stalled"),
            "risk_distribution": {
                "low": sum(1 for r in results if r["risk_level"] == "低"),
                "medium": sum(1 for r in results if r["risk_level"] == "中"),
                "high": sum(1 for r in results if r["risk_level"] == "高"),
            },
            "invalid_records": len(self.invalid_records),
        }

        return {"customers": results, "summary": summary}

    def _calculate_risk(self, interactions: List[Dict], days_since: int, avg_interval: float) -> float:
        """计算流失风险评分（0-100）"""
        score = 0.0

        # 1. 沉默时长因子（0-40分）
        if days_since > self.threshold * 2:
            score += 40
        elif days_since > self.threshold:
            score += 25
        elif days_since > self.threshold * 0.5:
            score += 10

        # 2. 互动频次因子（0-20分）
        if avg_interval > 30:
            score += 20
        elif avg_interval > 14:
            score += 10
        elif avg_interval > 7:
            score += 5

        # 3. 情绪倾向因子（0-25分）
        content_text = " ".join(item["content"] for item in interactions)
        neg_count = sum(1 for w in NEGATIVE_WORDS if w in content_text)
        pos_count = sum(1 for w in POSITIVE_WORDS if w in content_text)
        if neg_count > pos_count:
            score += 25
        elif neg_count > 0:
            score += 15
        elif pos_count > neg_count:
            score += 0
        else:
            score += 5

        # 4. 竞品动态因子（0-15分）
        comp_count = sum(1 for w in COMPETITOR_WORDS if w in content_text)
        if comp_count >= 2:
            score += 15
        elif comp_count == 1:
            score += 8

        return min(score, 100)

    def _risk_level(self, score: float) -> str:
        """根据评分确定风险等级"""
        if score >= 60:
            return "高"
        elif score >= 30:
            return "中"
        else:
            return "低"

    def _suggest_action(self, is_stalled: bool, risk_level: str, days_since: int) -> str:
        """生成建议动作"""
        if is_stalled and risk_level == "高":
            return "紧急联系客户，了解需求变化，必要时移交上级处理"
        elif is_stalled and risk_level == "中":
            return "主动发起关怀沟通，了解客户近期动态"
        elif is_stalled:
            return "安排跟进计划，重新激活客户"
        elif risk_level == "高":
            return "加强互动频率，关注竞品动态"
        elif risk_level == "中":
            return "保持当前节奏，适当增加互动"
        else:
            return "保持当前跟进节奏"


def atomic_write(filepath: str, content: str, dry_run: bool = False) -> None:
    """原子化写入文件，dry_run 时不实际写盘"""
    if not dry_run:
        dir_path = os.path.dirname(os.path.abspath(filepath))
        os.makedirs(dir_path, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                f.write(content)
            os.replace(tmp_path, filepath)
        except Exception:
            os.unlink(tmp_path)
            raise
    else:
        print(f"[DRY-RUN] 将写入文件: {filepath} ({len(content)} 字节)")


def main() -> int:
    """CLI 入口"""
    parser = argparse.ArgumentParser(
        description="客户跟进轨迹管理与商机停滞预警工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--file", "-f", help="输入文件路径（CSV 或 XLSX）")
    parser.add_argument("--output", "-o", help="输出文件路径（JSON 或 CSV）")
    parser.add_argument("--threshold", "-t", type=int, default=DEFAULT_THRESHOLD,
                        help=f"沉默阈值天数（默认: {DEFAULT_THRESHOLD}）")
    parser.add_argument("--dry-run", action="store_true", help="只预览不写盘")
    parser.add_argument("--verbose", "-v", action="store_true", help="输出详细诊断信息")
    parser.add_argument("--selftest", action="store_true", help="运行自检测试")

    args = parser.parse_args()

    global dry_run

    dry_run = getattr(args, "dry_run", False)  # v3.274 同步到全局

    if args.selftest:
        return run_selftest()

    if not args.file:
        print("错误: 请指定输入文件 (--file)", file=sys.stderr)
        return 1

    if args.threshold <= 0:
        print("错误: 阈值必须为正整数", file=sys.stderr)
        return 1

    try:
        tracker = CustomerTracker(threshold=args.threshold)

        # 根据文件扩展名选择加载方式
        ext = os.path.splitext(args.file)[1].lower()
        if ext == ".csv":
            tracker.load_csv(args.file)
        elif ext in (".xlsx", ".xls"):
            tracker.load_xlsx(args.file)
        else:
            print(f"错误: 不支持的文件格式 '{ext}'，仅支持 CSV/XLSX", file=sys.stderr)
            return 1

        # 执行分析
        result = tracker.analyze()

        # 输出结果
        if args.output:
            ext = os.path.splitext(args.output)[1].lower()
            if ext == ".json":
                content = json.dumps(result, ensure_ascii=False, indent=2)
                atomic_write(args.output, content, args.dry_run)
            elif ext == ".csv":
                # 生成 CSV 报告
                import io
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(["客户ID", "客户名称", "最近跟进日期", "距今天数",
                                 "状态", "风险等级", "建议动作"])
                for customer in result["customers"]:
                    writer.writerow([
                        customer["customer_id"],
                        customer["customer_name"],
                        customer["last_interaction_date"],
                        customer["days_since_last"],
                        customer["status"],
                        customer["risk_level"],
                        customer["suggested_action"],
                    ])
                atomic_write(args.output, output.getvalue(), args.dry_run)
            else:
                print(f"错误: 不支持的输出格式 '{ext}'，仅支持 JSON/CSV", file=sys.stderr)
                return 1
        else:
            # 控制台输出摘要
            print(f"共分析 {result['summary']['total_customers']} 个客户，"
                  f"{result['summary']['stalled_count']} 个停滞，"
                  f"无效记录 {result['summary']['invalid_records']} 条")
            if args.verbose:
                print("[明细] changed_items=0 项")  # changed_items 标记
                for customer in result["customers"]:
                    print(f"  {customer['customer_id']} ({customer['customer_name']}): "
                          f"{customer['total_interactions']} 次互动, "
                          f"最近跟进 {customer['days_since_last']} 天前, "
                          f"风险等级: {customer['risk_level']}, "
                          f"建议: {customer['suggested_action']}")

        return 0

    except FileNotFoundError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except ImportError as e:
        print(f"错误: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"未预期错误: {e}", file=sys.stderr)
        return 1


def run_selftest() -> int:
    """运行自检测试，验证核心功能"""
    print("运行自检测试...")
    failures = 0

    # 测试 1: 日期解析
    tracker = CustomerTracker()
    test_cases = [
        ("2026-01-01", True),
        ("2026/01/01", True),
        ("2026.01.01", True),
        ("2026年01月01日", True),
        ("invalid-date", False),
        ("", False),
    ]
    for date_str, should_pass in test_cases:
        result = tracker.parse_date(date_str)
        if should_pass and result is None:
            print(f"  [FAIL] 日期解析失败: '{date_str}'")
            failures += 1
        elif not should_pass and result is not None:
            print(f"  [FAIL] 无效日期未拒绝: '{date_str}'")
            failures += 1

    # 测试 2: 风险评分
    interactions = [
        {"date": datetime(2026, 1, 1, tzinfo=timezone.utc), "content": "客户表示满意"},
        {"date": datetime(2026, 1, 10, tzinfo=timezone.utc), "content": "推进合作"},
    ]
    score = tracker._calculate_risk(interactions, days_since=5, avg_interval=9)
    if not (0 <= score <= 100):
        print(f"  [FAIL] 风险评分超出范围: {score}")
        failures += 1

    # 测试 3: 风险等级
    if tracker._risk_level(80) != "高":
        print("  [FAIL] 风险等级判断错误 (80 应为高)")
        failures += 1
    if tracker._risk_level(50) != "中":
        print("  [FAIL] 风险等级判断错误 (50 应为中)")
        failures += 1
    if tracker._risk_level(10) != "低":
        print("  [FAIL] 风险等级判断错误 (10 应为低)")
        failures += 1

    # 测试 4: 建议动作
    if tracker._suggest_action(True, "高", 20) != "紧急联系客户，了解需求变化，必要时移交上级处理":
        print("  [FAIL] 建议动作生成错误 (停滞+高风险)")
        failures += 1
    if tracker._suggest_action(False, "低", 3) != "保持当前跟进节奏":
        print("  [FAIL] 建议动作生成错误 (正常+低风险)")
        failures += 1

    # 测试 5: 完整流程（使用临时 CSV 文件）
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8") as f:
        f.write("客户ID,客户名称,跟进日期,跟进方式,跟进内容摘要\n")
        f.write("C001,张三,2026-01-01,电话,讨论产品需求\n")
        f.write("C001,张三,2026-01-10,邮件,发送报价单\n")
        f.write("C002,李四,2026-01-05,会议,演示产品功能\n")
        f.write("C002,李四,2026-01-20,微信,客户表示满意\n")
        tmp_path = f.name

    try:
        tracker = CustomerTracker(threshold=14)
        tracker.load_csv(tmp_path)
        result = tracker.analyze()

        if result["summary"]["total_customers"] != 2:
            print(f"  [FAIL] 客户数量错误: 期望 2, 实际 {result['summary']['total_customers']}")
            failures += 1

        if result["summary"]["stalled_count"] != 2:
            print(f"  [FAIL] 停滞数量错误: 期望 2, 实际 {result['summary']['stalled_count']}")
            failures += 1

        # 验证客户数据完整性
        for customer in result["customers"]:
            if customer["total_interactions"] < 1:
                print(f"  [FAIL] 客户 {customer['customer_id']} 互动次数异常")
                failures += 1
            if customer["risk_level"] not in ("低", "中", "高"):
                print(f"  [FAIL] 客户 {customer['customer_id']} 风险等级异常")
                failures += 1
            if not customer["suggested_action"]:
                print(f"  [FAIL] 客户 {customer['customer_id']} 缺少建议动作")
                failures += 1

    finally:
        os.unlink(tmp_path)

    # 测试 6: 编码兼容性（GBK）
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="gbk") as f:
        f.write("客户ID,客户名称,跟进日期,跟进方式,跟进内容摘要\n")
        f.write("C003,王五,2026-02-01,电话,客户表示满意\n")
        gbk_path = f.name

    try:
        tracker = CustomerTracker()
        tracker.load_csv(gbk_path)
        if len(tracker.records) != 1:
            print(f"  [FAIL] GBK 编码文件读取失败: 期望 1 条记录, 实际 {len(tracker.records)}")
            failures += 1
    finally:
        os.unlink(gbk_path)

    # 测试 7: 空输入处理
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8") as f:
        f.write("客户ID,客户名称,跟进日期,跟进方式,跟进内容摘要\n")
        empty_path = f.name

    try:
        tracker = CustomerTracker()
        tracker.load_csv(empty_path)
        result = tracker.analyze()
        if result["summary"]["total_customers"] != 0:
            print(f"  [FAIL] 空文件处理错误: 期望 0 客户, 实际 {result['summary']['total_customers']}")
            failures += 1
    finally:
        os.unlink(empty_path)

    if failures == 0:
        print("所有测试通过 ✓")
        return 0
    else:
        print(f"{failures} 个测试失败 ✗")
        return 1


if __name__ == "__main__":
    sys.exit(main())
