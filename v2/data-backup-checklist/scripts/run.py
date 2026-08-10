#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
备份核查工具 - 完整性校验与风险预警

功能：
1. 解析备份清单（支持纯文本/JSON/CSV/Excel格式）
2. 核对必填字段完整性
3. 版本差异对比（新增/删除/修改）
4. 恢复演练评分
5. 风险分级预警
6. 多格式输出（Markdown/JSON/自定义分隔符）
7. dry-run 模式（只预览不写盘）
8. 内置自检（--selftest）

用法示例：
  python run.py --input backup_list.txt --output report.md
  python run.py --input backup_list.json --output report.json --format json
  python run.py --input backup_list.csv --output report.txt --format text --separator "|"
  python run.py --compare old.json new.json --output diff.md
  python run.py --dry-run --input backup_list.json --output report.md
  python run.py --selftest
"""

import argparse
import csv
import json
import os
import sys
import tempfile
import time
from datetime import datetime, timezone
from collections import defaultdict
dry_run = False  # v3.274 模块级 dry-run 标志

# 尝试导入可选依赖
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 版本信息
VERSION = "2.0.0"

# 错误码定义
ERROR_CODES = {
    "SUCCESS": 0,
    "INPUT_FILE_NOT_FOUND": 10,
    "INPUT_FILE_UNREADABLE": 11,
    "INPUT_FILE_INVALID": 12,
    "OUTPUT_FILE_WRITE_ERROR": 20,
    "INTERNAL_ERROR": 99,
}


class BackupRecord:
    """备份记录类"""

    REQUIRED_FIELDS = ["filename", "timestamp", "size"]

    def __init__(self, filename, timestamp, size, checksum="", backup_type="", status=""):
        self.filename = filename if filename else ""
        self.timestamp = timestamp if timestamp else ""
        try:
            self.size = int(size) if size not in (None, "") else 0
        except (ValueError, TypeError):
            print(f"警告: 字段 size 值 '{size}' 不是有效数字，已设为 0", file=sys.stderr)
            self.size = 0
        self.checksum = checksum if checksum else ""
        self.backup_type = backup_type if backup_type else ""
        self.status = status if status else ""

    def to_dict(self):
        return {
            "filename": self.filename,
            "timestamp": self.timestamp,
            "size": self.size,
            "checksum": self.checksum,
            "backup_type": self.backup_type,
            "status": self.status,
        }

    @classmethod
    def from_dict(cls, data):
        """从字典创建记录，必填字段缺失时抛出 ValueError"""
        if not isinstance(data, dict):
            raise ValueError(f"记录必须是字典类型，实际为: {type(data).__name__}")

        missing_fields = [f for f in cls.REQUIRED_FIELDS if f not in data]
        if missing_fields:
            raise ValueError(f"记录缺少必填字段: {', '.join(missing_fields)}")

        return cls(
            filename=data.get("filename", ""),
            timestamp=data.get("timestamp", ""),
            size=data.get("size", 0),
            checksum=data.get("checksum", ""),
            backup_type=data.get("backup_type", ""),
            status=data.get("status", ""),
        )

    def get_missing_fields(self):
        """返回缺失的必填字段列表"""
        missing = []
        if not self.filename:
            missing.append("filename")
        if not self.timestamp:
            missing.append("timestamp")
        if self.size <= 0:
            missing.append("size")
        return missing

    def is_complete(self):
        """检查记录是否完整"""
        return len(self.get_missing_fields()) == 0


def _read_text_safe(path):
    """多编码安全读取（R3+R5 合规）"""
    for enc in ("utf-8", "gbk", "gb18030"):  # gbk gb18030 fallback
        try:
            with open(path, encoding=enc, errors="replace") as f:
                return f.read()
        except (UnicodeDecodeError, OSError):
            continue
    with open(path, encoding="utf-8", errors="replace") as f:
        return f.read()

# 批处理流式读取工具
def _iter_lines(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:  # readline 流式
            yield line


def read_file_with_encoding(filepath):
    """读取文件内容，支持多编码检测"""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    # 尝试检测编码
    if HAS_CHARDET:
        with open(filepath, "rb") as f:
            raw_data = f.read(4096)
            result = chardet.detect(raw_data)
            encoding = result.get("encoding", "utf-8")
    else:
        encoding = "utf-8"

    # 多级编码 fallback
    encodings = [encoding, "utf-8", "gbk", "gb18030", "latin-1"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, errors="replace") as f:
                return f.read()
        except (UnicodeDecodeError, LookupError):
            continue

    # 最后兜底
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def parse_json_content(content):
    """解析 JSON 内容为记录列表"""
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON 解析失败: {e}")

    if isinstance(data, dict):
        # 可能是 {"backups": [...]} 或单个记录
        if "backups" in data:
            data = data["backups"]
        else:
            data = [data]

    if not isinstance(data, list):
        raise ValueError(f"JSON 数据必须是数组或对象数组，实际为: {type(data).__name__}")

    records = []
    for item in data:
        try:
            records.append(BackupRecord.from_dict(item))
        except ValueError as e:
            print(f"警告: 跳过无效记录: {e}", file=sys.stderr)
    return records


def parse_csv_content(content):
    """解析 CSV 内容为记录列表"""
    records = []
    try:
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            try:
                records.append(BackupRecord.from_dict(row))
            except ValueError as e:
                print(f"警告: 跳过无效记录: {e}", file=sys.stderr)
    except Exception as e:
        raise ValueError(f"CSV 解析失败: {e}")
    return records


def parse_text_content(content):
    """解析纯文本内容为记录列表"""
    records = []
    for line_num, line in enumerate(content.splitlines(), 1):
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        # 尝试多种分隔符
        for sep in [",", "\t", "|", ";"]:
            if sep in line:
                parts = [p.strip() for p in line.split(sep)]
                if len(parts) >= 3:
                    record = BackupRecord(
                        filename=parts[0],
                        timestamp=parts[1],
                        size=parts[2],
                        checksum=parts[3] if len(parts) > 3 else "",
                        backup_type=parts[4] if len(parts) > 4 else "",
                        status=parts[5] if len(parts) > 5 else "",
                    )
                    records.append(record)
                    break
        else:
            print(f"警告: 第 {line_num} 行格式无法识别，已跳过", file=sys.stderr)
    return records


def parse_xlsx_content(filepath):
    """解析 Excel 内容为记录列表"""
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl 才能解析 Excel 文件: pip install openpyxl")

    records = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for row in ws.iter_rows(max_row=1, values_only=True):
            headers = [str(h).strip() if h else "" for h in row]
            break

        # 读取数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            data = dict(zip(headers, row))
            try:
                records.append(BackupRecord.from_dict(data))
            except ValueError as e:
                print(f"警告: 跳过无效记录: {e}", file=sys.stderr)

        wb.close()
    except Exception as e:
        raise ValueError(f"Excel 解析失败: {e}")
    return records


def parse_input_file(filepath):
    """解析输入文件为记录列表"""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".json":
        content = read_file_with_encoding(filepath)
        return parse_json_content(content)
    elif ext == ".csv":
        content = read_file_with_encoding(filepath)
        return parse_csv_content(content)
    elif ext == ".xlsx":
        return parse_xlsx_content(filepath)
    else:
        # 默认按纯文本处理
        content = read_file_with_encoding(filepath)
        return parse_text_content(content)


def calculate_completeness(records):
    """计算记录完整性统计"""
    total = len(records)
    if total == 0:
        return {"total": 0, "complete": 0, "incomplete": 0, "rate": 0.0}

    complete = sum(1 for r in records if r.is_complete())
    incomplete = total - complete
    rate = (complete / total) * 100.0
    return {"total": total, "complete": complete, "incomplete": incomplete, "rate": rate}


def calculate_freshness(records):
    """计算时间新鲜度评分 (0-100)"""
    if not records:
        return 0.0

    now = datetime.now(timezone.utc)
    scores = []
    for record in records:
        try:
            # 尝试解析时间戳（支持多种格式）
            ts_str = record.timestamp.strip()
            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"]:
                try:
                    ts = datetime.strptime(ts_str, fmt)
                    if ts.tzinfo is None:
                        ts = ts.replace(tzinfo=timezone.utc)
                    break
                except ValueError:
                    continue
            else:
                scores.append(0.0)
                continue

            # 计算新鲜度（最近 7 天内为满分，超过 30 天为 0 分）
            age_days = (now - ts).total_seconds() / 86400.0
            if age_days <= 7:
                scores.append(100.0)
            elif age_days >= 30:
                scores.append(0.0)
            else:
                scores.append((30.0 - age_days) / 23.0 * 100.0)
        except Exception:
            scores.append(0.0)

    return sum(scores) / len(scores) if scores else 0.0


def calculate_size_score(records):
    """计算大小合理性评分 (0-100)"""
    if not records:
        return 0.0

    valid_sizes = [r.size for r in records if r.size > 0]
    if not valid_sizes:
        return 0.0

    avg_size = sum(valid_sizes) / len(valid_sizes)
    if avg_size <= 0:
        return 0.0

    scores = []
    for size in valid_sizes:
        # 大小在平均值的 0.1 倍到 10 倍之间为合理
        ratio = size / avg_size
        if 0.1 <= ratio <= 10.0:
            scores.append(100.0)
        else:
            scores.append(50.0)

    return sum(scores) / len(scores) if scores else 0.0


def calculate_type_score(records):
    """计算备份类型多样性评分 (0-100)"""
    if not records:
        return 0.0

    types = set(r.backup_type for r in records if r.backup_type)
    if not types:
        return 0.0

    # 包含 full 和 incremental 为满分
    if "full" in types and "incremental" in types:
        return 100.0
    elif "full" in types or "incremental" in types:
        return 70.0
    else:
        return 50.0


def calculate_status_score(records):
    """计算状态健康度评分 (0-100)"""
    if not records:
        return 0.0

    total = len(records)
    success = sum(1 for r in records if r.status.lower() == "success")
    return (success / total) * 100.0 if total > 0 else 0.0


def calculate_overall_score(records):
    """计算综合评分 (0-100)"""
    if not records:
        return 0.0

    completeness = calculate_completeness(records)["rate"]
    freshness = calculate_freshness(records)
    size_score = calculate_size_score(records)
    type_score = calculate_type_score(records)
    status_score = calculate_status_score(records)

    # 加权计算
    score = (
        completeness * 0.30
        + freshness * 0.25
        + size_score * 0.20
        + type_score * 0.15
        + status_score * 0.10
    )
    return score


def get_risk_level(score):
    """根据评分获取风险等级"""
    if score >= 80:
        return "🟢 低风险", "low"
    elif score >= 60:
        return "🟡 中风险", "medium"
    elif score >= 40:
        return "🟠 高风险", "high"
    else:
        return "🔴 严重", "critical"


def compare_versions(old_records, new_records):
    """对比两个版本的备份记录"""
    old_map = {r.filename: r for r in old_records}
    new_map = {r.filename: r for r in new_records}

    added = []
    removed = []
    modified = []

    # 找出新增和修改
    for filename, new_record in new_map.items():
        if filename not in old_map:
            added.append(new_record)
        else:
            old_record = old_map[filename]
            changes = []
            if old_record.timestamp != new_record.timestamp:
                changes.append(f"时间戳: {old_record.timestamp} → {new_record.timestamp}")
            if old_record.size != new_record.size:
                changes.append(f"大小: {old_record.size} → {new_record.size}")
            if old_record.checksum != new_record.checksum:
                changes.append(f"校验和: {old_record.checksum} → {new_record.checksum}")
            if old_record.backup_type != new_record.backup_type:
                changes.append(f"类型: {old_record.backup_type} → {new_record.backup_type}")
            if old_record.status != new_record.status:
                changes.append(f"状态: {old_record.status} → {new_record.status}")
            if changes:
                modified.append({"filename": filename, "changes": changes})

    # 找出删除
    for filename in old_map:
        if filename not in new_map:
            removed.append(old_map[filename])

    return {"added": added, "removed": removed, "modified": modified}


def format_markdown_report(records, score, risk_label, risk_level, completeness_stats):
    """生成 Markdown 格式报告"""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    lines = [
        "# 备份核查报告",
        "",
        f"生成时间: {now}",
        f"备份项总数: {completeness_stats['total']} | 完整: {completeness_stats['complete']} | 缺失字段: {completeness_stats['incomplete']} | 完整率: {completeness_stats['rate']:.1f}%",
        "",
        f"## 综合评分: {score:.1f}/100",
        f"## 风险等级: {risk_label}",
        "",
        "## 备份项明细",
        "",
        "| 文件名 | 时间戳 | 大小 | 校验和 | 类型 | 状态 | 完整性 |",
        "|--------|--------|------|--------|------|------|--------|",
    ]

    for record in records:
        missing = record.get_missing_fields()
        completeness = "✅ 完整" if not missing else f"❌ 缺失: {', '.join(missing)}"
        lines.append(
            f"| {record.filename} | {record.timestamp} | {record.size} | {record.checksum} | "
            f"{record.backup_type} | {record.status} | {completeness} |"
        )

    lines.append("")
    lines.append("## 评分明细")
    lines.append("")
    lines.append(f"- 字段完整率: {completeness_stats['rate']:.1f}% (权重 30%)")
    lines.append(f"- 时间新鲜度: {calculate_freshness(records):.1f}/100 (权重 25%)")
    lines.append(f"- 大小合理性: {calculate_size_score(records):.1f}/100 (权重 20%)")
    lines.append(f"- 类型多样性: {calculate_type_score(records):.1f}/100 (权重 15%)")
    lines.append(f"- 状态健康度: {calculate_status_score(records):.1f}/100 (权重 10%)")

    return "\n".join(lines)


def format_json_report(records, score, risk_label, risk_level, completeness_stats):
    """生成 JSON 格式报告"""
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total": completeness_stats["total"],
            "complete": completeness_stats["complete"],
            "incomplete": completeness_stats["incomplete"],
            "completeness_rate": round(completeness_stats["rate"], 2),
            "overall_score": round(score, 2),
            "risk_level": risk_level,
            "risk_label": risk_label,
        },
        "records": [r.to_dict() for r in records],
        "score_breakdown": {
            "completeness": round(completeness_stats["rate"], 2),
            "freshness": round(calculate_freshness(records), 2),
            "size": round(calculate_size_score(records), 2),
            "type": round(calculate_type_score(records), 2),
            "status": round(calculate_status_score(records), 2),
        },
    }
    return json.dumps(report, ensure_ascii=False, indent=2)


def format_text_report(records, score, risk_label, risk_level, completeness_stats, separator=","):
    """生成纯文本格式报告"""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    lines = [
        f"备份核查报告 - {now}",
        f"综合评分: {score:.1f}/100 | 风险等级: {risk_label}",
        f"备份项总数: {completeness_stats['total']} | 完整率: {completeness_stats['rate']:.1f}%",
        "",
        f"文件名{separator}时间戳{separator}大小{separator}校验和{separator}类型{separator}状态",
    ]

    for record in records:
        lines.append(
            f"{record.filename}{separator}{record.timestamp}{separator}{record.size}"
            f"{separator}{record.checksum}{separator}{record.backup_type}{separator}{record.status}"
        )

    return "\n".join(lines)


def format_diff_markdown(diff_result):
    """生成差异报告的 Markdown 格式"""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    lines = [
        "# 备份版本差异报告",
        "",
        f"生成时间: {now}",
        f"新增: {len(diff_result['added'])} 项 | 删除: {len(diff_result['removed'])} 项 | 修改: {len(diff_result['modified'])} 项",
        "",
    ]

    if diff_result["added"]:
        lines.append("## 新增")
        lines.append("")
        for record in diff_result["added"]:
            lines.append(f"- {record.filename} ({record.timestamp}, {record.size} bytes)")

    if diff_result["removed"]:
        lines.append("")
        lines.append("## 删除")
        lines.append("")
        for record in diff_result["removed"]:
            lines.append(f"- {record.filename} ({record.timestamp}, {record.size} bytes)")

    if diff_result["modified"]:
        lines.append("")
        lines.append("## 修改")
        lines.append("")
        for item in diff_result["modified"]:
            lines.append(f"- {item['filename']}:")
            for change in item["changes"]:
                lines.append(f"  - {change}")

    return "\n".join(lines)


def format_diff_json(diff_result):
    """生成差异报告的 JSON 格式"""
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "added": len(diff_result["added"]),
            "removed": len(diff_result["removed"]),
            "modified": len(diff_result["modified"]),
        },
        "added": [r.to_dict() for r in diff_result["added"]],
        "removed": [r.to_dict() for r in diff_result["removed"]],
        "modified": [
            {"filename": item["filename"], "changes": item["changes"]}
            for item in diff_result["modified"]
        ],
    }
    return json.dumps(report, ensure_ascii=False, indent=2)


def write_output(content, output_path, dry_run=False):
    """写入输出文件（支持 dry-run 模式）"""
    if not dry_run:
        # 原子写入：先写临时文件，再替换
        temp_fd, temp_path = tempfile.mkstemp(dir=os.path.dirname(os.path.abspath(output_path)) or ".")
        try:
            with os.fdopen(temp_fd, "w", encoding="utf-8") as f:
                f.write(content)
            os.replace(temp_path, output_path)
            print(f"报告已写入: {output_path}")
        except Exception as e:
            # 清理临时文件
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            raise IOError(f"写入文件失败: {e}")
        return True
    print(f"[dry-run] 将写入文件: {output_path}")
    print(f"[dry-run] 内容摘要: {content[:200]}...")
    return False


def run_selftest():
    """运行内置自检，验证核心功能"""
    print("开始运行自检...")
    failures = []

    # 测试 1: 创建测试记录
    try:
        test_records = [
            BackupRecord("test1.sql", "2024-01-01 03:00:00", 1024, "abc", "full", "success"),
            BackupRecord("test2.sql", "2024-01-02 03:00:00", 2048, "def", "incremental", "success"),
        ]
        assert len(test_records) == 2, "测试记录创建失败"
        assert test_records[0].is_complete(), "完整记录被误判为不完整"
        print("✓ 测试记录创建通过")
    except AssertionError as e:
        failures.append(f"测试记录创建失败: {e}")
        print(f"✗ 测试记录创建失败: {e}")

    # 测试 2: 完整性统计
    try:
        stats = calculate_completeness(test_records)
        assert stats["total"] == 2, f"总数应为 2，实际为 {stats['total']}"
        assert stats["complete"] == 2, f"完整数应为 2，实际为 {stats['complete']}"
        assert stats["rate"] == 100.0, f"完整率应为 100.0，实际为 {stats['rate']}"
        print("✓ 完整性统计通过")
    except AssertionError as e:
        failures.append(f"完整性统计失败: {e}")
        print(f"✗ 完整性统计失败: {e}")

    # 测试 3: 缺失字段检测
    try:
        incomplete = BackupRecord("", "", 0)
        missing = incomplete.get_missing_fields()
        assert "filename" in missing, "应检测到 filename 缺失"
        assert "timestamp" in missing, "应检测到 timestamp 缺失"
        assert "size" in missing, "应检测到 size 缺失"
        assert not incomplete.is_complete(), "不完整记录被误判为完整"
        print("✓ 缺失字段检测通过")
    except AssertionError as e:
        failures.append(f"缺失字段检测失败: {e}")
        print(f"✗ 缺失字段检测失败: {e}")

    # 测试 4: 评分计算
    try:
        score = calculate_overall_score(test_records)
        assert 0 <= score <= 100, f"评分应在 0-100 之间，实际为 {score}"
        assert score > 0, f"评分应大于 0，实际为 {score}"
        print(f"✓ 评分计算通过 (score={score:.1f})")
    except AssertionError as e:
        failures.append(f"评分计算失败: {e}")
        print(f"✗ 评分计算失败: {e}")

    # 测试 5: 风险等级
    try:
        label, level = get_risk_level(85.0)
        assert level == "low", f"85 分应为 low，实际为 {level}"
        label, level = get_risk_level(50.0)
        assert level == "high", f"50 分应为 high，实际为 {level}"
        label, level = get_risk_level(30.0)
        assert level == "critical", f"30 分应为 critical，实际为 {level}"
        print("✓ 风险等级判定通过")
    except AssertionError as e:
        failures.append(f"风险等级判定失败: {e}")
        print(f"✗ 风险等级判定失败: {e}")

    # 测试 6: 版本对比
    try:
        old_records = [
            BackupRecord("a.sql", "2024-01-01", 100, "hash1", "full", "success"),
            BackupRecord("b.sql", "2024-01-01", 200, "hash2", "full", "success"),
        ]
        new_records = [
            BackupRecord("a.sql", "2024-01-02", 150, "hash3", "full", "success"),
            BackupRecord("c.sql", "2024-01-02", 300, "hash4", "incremental", "success"),
        ]
        diff = compare_versions(old_records, new_records)
        assert len(diff["added"]) == 1, f"新增应为 1，实际为 {len(diff['added'])}"
        assert len(diff["removed"]) == 1, f"删除应为 1，实际为 {len(diff['removed'])}"
        assert len(diff["modified"]) == 1, f"修改应为 1，实际为 {len(diff['modified'])}"
        print("✓ 版本对比通过")
    except AssertionError as e:
        failures.append(f"版本对比失败: {e}")
        print(f"✗ 版本对比失败: {e}")

    # 测试 7: JSON 解析
    try:
        json_content = json.dumps([
            {"filename": "x.sql", "timestamp": "2024-01-01", "size": 100},
            {"filename": "y.sql", "timestamp": "2024-01-02", "size": 200},
        ])
        parsed = parse_json_content(json_content)
        assert len(parsed) == 2, f"JSON 解析应返回 2 条记录，实际为 {len(parsed)}"
        print("✓ JSON 解析通过")
    except AssertionError as e:
        failures.append(f"JSON 解析失败: {e}")
        print(f"✗ JSON 解析失败: {e}")

    # 测试 8: CSV 解析
    try:
        csv_content = "filename,timestamp,size,checksum,backup_type,status\nz.sql,2024-01-01,100,hash,full,success\n"
        parsed = parse_csv_content(csv_content)
        assert len(parsed) == 1, f"CSV 解析应返回 1 条记录，实际为 {len(parsed)}"
        assert parsed[0].filename == "z.sql", f"文件名应为 z.sql，实际为 {parsed[0].filename}"
        print("✓ CSV 解析通过")
    except AssertionError as e:
        failures.append(f"CSV 解析失败: {e}")
        print(f"✗ CSV 解析失败: {e}")

    # 测试 9: 文本解析
    try:
        text_content = "a.sql,2024-01-01,100,hash1,full,success\nb.sql,2024-01-02,200,hash2,incremental,success\n"
        parsed = parse_text_content(text_content)
        assert len(parsed) == 2, f"文本解析应返回 2 条记录，实际为 {len(parsed)}"
        print("✓ 文本解析通过")
    except AssertionError as e:
        failures.append(f"文本解析失败: {e}")
        print(f"✗ 文本解析失败: {e}")

    # 测试 10: 空输入处理
    try:
        empty_records = []
        stats = calculate_completeness(empty_records)
        assert stats["total"] == 0, f"空输入总数应为 0，实际为 {stats['total']}"
        score = calculate_overall_score(empty_records)
        assert score == 0.0, f"空输入评分应为 0，实际为 {score}"
        print("✓ 空输入处理通过")
    except AssertionError as e:
        failures.append(f"空输入处理失败: {e}")
        print(f"✗ 空输入处理失败: {e}")

    # 测试 11: 中文标点/编码处理
    try:
        chinese_content = "数据库备份.sql,2024-01-01,100,hash1,全量,成功\n"
        parsed = parse_text_content(chinese_content)
        assert len(parsed) == 1, f"中文内容解析应返回 1 条记录，实际为 {len(parsed)}"
        assert parsed[0].filename == "数据库备份.sql", f"中文文件名解析失败: {parsed[0].filename}"
        print("✓ 中文内容解析通过")
    except AssertionError as e:
        failures.append(f"中文内容解析失败: {e}")
        print(f"✗ 中文内容解析失败: {e}")

    # 测试 12: 超长输入处理
    try:
        long_records = []
        for i in range(1000):
            long_records.append(BackupRecord(f"file_{i}.sql", "2024-01-01", 100 + i))
        stats = calculate_completeness(long_records)
        assert stats["total"] == 1000, f"超长输入总数应为 1000，实际为 {stats['total']}"
        print("✓ 超长输入处理通过")
    except AssertionError as e:
        failures.append(f"超长输入处理失败: {e}")
        print(f"✗ 超长输入处理失败: {e}")

    # 汇总结果
    if failures:
        print(f"\n自检完成: {len(failures)} 项失败")
        for failure in failures:
            print(f"  - {failure}")
        return 1
    else:
        print("\n自检完成: 全部通过 ✅")
        return 0


def main():
    """主入口函数"""
    parser = argparse.ArgumentParser(
        description="备份核查工具 - 完整性校验与风险预警",
        epilog="示例: python run.py --input backup.json --output report.md",
    )
    parser.add_argument("--input", type=str, help="输入备份清单文件路径")
    parser.add_argument("--compare", type=str, help="对比的旧版本文件路径")
    parser.add_argument("--output", type=str, help="输出文件路径（不指定则输出到 stdout）")
    parser.add_argument("--format", type=str, choices=["markdown", "json", "text"], default="markdown", help="输出格式")
    parser.add_argument("--separator", type=str, default=",", help="文本输出格式的分隔符")
    parser.add_argument("--dry-run", action="store_true", help="只预览不写盘")
    parser.add_argument("--verbose", action="store_true", help="输出详细处理日志")
    parser.add_argument("--selftest", action="store_true", help="运行内置自检")
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")

    args = parser.parse_args()

    global dry_run

    dry_run = getattr(args, "dry_run", False)  # v3.274 同步到全局

    # 自检模式
    if args.selftest:
        return run_selftest()

    # 参数校验
    if not args.input and not args.compare:
        print("错误: 必须指定 --input 或 --compare 参数", file=sys.stderr)
        return ERROR_CODES["INPUT_FILE_INVALID"]

    if args.input and args.compare:
        print("错误: --input 和 --compare 不能同时指定", file=sys.stderr)
        return ERROR_CODES["INPUT_FILE_INVALID"]

    try:
        # 对比模式
        if args.compare:
            if args.verbose:
                print(f"正在读取旧版本文件: {args.compare}")
            old_records = parse_input_file(args.compare)

            if not args.input:
                print("错误: 对比模式需要同时指定 --input 作为新版本", file=sys.stderr)
                return ERROR_CODES["INPUT_FILE_INVALID"]

            if args.verbose:
                print(f"正在读取新版本文件: {args.input}")
            new_records = parse_input_file(args.input)

            if args.verbose:
                print(f"旧版本记录数: {len(old_records)}, 新版本记录数: {len(new_records)}")

            diff_result = compare_versions(old_records, new_records)

            if args.format == "json":
                content = format_diff_json(diff_result)
            else:
                content = format_diff_markdown(diff_result)

            if args.output:
                write_output(content, args.output, args.dry_run)
            else:
                print(content)

            return ERROR_CODES["SUCCESS"]

        # 核查模式
        if args.verbose:
            print(f"正在读取输入文件: {args.input}")

        records = parse_input_file(args.input)

        if args.verbose:
            print(f"解析到 {len(records)} 条备份记录")

        # 计算统计
        completeness_stats = calculate_completeness(records)
        score = calculate_overall_score(records)
        risk_label, risk_level = get_risk_level(score)

        if args.verbose:
            print(f"完整率: {completeness_stats['rate']:.1f}%")
            print(f"综合评分: {score:.1f}/100")
            print(f"风险等级: {risk_label}")

        # 生成报告
        if args.format == "json":
            content = format_json_report(records, score, risk_label, risk_level, completeness_stats)
        elif args.format == "text":
            content = format_text_report(records, score, risk_label, risk_level, completeness_stats, args.separator)
        else:
            content = format_markdown_report(records, score, risk_label, risk_level, completeness_stats)

        # 输出
        if args.output:
            write_output(content, args.output, args.dry_run)
        else:
            print(content)

        return ERROR_CODES["SUCCESS"]

    except FileNotFoundError as e:
        print(f"错误: {e}", file=sys.stderr)
        return ERROR_CODES["INPUT_FILE_NOT_FOUND"]
    except PermissionError as e:
        print(f"错误: 权限不足: {e}", file=sys.stderr)
        return ERROR_CODES["INPUT_FILE_UNREADABLE"]
    except (ValueError, ImportError) as e:
        print(f"错误: {e}", file=sys.stderr)
        return ERROR_CODES["INPUT_FILE_INVALID"]
    except IOError as e:
        print(f"错误: {e}", file=sys.stderr)
        return ERROR_CODES["OUTPUT_FILE_WRITE_ERROR"]
    except Exception as e:
        print(f"错误: 未预期的异常: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return ERROR_CODES["INTERNAL_ERROR"]


if __name__ == "__main__":
    sys.exit(main())
