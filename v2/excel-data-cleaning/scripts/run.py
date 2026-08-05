#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 数据清洗工具 - 表格整理、数据规范化、清洗校验

功能：
1. 日期字段格式统一
2. 数字字段格式规范化
3. 全字段去重
4. 空值标记
5. 字段拆分（按分隔符/正则）
6. 单文件/批量处理
7. 生成清洗日志和去重报告

用法示例：
    python run.py --config 清洗规则.json --input 样本文件.xlsx --output 样本_清洗后.xlsx
    python run.py --config 清洗规则.json --input-dir ./data/ --output-dir ./cleaned/
"""

import argparse
import csv
import json
import os
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl 库。请安装: pip install openpyxl")
    sys.exit(1)

# ============ 核心清洗逻辑 ============

def load_config(config_path):
    """加载清洗规则配置文件"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在: {config_path}")
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    # 校验必要字段
    required = ['date_fields', 'numeric_fields', 'deduplicate']
    for field in required:
        if field not in config:
            raise ValueError(f"配置缺少必要字段: {field}")
    return config


def normalize_date(value, date_format):
    """统一日期格式"""
    if value is None or str(value).strip() == '':
        return value
    # 尝试多种常见格式解析
    date_str = str(value).strip()
    # 处理 Excel 序列号
    if isinstance(value, (int, float)) and value > 20000:
        try:
            return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(value) - 2).strftime(date_format)
        except:
            pass
    # 常见日期格式
    formats = [
        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
        '%Y年%m月%d日', '%m/%d/%Y', '%d/%m/%Y',
        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S'
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime(date_format)
        except:
            continue
    # 尝试正则提取
    match = re.search(r'(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})', date_str)
    if match:
        try:
            dt = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            return dt.strftime(date_format)
        except:
            pass
    return f"[无法解析日期]{date_str}"


def normalize_number(value):
    """规范化数字格式"""
    if value is None or str(value).strip() == '':
        return value
    if isinstance(value, (int, float)):
        return value
    num_str = str(value).strip()
    # 去除千分位逗号和货币符号
    num_str = num_str.replace(',', '').replace('¥', '').replace('$', '').replace('%', '')
    try:
        if '%' in str(value):
            return float(num_str) / 100
        return float(num_str)
    except:
        return f"[无法解析数字]{value}"


def split_field(value, separator, new_fields):
    """按分隔符拆分字段"""
    if value is None:
        return [None] * len(new_fields)
    parts = str(value).split(separator)
    result = []
    for i, field in enumerate(new_fields):
        if i < len(parts):
            result.append(parts[i].strip())
        else:
            result.append(None)
    return result


def clean_row(row, headers, config):
    """清洗单行数据，返回 (清洗后行, 状态, 说明)"""
    try:
        cleaned = list(row)
        # 日期字段格式化
        for date_field in config.get('date_fields', []):
            if date_field in headers:
                idx = headers.index(date_field)
                if idx < len(cleaned) and cleaned[idx] is not None:
                    cleaned[idx] = normalize_date(cleaned[idx], config.get('date_format', '%Y-%m-%d'))
        
        # 数字字段规范化
        for num_field in config.get('numeric_fields', []):
            if num_field in headers:
                idx = headers.index(num_field)
                if idx < len(cleaned) and cleaned[idx] is not None:
                    cleaned[idx] = normalize_number(cleaned[idx])
        
        # 空值标记
        empty_marker = config.get('empty_marker', '[缺失]')
        for i in range(len(cleaned)):
            if cleaned[i] is None or str(cleaned[i]).strip() == '':
                cleaned[i] = empty_marker
        
        # 字段拆分
        for rule in config.get('split_rules', []):
            field = rule.get('field')
            separator = rule.get('separator', ' ')
            new_fields = rule.get('new_fields', [])
            if field in headers and new_fields:
                idx = headers.index(field)
                if idx < len(cleaned) and cleaned[idx] is not None:
                    split_values = split_field(cleaned[idx], separator, new_fields)
                    # 在原始字段后插入新字段
                    for j, val in enumerate(split_values):
                        cleaned.insert(idx + 1 + j, val)
        
        return cleaned, '成功', ''
    except Exception as e:
        return row, '失败', str(e)


def process_file(input_path, output_path, config, log_path=None, report_path=None):
    """处理单个 Excel 文件"""
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    
    # 读取 Excel
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        raise ValueError(f"无法读取 Excel 文件 {input_path}: {e}")
    
    ws = wb.active
    # 获取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())
        else:
            headers.append(f"列{len(headers)+1}")
    
    # 处理数据行
    cleaned_rows = []
    log_rows = []
    removed_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            row_data.append(ws.cell(row=row_idx, column=col_idx).value)
        
        # 跳过全空行
        if all(v is None or str(v).strip() == '' for v in row_data):
            continue
        
        cleaned, status, note = clean_row(row_data, headers, config)
        
        if status == '成功':
            # 去重检查
            if config.get('deduplicate', False):
                if cleaned in cleaned_rows:
                    removed_count += 1
                    log_rows.append([os.path.basename(input_path), row_idx, '去重移除', '重复行'])
                    continue
            cleaned_rows.append(cleaned)
            log_rows.append([os.path.basename(input_path), row_idx, '成功', ''])
        else:
            log_rows.append([os.path.basename(input_path), row_idx, '失败', note])
    
    # 写入输出文件
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    
    # 写入表头（考虑拆分后的新字段）
    final_headers = list(headers)
    for rule in config.get('split_rules', []):
        if rule.get('field') in headers and rule.get('new_fields'):
            idx = headers.index(rule['field'])
            for j, new_field in enumerate(rule['new_fields']):
                final_headers.insert(idx + 1 + j, new_field)
    
    out_ws.append(final_headers)
    for row in cleaned_rows:
        out_ws.append(row)
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    out_wb.save(output_path)
    
    # 写入日志
    if log_path:
        os.makedirs(os.path.dirname(log_path) if os.path.dirname(log_path) else '.', exist_ok=True)
        with open(log_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['文件名', '行号', '状态', '说明'])
            writer.writerows(log_rows)
    
    # 写入去重报告
    if report_path and removed_count > 0:
        os.makedirs(os.path.dirname(report_path) if os.path.dirname(report_path) else '.', exist_ok=True)
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"去重报告\n")
            f.write(f"文件: {os.path.basename(input_path)}\n")
            f.write(f"移除重复行数: {removed_count}\n")
            f.write(f"原因: 全字段匹配重复\n")
    
    return len(cleaned_rows), removed_count, len(log_rows) - removed_count


# ============ 主程序 ============

def main():
    parser = argparse.ArgumentParser(description='Excel 数据清洗工具')
    parser.add_argument('--config', required=True, help='清洗规则 JSON 配置文件路径')
    parser.add_argument('--input', help='输入 Excel 文件路径（单文件模式）')
    parser.add_argument('--output', help='输出 Excel 文件路径（单文件模式）')
    parser.add_argument('--input-dir', help='输入目录（批量模式）')
    parser.add_argument('--output-dir', help='输出目录（批量模式）')
    parser.add_argument('--selftest', action='store_true', help='运行自检')
    
    args = parser.parse_args()
    
    if args.selftest:
        selftest()
        return
    
    # 校验参数
    if not args.config:
        print("错误: 必须指定 --config 参数")
        sys.exit(1)
    
    if args.input and args.input_dir:
        print("错误: --input 和 --input-dir 不能同时使用")
        sys.exit(1)
    
    if not args.input and not args.input_dir:
        print("错误: 必须指定 --input 或 --input-dir")
        sys.exit(1)
    
    if args.input and not args.output:
        print("错误: 单文件模式必须指定 --output")
        sys.exit(1)
    
    if args.input_dir and not args.output_dir:
        print("错误: 批量模式必须指定 --output-dir")
        sys.exit(1)
    
    try:
        config = load_config(args.config)
    except Exception as e:
        print(f"错误: 加载配置失败: {e}")
        sys.exit(1)
    
    try:
        if args.input:
            # 单文件模式
            process_file(args.input, args.output, config,
                        log_path=os.path.splitext(args.output)[0] + '_清洗日志.csv',
                        report_path=os.path.splitext(args.output)[0] + '_去重报告.txt')
            print(f"清洗完成: {args.output}")
        else:
            # 批量模式
            input_dir = Path(args.input_dir)
            output_dir = Path(args.output_dir)
            if not input_dir.exists():
                print(f"错误: 输入目录不存在: {input_dir}")
                sys.exit(1)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            excel_files = list(input_dir.glob('*.xlsx')) + list(input_dir.glob('*.xls'))
            if not excel_files:
                print(f"错误: 输入目录中没有 Excel 文件: {input_dir}")
                sys.exit(1)
            
            for file in excel_files:
                output_file = output_dir / (file.stem + '_清洗后.xlsx')
                log_file = output_dir / (file.stem + '_清洗日志.csv')
                report_file = output_dir / (file.stem + '_去重报告.txt')
                try:
                    success, removed, failed = process_file(str(file), str(output_file), config,
                                                           log_path=str(log_file),
                                                           report_path=str(report_file))
                    print(f"处理 {file.name}: 成功 {success} 行, 去重移除 {removed} 行, 失败 {failed} 行")
                except Exception as e:
                    print(f"处理 {file.name} 失败: {e}")
            
            print(f"批量清洗完成，输出目录: {output_dir}")
    
    except Exception as e:
        print(f"错误: {e}")
        traceback.print_exc()
        sys.exit(1)


def selftest():
    """自检函数 - 不联网，纯本地测试"""
    print("运行自检...")
    
    # 创建临时测试文件
    import tempfile
    import shutil
    
    temp_dir = tempfile.mkdtemp()
    try:
        # 创建测试配置
        config = {
            "date_fields": ["日期"],
            "date_format": "%Y-%m-%d",
            "numeric_fields": ["金额"],
            "deduplicate": True,
            "empty_marker": "[缺失]",
            "split_rules": [
                {"field": "姓名", "separator": " ", "new_fields": ["姓", "名"]}
            ]
        }
        config_path = os.path.join(temp_dir, 'test_config.json')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False)
        
        # 创建测试 Excel
        test_input = os.path.join(temp_dir, 'test_input.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['姓名', '日期', '金额'])
        ws.append(['张三 三', '2023/01/15', '1,234.56'])
        ws.append(['李四 四', '2023-02-20', '2,345.67'])
        ws.append(['张三 三', '2023/01/15', '1,234.56'])  # 重复行
        ws.append(['王五', '', 'abc'])  # 空值和非法数字
        wb.save(test_input)
        
        # 执行清洗
        test_output = os.path.join(temp_dir, 'test_output.xlsx')
        process_file(test_input, test_output, config,
                    log_path=os.path.join(temp_dir, 'test_log.csv'),
                    report_path=os.path.join(temp_dir, 'test_report.txt'))
        
        # 验证结果
        wb2 = openpyxl.load_workbook(test_output)
        ws2 = wb2.active
        headers = [cell.value for cell in ws2[1]]
        rows = list(ws2.iter_rows(min_row=2, values_only=True))
        
        assert '姓' in headers and '名' in headers, "字段拆分失败"
        assert len(rows) == 3, f"去重失败，期望3行，实际{len(rows)}行"
        assert rows[0][1] == '2023-01-15', f"日期格式化失败: {rows[0][1]}"
        assert rows[0][2] == 1234.56, f"数字格式化失败: {rows[0][2]}"
        assert rows[2][1] == '[缺失]', f"空值标记失败: {rows[2][1]}"
        
        # 验证日志
        with open(os.path.join(temp_dir, 'test_log.csv'), 'r', encoding='utf-8-sig') as f:
            log_lines = f.readlines()
        assert len(log_lines) >= 5, "日志行数不足"
        
        print("✅ 自检通过！所有功能正常。")
        print("   - 字段拆分: 正常")
        print("   - 日期格式化: 正常")
        print("   - 数字规范化: 正常")
        print("   - 去重: 正常")
        print("   - 空值标记: 正常")
        print("   - 日志生成: 正常")
        
    except AssertionError as e:
        print(f"❌ 自检失败: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 自检异常: {e}")
        traceback.print_exc()
        sys.exit(1)
    finally:
        shutil.rmtree(temp_dir)


if __name__ == "__main__":
    main()
