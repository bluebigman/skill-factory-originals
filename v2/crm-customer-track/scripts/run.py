#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客户旅程 商机预警 跟进复盘 - 客户跟进轨迹管理工具

功能：
1. 读取客户跟进记录（CSV/XLSX）
2. 按客户归并时间线，计算跟进频次、最近跟进时间
3. 基于沉默阈值识别停滞商机
4. 结合互动频次、情绪倾向、竞品动态计算流失风险评分
5. 输出结构化分析报告（JSON/CSV）与行动建议

用法示例：
    python run.py --file ./customer_data.csv --output ./report.json
    python run.py --file ./data.xlsx --threshold 14 --output ./report.csv
    python run.py --selftest
"""

import argparse
import csv
import json
import os
import sys
import tempfile
from datetime import datetime, timedelta
from collections import defaultdict

# 尝试导入可选依赖
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 情绪关键词词典（用于流失评分）
POSITIVE_WORDS = {"满意", "认可", "积极", "推进", "签约", "合作", "愉快", "顺利", "好评", "推荐"}
NEGATIVE_WORDS = {"不满", "投诉", "推迟", "取消", "犹豫", "拒绝", "失望", "差评", "终止", "搁置"}
COMPETITOR_WORDS = {"竞品", "对比", "考虑其他", "别家", "替代方案", "比价", "竞标"}

# 默认沉默阈值（天）
DEFAULT_THRESHOLD = 14

class CustomerTracker:
    """客户跟进轨迹分析引擎"""
    
    def __init__(self, threshold=DEFAULT_THRESHOLD):
        self.threshold = threshold
        self.records = []
        self.customers = defaultdict(list)
    
    def load_csv(self, filepath):
        """加载CSV文件"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")
        
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                required = {"客户ID", "客户名称", "跟进日期", "跟进方式", "跟进内容摘要"}
                if not required.issubset(reader.fieldnames or []):
                    missing = required - set(reader.fieldnames or [])
                    raise ValueError(f"CSV缺少必填字段: {missing}")
                self.records = list(reader)
        except UnicodeDecodeError:
            raise ValueError("CSV文件编码错误，请使用UTF-8编码")
        
        self._process_records()
    
    def load_xlsx(self, filepath):
        """加载XLSX文件"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")
        
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            required = {"客户ID", "客户名称", "跟进日期", "跟进方式", "跟进内容摘要"}
            if not required.issubset(set(headers)):
                missing = required - set(headers)
                raise ValueError(f"XLSX缺少必填字段: {missing}")
            
            self.records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = dict(zip(headers, row))
                if all(record.get(k) for k in required):
                    self.records.append(record)
            wb.close()
        except Exception as e:
            raise ValueError(f"读取XLSX失败: {e}")
        
        self._process_records()
    
    def _process_records(self):
        """处理原始记录，按客户分组"""
        self.customers = defaultdict(list)
        for rec in self.records:
            try:
                date_str = str(rec["跟进日期"]).strip()
                # 支持多种日期格式
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
                    try:
                        rec["_date"] = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    raise ValueError(f"无法解析日期: {date_str}")
                
                rec["_content"] = str(rec.get("跟进内容摘要", "")).lower()
                rec["_method"] = str(rec.get("跟进方式", ""))
                self.customers[rec["客户ID"]].append(rec)
            except (KeyError, ValueError) as e:
                print(f"警告: 跳过无效记录 {rec}: {e}")
        
        # 按日期排序
        for cid in self.customers:
            self.customers[cid].sort(key=lambda x: x["_date"])
    
    def analyze(self):
        """执行完整分析"""
        today = datetime.now()
        results = []
        
        for cid, records in self.customers.items():
            if not records:
                continue
            
            name = records[0]["客户名称"]
            last_date = records[-1]["_date"]
            days_since = (today - last_date).days
            total_count = len(records)
            
            # 计算互动频次（平均间隔天数）
            if total_count > 1:
                intervals = [(records[i]["_date"] - records[i-1]["_date"]).days 
                           for i in range(1, len(records))]
                avg_interval = sum(intervals) / len(intervals)
            else:
                avg_interval = days_since if days_since > 0 else 0
            
            # 停滞判断
            is_stagnant = days_since > self.threshold
            
            # 流失评分（0-100）
            risk_score = self._calc_risk_score(records, days_since, avg_interval)
            
            # 风险等级
            if risk_score >= 70:
                risk_level = "高"
            elif risk_score >= 40:
                risk_level = "中"
            else:
                risk_level = "低"
            
            # 行动建议
            actions = self._generate_actions(is_stagnant, risk_score, days_since, records)
            
            results.append({
                "客户ID": cid,
                "客户名称": name,
                "跟进次数": total_count,
                "最近跟进日期": last_date.strftime("%Y-%m-%d"),
                "距今天数": days_since,
                "平均间隔天数": round(avg_interval, 1),
                "停滞状态": "是" if is_stagnant else "否",
                "流失风险评分": risk_score,
                "风险等级": risk_level,
                "行动建议": actions
            })
        
        # 按风险评分降序排列
        results.sort(key=lambda x: x["流失风险评分"], reverse=True)
        return results
    
    def _calc_risk_score(self, records, days_since, avg_interval):
        """计算流失风险评分"""
        score = 0
        
        # 1. 时间因素（0-40分）
        if days_since > 30:
            score += 40
        elif days_since > 14:
            score += 30
        elif days_since > 7:
            score += 20
        elif days_since > 3:
            score += 10
        
        # 2. 频次因素（0-20分）
        if avg_interval > 30:
            score += 20
        elif avg_interval > 14:
            score += 15
        elif avg_interval > 7:
            score += 10
        elif avg_interval > 3:
            score += 5
        
        # 3. 内容情绪分析（0-40分）
        content = " ".join(rec["_content"] for rec in records[-3:])  # 最近3条
        neg_count = sum(1 for w in NEGATIVE_WORDS if w in content)
        pos_count = sum(1 for w in POSITIVE_WORDS if w in content)
        comp_count = sum(1 for w in COMPETITOR_WORDS if w in content)
        
        score += min(neg_count * 10, 20)  # 负面词最多20分
        score += min(comp_count * 10, 10)  # 竞品词最多10分
        if pos_count > neg_count:
            score -= 10  # 正面情绪减分
        
        return max(0, min(100, score))
    
    def _generate_actions(self, is_stagnant, risk_score, days_since, records):
        """生成行动建议"""
        actions = []
        
        if is_stagnant:
            if days_since > 30:
                actions.append("紧急联系客户，了解当前需求状态")
                actions.append("考虑升级处理或移交上级")
            elif days_since > 14:
                actions.append("安排主动关怀，询问项目进展")
                actions.append("发送最新产品资料或行业动态")
        
        if risk_score >= 70:
            actions.append("立即安排高层拜访")
            actions.append("准备竞品对比方案")
        elif risk_score >= 40:
            actions.append("增加跟进频率至每周1次")
            actions.append("提供定制化解决方案")
        
        if not actions:
            actions.append("保持正常跟进节奏")
        
        # 根据最近记录内容补充建议
        if records:
            last_content = records[-1]["_content"]
            if "预算" in last_content:
                actions.append("确认预算审批流程")
            if "时间" in last_content or "排期" in last_content:
                actions.append("确认时间安排")
        
        return actions
    
    def export_json(self, results, output_path):
        """导出JSON报告"""
        report = {
            "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "沉默阈值": f"{self.threshold}天",
            "客户总数": len(results),
            "停滞商机数": sum(1 for r in results if r["停滞状态"] == "是"),
            "高风险客户数": sum(1 for r in results if r["风险等级"] == "高"),
            "客户分析": results
        }
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
    
    def export_csv(self, results, output_path):
        """导出CSV报告"""
        if not results:
            raise ValueError("没有可导出的数据")
        
        fieldnames = ["客户ID", "客户名称", "跟进次数", "最近跟进日期", "距今天数",
                     "平均间隔天数", "停滞状态", "流失风险评分", "风险等级", "行动建议"]
        
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in results:
                row = dict(r)
                row["行动建议"] = "; ".join(r["行动建议"])
                writer.writerow(row)

def selftest():
    """自检函数：验证核心功能"""
    print("开始自检...")
    
    # 创建测试数据
    test_data = [
        {"客户ID": "C001", "客户名称": "测试客户A", "跟进日期": "2024-01-01", 
         "跟进方式": "电话", "跟进内容摘要": "客户表示满意，推进顺利"},
        {"客户ID": "C001", "客户名称": "测试客户A", "跟进日期": "2024-01-15", 
         "跟进方式": "邮件", "跟进内容摘要": "发送方案，客户认可"},
        {"客户ID": "C002", "客户名称": "测试客户B", "跟进日期": "2024-01-01", 
         "跟进方式": "会议", "跟进内容摘要": "客户投诉价格过高，考虑竞品"},
        {"客户ID": "C002", "客户名称": "测试客户B", "跟进日期": "2024-01-02", 
         "跟进方式": "微信", "跟进内容摘要": "客户表示不满，推迟决策"},
    ]
    
    # 写入临时CSV
    tmp_dir = tempfile.mkdtemp()
    csv_path = os.path.join(tmp_dir, "test_data.csv")
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=list(test_data[0].keys()))
        writer.writeheader()
        writer.writerows(test_data)
    
    # 执行分析
    tracker = CustomerTracker(threshold=7)
    tracker.load_csv(csv_path)
    results = tracker.analyze()
    
    # 验证结果
    assert len(results) == 2, "应分析2个客户"
    assert results[0]["客户ID"] == "C002", "高风险客户应排前面"
    assert results[0]["风险等级"] == "高", "C002应为高风险"
    assert results[1]["客户ID"] == "C001", "C001应排第二"
    assert results[1]["停滞状态"] == "是", "C001应标记为停滞"
    
    # 测试导出
    json_path = os.path.join(tmp_dir, "test_report.json")
    tracker.export_json(results, json_path)
    assert os.path.exists(json_path), "JSON导出失败"
    
    csv_out = os.path.join(tmp_dir, "test_report.csv")
    tracker.export_csv(results, csv_out)
    assert os.path.exists(csv_out), "CSV导出失败"
    
    # 清理
    os.remove(csv_path)
    os.remove(json_path)
    os.remove(csv_out)
    os.rmdir(tmp_dir)
    
    print("自检通过！所有功能正常。")
    return True

def main():
    parser = argparse.ArgumentParser(
        description="客户旅程 商机预警 跟进复盘 - 客户跟进轨迹管理工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python run.py --file ./customer_data.csv --output ./report.json
  python run.py --file ./data.xlsx --threshold 14 --output ./report.csv
  python run.py --selftest
        """
    )
    
    parser.add_argument("--version", action="version", version="crm-customer-track 1.0.0")
    parser.add_argument("--selftest", action="store_true", help="运行自检")
    parser.add_argument("--file", "-f", help="输入文件路径（CSV或XLSX）")
    parser.add_argument("--output", "-o", help="输出文件路径（JSON或CSV，根据扩展名自动判断）")
    parser.add_argument("--threshold", "-t", type=int, default=DEFAULT_THRESHOLD,
                       help=f"沉默阈值天数（默认: {DEFAULT_THRESHOLD}）")
    
    args = parser.parse_args()
    
    # 自检模式
    if args.selftest:
        try:
            selftest()
            sys.exit(0)
        except Exception as e:
            print(f"自检失败: {e}")
            sys.exit(1)
    
    # 参数校验
    if not args.file:
        parser.error("必须指定 --file 参数")
    if not args.output:
        parser.error("必须指定 --output 参数")
    if args.threshold <= 0:
        parser.error("--threshold 必须为正整数")
    
    # 执行分析
    try:
        tracker = CustomerTracker(threshold=args.threshold)
        
        # 根据扩展名选择加载方式
        ext = os.path.splitext(args.file)[1].lower()
        if ext == '.csv':
            tracker.load_csv(args.file)
        elif ext in ('.xlsx', '.xlsm'):
            tracker.load_xlsx(args.file)
        else:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持CSV或XLSX")
        
        results = tracker.analyze()
        
        # 输出
        out_ext = os.path.splitext(args.output)[1].lower()
        if out_ext == '.json':
            tracker.export_json(results, args.output)
        elif out_ext == '.csv':
            tracker.export_csv(results, args.output)
        else:
            raise ValueError(f"不支持的输出格式: {out_ext}，仅支持JSON或CSV")
        
        # 打印摘要
        stagnant = sum(1 for r in results if r["停滞状态"] == "是")
        high_risk = sum(1 for r in results if r["风险等级"] == "高")
        print(f"分析完成！共 {len(results)} 个客户")
        print(f"停滞商机: {stagnant} 个")
        print(f"高风险客户: {high_risk} 个")
        print(f"报告已保存至: {args.output}")
        
        sys.exit(0)
        
    except FileNotFoundError as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
    except ImportError as e:
        print(f"依赖缺失: {e}", file=sys.stderr)
        print("请安装所需依赖: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"数据错误: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"未知错误: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
