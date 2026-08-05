#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
竞品拆解与差异化策略生成工具 (competitor-analysis-ai)

功能：
1. 读取竞品数据（CSV/Excel/纯文本）
2. 多维度拆解：功能、定价、用户体验、市场定位、技术架构、运营策略
3. 生成结构化对比报告（Markdown格式）
4. 输出差异化策略建议与风险提示

用法示例：
    python run.py --input competitors.csv --output report.md
    python run.py --input competitors.xlsx --output report.md --mode detailed
    python run.py --input data.txt --output report.md --top 5
    python run.py --selftest
"""

import argparse
import csv
import json
import os
import sys
import re
from datetime import datetime
from collections import OrderedDict

# 尝试导入可选依赖
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ========== 核心业务逻辑 ==========

# 分析维度定义
ANALYSIS_DIMENSIONS = [
    "功能特性",
    "定价策略",
    "用户体验",
    "市场定位",
    "技术架构",
    "运营策略"
]

# 内置行业知识库（用于策略建议）
INDUSTRY_INSIGHTS = {
    "SaaS": {
        "趋势": "订阅制向按量付费演进，AI功能成为差异化关键",
        "风险": "客户获取成本持续上升，需关注留存率",
        "建议": "聚焦垂直场景，提供可量化的ROI证明"
    },
    "电商": {
        "趋势": "社交电商与直播带货融合，供应链效率决定利润",
        "风险": "价格战激烈，平台规则变化频繁",
        "建议": "强化私域运营，建立品牌护城河"
    },
    "金融科技": {
        "趋势": "合规成本上升，开放银行成为主流",
        "风险": "监管政策不确定性，数据安全要求提高",
        "建议": "与持牌机构合作，注重风控能力建设"
    },
    "教育": {
        "趋势": "AI个性化学习，线上线下融合OMO模式",
        "风险": "获客成本高，政策监管趋严",
        "建议": "深耕内容质量，打造口碑传播"
    },
    "医疗健康": {
        "趋势": "远程医疗普及，AI辅助诊断加速",
        "风险": "数据隐私敏感，审批流程漫长",
        "建议": "与医疗机构深度绑定，积累临床数据"
    },
    "默认": {
        "趋势": "数字化转型加速，用户体验成为核心竞争点",
        "风险": "同质化竞争严重，需持续创新",
        "建议": "聚焦细分市场，快速迭代验证"
    }
}


class CompetitorAnalyzer:
    """竞品分析核心引擎"""
    
    def __init__(self, input_file, output_file, mode="standard", top=10):
        self.input_file = input_file
        self.output_file = output_file
        self.mode = mode
        self.top = min(top, 10)  # 最多分析10个
        self.data = []
        self.errors = []
        
    def load_data(self):
        """加载竞品数据（支持CSV/Excel/纯文本）"""
        if not os.path.exists(self.input_file):
            raise FileNotFoundError(f"输入文件不存在: {self.input_file}")
        
        ext = os.path.splitext(self.input_file)[1].lower()
        
        if ext == ".csv":
            self._load_csv()
        elif ext in (".xlsx", ".xls"):
            if not HAS_OPENPYXL:
                raise ImportError("处理Excel文件需要安装openpyxl库: pip install openpyxl")
            self._load_excel()
        elif ext in (".txt", ".md"):
            self._load_text()
        else:
            raise ValueError(f"不支持的文件格式: {ext}（支持CSV/Excel/纯文本）")
        
        if not self.data:
            raise ValueError("未解析到任何竞品数据，请检查文件内容格式")
        
        # 限制数量
        self.data = self.data[:self.top]
        
    def _load_csv(self):
        """解析CSV文件"""
        try:
            with open(self.input_file, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # 清理空字段
                    clean_row = {k: (v.strip() if v else "") for k, v in row.items()}
                    if clean_row.get("产品名称") or clean_row.get("name"):
                        self.data.append(clean_row)
        except Exception as e:
            raise ValueError(f"CSV解析失败: {e}")
    
    def _load_excel(self):
        """解析Excel文件"""
        try:
            wb = openpyxl.load_workbook(self.input_file, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                item = {headers[i]: (str(row[i]).strip() if row[i] else "") for i in range(len(headers))}
                if item.get("产品名称") or item.get("name"):
                    self.data.append(item)
        except Exception as e:
            raise ValueError(f"Excel解析失败: {e}")
    
    def _load_text(self):
        """解析纯文本文件（按段落分割）"""
        try:
            with open(self.input_file, "r", encoding="utf-8") as f:
                content = f.read()
            
            # 按空行分割段落
            paragraphs = [p.strip() for p in re.split(r'\n\s*\n', content) if p.strip()]
            
            for para in paragraphs:
                item = {"产品名称": para.split("\n")[0][:50]}
                # 尝试提取关键信息
                for line in para.split("\n"):
                    if ":" in line:
                        key, value = line.split(":", 1)
                        item[key.strip()] = value.strip()
                self.data.append(item)
        except Exception as e:
            raise ValueError(f"文本解析失败: {e}")
    
    def analyze(self):
        """执行多维度分析"""
        results = []
        
        for idx, competitor in enumerate(self.data, 1):
            name = competitor.get("产品名称") or competitor.get("name") or f"竞品{idx}"
            
            # 提取各维度信息（缺失字段标注占位符）
            analysis = {
                "序号": idx,
                "产品名称": name,
                "功能特性": self._extract_field(competitor, ["功能", "features", "核心功能"]),
                "定价策略": self._extract_field(competitor, ["定价", "价格", "price"]),
                "用户体验": self._extract_field(competitor, ["体验", "UX", "用户评价"]),
                "市场定位": self._extract_field(competitor, ["定位", "市场", "target"]),
                "技术架构": self._extract_field(competitor, ["技术", "架构", "tech"]),
                "运营策略": self._extract_field(competitor, ["运营", "策略", "operation"]),
            }
            
            # 计算综合评分（基于字段完整度）
            filled = sum(1 for v in analysis.values() if v and not v.startswith("[需核实"))
            analysis["数据完整度"] = f"{filled}/{len(ANALYSIS_DIMENSIONS)}"
            
            results.append(analysis)
        
        return results
    
    def _extract_field(self, data, keys):
        """从数据中提取字段值"""
        for key in keys:
            if key in data and data[key]:
                return data[key]
        # 尝试模糊匹配
        for k, v in data.items():
            if any(word in k for word in keys):
                return v
        return f"[需核实:{keys[0]}]"
    
    def generate_strategy(self, results):
        """生成差异化策略建议"""
        strategies = []
        
        # 识别行业（通过关键词匹配）
        industry = "默认"
        all_text = json.dumps(results, ensure_ascii=False)
        for key in INDUSTRY_INSIGHTS:
            if key != "默认" and key.lower() in all_text.lower():
                industry = key
                break
        
        insights = INDUSTRY_INSIGHTS[industry]
        
        # 分析竞品数量
        count = len(results)
        if count == 0:
            return []
        
        # 生成基础策略
        strategies.append({
            "行业洞察": insights["趋势"],
            "风险提示": insights["风险"],
            "核心建议": insights["建议"]
        })
        
        # 基于数据完整度生成建议
        complete_count = sum(1 for r in results if r["数据完整度"].startswith("6/6"))
        if complete_count < count * 0.5:
            strategies.append({
                "数据建议": "竞品数据完整度不足50%，建议补充功能、定价等关键字段后再做深度分析"
            })
        
        # 定价策略分析
        prices = []
        for r in results:
            price_str = r["定价策略"]
            if price_str and not price_str.startswith("[需核实"):
                numbers = re.findall(r'\d+', price_str)
                if numbers:
                    prices.append(int(numbers[0]))
        
        if len(prices) >= 2:
            avg_price = sum(prices) / len(prices)
            strategies.append({
                "定价洞察": f"竞品平均定价约{avg_price:.0f}元，若你的产品定价在此区间，建议突出差异化价值"
            })
        
        return strategies
    
    def generate_report(self, results, strategies):
        """生成Markdown格式报告"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        lines = [
            "# 竞品分析报告",
            "",
            f"**生成时间**: {now}",
            f"**分析竞品数**: {len(results)}",
            f"**分析模式**: {self.mode}",
            "",
            "---",
            "",
            "## 一、竞品概览",
            "",
            "| 序号 | 产品名称 | 数据完整度 | 功能特性 | 定价策略 |",
            "|------|----------|------------|----------|----------|",
        ]
        
        for r in results:
            lines.append(
                f"| {r['序号']} | {r['产品名称']} | {r['数据完整度']} | "
                f"{r['功能特性'][:30] if r['功能特性'] else 'N/A'} | "
                f"{r['定价策略'][:20] if r['定价策略'] else 'N/A'} |"
            )
        
        lines.extend(["", "## 二、详细对比", ""])
        
        for dim in ANALYSIS_DIMENSIONS:
            lines.append(f"### {dim}")
            lines.append("")
            lines.append("| 产品 | 详情 |")
            lines.append("|------|------|")
            for r in results:
                value = r.get(dim, "N/A")
                lines.append(f"| {r['产品名称']} | {value} |")
            lines.append("")
        
        lines.extend(["## 三、差异化策略建议", ""])
        
        for i, strategy in enumerate(strategies, 1):
            for key, value in strategy.items():
                lines.append(f"### 建议{i}.{key}")
                lines.append("")
                lines.append(f"> {value}")
                lines.append("")
        
        lines.extend(["---", "", "*本报告由AI自动生成，数据来源于用户提供的素材，仅供参考。*"])
        
        return "\n".join(lines)
    
    def run(self):
        """执行完整分析流程"""
        # 1. 加载数据
        self.load_data()
        
        # 2. 执行分析
        results = self.analyze()
        
        # 3. 生成策略
        strategies = self.generate_strategy(results)
        
        # 4. 生成报告
        report = self.generate_report(results, strategies)
        
        # 5. 输出报告
        with open(self.output_file, "w", encoding="utf-8") as f:
            f.write(report)
        
        return len(results)


def selftest():
    """自检函数：验证核心功能正常"""
    print("=== 竞品分析工具自检 ===")
    
    # 创建测试数据
    test_data = [
        {"产品名称": "竞品A", "功能特性": "AI客服、数据分析", "定价策略": "99元/月", "用户体验": "界面简洁", "市场定位": "中小企业", "技术架构": "SaaS", "运营策略": "内容营销"},
        {"产品名称": "竞品B", "功能特性": "CRM、自动化", "定价策略": "199元/月", "用户体验": "功能强大", "市场定位": "大型企业", "技术架构": "私有化部署", "运营策略": "渠道合作"},
    ]
    
    # 测试分析功能
    analyzer = CompetitorAnalyzer.__new__(CompetitorAnalyzer)
    analyzer.data = test_data
    analyzer.mode = "standard"
    
    results = analyzer.analyze()
    assert len(results) == 2, "分析结果数量错误"
    assert results[0]["产品名称"] == "竞品A", "产品名称解析错误"
    assert results[0]["功能特性"] == "AI客服、数据分析", "功能特性解析错误"
    
    # 测试策略生成
    strategies = analyzer.generate_strategy(results)
    assert len(strategies) > 0, "策略生成失败"
    
    # 测试报告生成
    report = analyzer.generate_report(results, strategies)
    assert "竞品分析报告" in report, "报告生成失败"
    assert "差异化策略建议" in report, "策略建议缺失"
    
    # 测试文件IO
    test_input = "selftest_competitors.csv"
    test_output = "selftest_report.md"
    
    try:
        with open(test_input, "w", encoding="utf-8") as f:
            f.write("产品名称,功能特性,定价策略,用户体验,市场定位,技术架构,运营策略\n")
            f.write("测试产品1,AI功能,50元,良好,初创,云原生,社交媒体\n")
            f.write("测试产品2,数据分析,80元,优秀,中型,混合架构,线下活动\n")
        
        analyzer2 = CompetitorAnalyzer(test_input, test_output)
        count = analyzer2.run()
        assert count == 2, "文件分析数量错误"
        assert os.path.exists(test_output), "输出文件未生成"
        
        print("✅ 所有自检测试通过！")
        return 0
    finally:
        # 清理测试文件
        for f in [test_input, test_output]:
            if os.path.exists(f):
                os.remove(f)


def main():
    parser = argparse.ArgumentParser(
        description="竞品拆解与差异化策略生成工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python run.py --input competitors.csv --output report.md
  python run.py --input data.xlsx --output report.md --mode detailed
  python run.py --input data.txt --output report.md --top 5
  python run.py --selftest
        """
    )
    
    parser.add_argument("--input", "-i", help="输入文件路径（CSV/Excel/纯文本）")
    parser.add_argument("--output", "-o", default="competitor_report.md", help="输出报告路径（默认: competitor_report.md）")
    parser.add_argument("--mode", "-m", choices=["standard", "detailed"], default="standard", help="分析模式（默认: standard）")
    parser.add_argument("--top", "-t", type=int, default=10, help="最多分析竞品数量，1-10（默认: 10）")
    parser.add_argument("--selftest", action="store_true", help="运行自检")
    
    args = parser.parse_args()
    
    # 自检模式
    if args.selftest:
        sys.exit(selftest())
    
    # 参数校验
    if not args.input:
        parser.error("必须指定 --input 参数")
    
    if not 1 <= args.top <= 10:
        parser.error("--top 参数必须在1-10之间")
    
    # 执行分析
    try:
        analyzer = CompetitorAnalyzer(args.input, args.output, args.mode, args.top)
        count = analyzer.run()
        print(f"✅ 分析完成！共分析 {count} 个竞品")
        print(f"📄 报告已生成: {args.output}")
        return 0
    except FileNotFoundError as e:
        print(f"❌ 错误: {e}", file=sys.stderr)
        return 1
    except ImportError as e:
        print(f"❌ 依赖缺失: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"❌ 数据错误: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"❌ 未知错误: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
